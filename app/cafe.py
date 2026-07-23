import json
import calendar
import math
import os
import secrets
from datetime import date, datetime, timedelta, time
from io import BytesIO
from uuid import uuid4
from zoneinfo import ZoneInfo

import qrcode
from PIL import Image, ImageDraw, ImageFont
from flask import Blueprint, Response, current_app, flash, g, jsonify, redirect, render_template, request, session, url_for
from openpyxl import Workbook
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from .attendance_logic import (
    ATTENDANCE_STATUS_OPTIONS,
    SELF_LEAVE_TYPE_OPTIONS,
    attendance_datetime_for,
    attendance_flags_for_row,
    attendance_pay_fraction,
    attendance_status_label,
    build_attendance_summary,
    refresh_attendance_row,
    worked_hours_for_row,
)
from .auth_helpers import login_required, roles_required, user_has_any_role, user_has_permission
from .deploy_config import load_deployment_config, save_deployment_config
from .extensions import db, socketio
from .leave_logic import (
    apply_leave_decision,
    business_today,
    calculate_leave_duration,
    ensure_leave_defaults,
    leave_dashboard_context,
    leave_policy,
    run_leave_maintenance,
    validate_leave_request,
    weekly_off_config,
)
from .models import (
    CafeFeedback,
    CafeFeedbackItem,
    CafeOrder,
    CafeOrderItem,
    CafeTable,
    InventoryCategory,
    InventoryExpenseLog,
    InventoryItem,
    InventoryDailyClosing,
    InventoryPurchase,
    InventoryPurchaseLine,
    InventoryRecipe,
    InventoryRecipeItem,
    InventoryToPurchase,
    InventoryVendor,
    InventoryWastage,
    MenuCategory,
    MenuItem,
    MenuType,
    StaffAttendance,
    StaffDocument,
    UserType,
    StaffLeaveRequest,
    StaffProfile,
    CompanyHoliday,
    LeaveBalance,
    LeaveTransaction,
    RoleLeaveRule,
    StaffNotification,
    AttendanceRuleBook,
    TableBooking,
    User,
    Workstation,
)
from .sms_gateway import send_sms_from_config
from .staff_lifecycle import retire_staff_account
from .rulebook import ensure_rulebook_default, next_rulebook_version

bp = Blueprint("cafe", __name__, url_prefix="/cafe")

DEFAULT_ROLE_OPTIONS = (
    "owner",
    "admin",
    "manager",
    "accountant",
    "cashier",
    "staff",
    "server",
    "chef",
    "barista",
    "librarian",
    "inventory_manager",
    "cleaner",
    "delivery_partner",
)
STAFF_ROLES = DEFAULT_ROLE_OPTIONS
DEFAULT_WORKSTATIONS = (
    ("kitchen", "Kitchen"),
    ("barista", "Barista Counter"),
)
WORKSTATION_COLOR_PALETTE = (
    "#6f4a35",
    "#2d7070",
    "#d4a574",
    "#8f6c96",
    "#4d8b53",
    "#9b5b5b",
    "#567fb5",
    "#a38656",
)
IST_TZ = ZoneInfo("Asia/Kolkata")
UTC_TZ = ZoneInfo("UTC")
PROTECTED_MENU_CATEGORY_NAMES = {"other", "utility"}
ITEM_PREP_STATUSES = ("pending", "preparing", "ready", "served")
DEFAULT_ATTENDANCE_CAFE_LAT = 25.207989477704068
DEFAULT_ATTENDANCE_CAFE_LNG = 80.87374457551877
DEFAULT_ATTENDANCE_RADIUS_METERS = 120.0
DEFAULT_RECEIPT_LOCATION = "Chitrakoot, Uttar Pradesh"
STAFF_CALL_COOLDOWN_SECONDS = 5 * 60


def _slugify_workstation(value: str) -> str:
    value = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value or "").strip())
    slug = "-".join(part for part in value.split("-") if part)
    return slug[:40]


def _ensure_workstations_seeded():
    changed = False
    for index, (slug, name) in enumerate(DEFAULT_WORKSTATIONS, start=1):
        station = Workstation.query.filter_by(slug=slug).first()
        if not station:
            db.session.add(Workstation(slug=slug, name=name, active=True, display_order=index))
            changed = True
            continue
        if not station.name:
            station.name = name
            changed = True
        if station.display_order != index:
            station.display_order = index
            changed = True
        if not station.active:
            station.active = True
            changed = True
    if changed:
        db.session.commit()


def _all_workstations(include_inactive: bool = False):
    _ensure_workstations_seeded()
    query = Workstation.query
    if not include_inactive:
        query = query.filter_by(active=True)
    return query.order_by(Workstation.display_order.asc(), Workstation.name.asc()).all()


def _chef_options(include_inactive: bool = False) -> list[User]:
    """Return active staff who can be accountable for menu preparation.

    A menu item's preparation responsibility is not limited to the kitchen:
    Baristas own beverage and counter-prepared items as well. Keep the helper
    name for compatibility with the existing review/statistics payloads while
    applying the shared Chef-or-Barista rule in one place.
    """
    users = User.query.order_by(User.full_name.asc(), User.email.asc()).all()
    result: list[User] = []
    for user in users:
        if not include_inactive and not user.active:
            continue
        if user.has_role("chef") or user.has_role("barista"):
            result.append(user)
    return result


def _is_preparation_responsibility_user(user: User | None) -> bool:
    return bool(user and (user.has_role("chef") or user.has_role("barista")))


def _workstation_slug_set(include_inactive: bool = False) -> set[str]:
    return {station.slug for station in _all_workstations(include_inactive=include_inactive) if station.slug}


def _normalize_prep_station(value: str | None) -> str:
    station = (value or "").strip().lower()
    return station if station in _workstation_slug_set(include_inactive=True) else ""


def _workstation_display_name(slug: str | None) -> str:
    normalized = (slug or "").strip().lower()
    if not normalized:
        return "No Workstation"
    for station in _all_workstations(include_inactive=True):
        if station.slug == normalized:
            return station.name
    return normalized.replace("-", " ").title()


def _station_action_label(slug: str | None, display_name: str | None = None) -> str:
    probe = f"{slug or ''} {display_name or ''}".lower()
    if "barista" in probe or "brew" in probe or "coffee" in probe:
        return "Brewing"
    return "Preparing"


def _inventory_area_options(include_all: bool = False) -> list[dict]:
    options = []
    if include_all:
        options.append({"value": "all", "label": "All"})
    for station in _all_workstations(include_inactive=True):
        options.append({"value": station.slug, "label": station.name})
    options.append({"value": "cafe", "label": "Cafe"})
    return options


def _inventory_area_name_map() -> dict[str, str]:
    return {row["value"]: row["label"] for row in _inventory_area_options(include_all=False)}


def _normalize_inventory_area(value: str | None, default: str | None = None) -> str:
    raw = (value or default or "").strip().lower()
    allowed = {row["value"] for row in _inventory_area_options(include_all=False)}
    if raw in allowed:
        return raw
    first_station = next((row["value"] for row in _inventory_area_options(include_all=False) if row["value"] != "cafe"), "cafe")
    return first_station or "cafe"


def _stats_station_registry(include_unassigned: bool = True) -> list[dict]:
    stations = []
    for index, station in enumerate(_all_workstations(include_inactive=True)):
        stations.append(
            {
                "slug": station.slug,
                "label": station.name,
                "color": WORKSTATION_COLOR_PALETTE[index % len(WORKSTATION_COLOR_PALETTE)],
            }
        )
    if include_unassigned:
        stations.append(
            {
                "slug": "unassigned",
                "label": "No Workstation",
                "color": "#7b7b7b",
            }
        )
    return stations


def _station_slug_for_stats(value: str | None) -> str:
    normalized = _normalize_prep_station(value)
    return normalized or "unassigned"


def _stats_sales_source_options() -> list[dict]:
    return [{"value": "all", "label": "All"}] + [
        {"value": row["slug"], "label": row["label"]}
        for row in _stats_station_registry(include_unassigned=True)
    ]


def _kiosk_token():
    return (load_deployment_config(current_app.instance_path).get("KDS_KIOSK_TOKEN") or "").strip()


def _reception_kiosk_token():
    return (load_deployment_config(current_app.instance_path).get("RECEPTION_KIOSK_TOKEN") or "").strip()


def _attendance_settings():
    cfg = load_deployment_config(current_app.instance_path)
    try:
        cafe_lat = float(cfg.get("ATTENDANCE_CAFE_LAT") or DEFAULT_ATTENDANCE_CAFE_LAT)
    except (TypeError, ValueError):
        cafe_lat = DEFAULT_ATTENDANCE_CAFE_LAT
    try:
        cafe_lng = float(cfg.get("ATTENDANCE_CAFE_LNG") or DEFAULT_ATTENDANCE_CAFE_LNG)
    except (TypeError, ValueError):
        cafe_lng = DEFAULT_ATTENDANCE_CAFE_LNG
    try:
        radius_m = float(cfg.get("ATTENDANCE_RADIUS_METERS") or DEFAULT_ATTENDANCE_RADIUS_METERS)
    except (TypeError, ValueError):
        radius_m = DEFAULT_ATTENDANCE_RADIUS_METERS
    return {
        "cafe_lat": cafe_lat,
        "cafe_lng": cafe_lng,
        "radius_m": max(20.0, radius_m),
    }


def _tax_settings():
    cfg = load_deployment_config(current_app.instance_path)

    def _read_rate(key: str) -> float:
        try:
            value = float(cfg.get(key) or 0)
        except (TypeError, ValueError):
            value = 0.0
        return round(value, 2)

    raw_service_rate = _read_rate("SERVICE_CHARGE_RATE")
    if raw_service_rate <= 0:
        raw_service_rate = _read_rate("SERVICE_TAX_RATE")
    service_charge_rate = min(max(raw_service_rate or 5.0, 5.0), 10.0)

    return {
        "service_charge_rate": service_charge_rate,
    }


def _selected_tax_flags():
    controls_present = request.form.get("tax_controls_present") == "1"

    def _flag(name: str) -> bool:
        if controls_present:
            return request.form.get(name) == "1"
        return True

    return {
        "apply_service_charge": _flag("apply_service_charge"),
    }


def _order_tax_breakdown(order: CafeOrder, *, base_amount: float | None = None, flags: dict | None = None):
    settings = _tax_settings()
    chosen = {
        "apply_service_charge": True,
    }
    if flags:
        chosen.update({key: bool(value) for key, value in flags.items()})
    taxable_base = round(float(order.total_amount if base_amount is None else base_amount) or 0, 2)
    service_tax_amount = round(
        taxable_base * settings["service_charge_rate"] / 100.0 if chosen["apply_service_charge"] else 0.0,
        2,
    )
    gst_amount = 0.0
    cst_amount = 0.0
    return {
        "base_amount": taxable_base,
        "service_charge_rate": settings["service_charge_rate"],
        "apply_service_charge": chosen["apply_service_charge"],
        "service_tax_amount": service_tax_amount,
        "gst_amount": gst_amount,
        "cst_amount": cst_amount,
        "tax_amount": 0.0,
        "grand_total": round(taxable_base + service_tax_amount, 2),
    }


def _apply_order_tax_breakdown(order: CafeOrder, flags: dict | None = None):
    breakdown = _order_tax_breakdown(order, flags=flags)
    order.service_tax_amount = breakdown["service_tax_amount"]
    order.gst_amount = breakdown["gst_amount"]
    order.cst_amount = breakdown["cst_amount"]
    order.total_amount = breakdown["grand_total"]
    return breakdown


def _receipt_location_text():
    return DEFAULT_RECEIPT_LOCATION


def _google_review_link() -> str:
    return "https://g.page/r/CZclLI_Be-puEAI/review"


def _clamp_star_rating(value, default: int = 3) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(1, min(5, parsed))


def _staff_call_cooldown_remaining_seconds(table: CafeTable | None) -> int:
    if not table or not table.last_staff_call_at:
        return 0
    elapsed = (datetime.utcnow() - table.last_staff_call_at).total_seconds()
    remaining = STAFF_CALL_COOLDOWN_SECONDS - int(elapsed)
    return max(0, remaining)


def _payment_breakdown_rows(order: CafeOrder) -> list[dict]:
    try:
        rows = json.loads(order.payment_breakdown_json or "[]")
        if isinstance(rows, list):
            cleaned = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                cleaned.append(
                    {
                        "method": str(row.get("method") or "").strip(),
                        "amount": round(float(row.get("amount") or 0), 2),
                        "reference": str(row.get("reference") or "").strip(),
                    }
                )
            if cleaned:
                return cleaned
    except (TypeError, ValueError, json.JSONDecodeError):
        pass
    if order.payment_type:
        return [
            {
                "method": str(order.payment_type or "").strip(),
                "amount": round(float(order.total_amount or 0), 2),
                "reference": str(order.payment_reference or "").strip(),
            }
        ]
    return []


def _receipt_orders_for_settlement(order: CafeOrder) -> list[CafeOrder]:
    if not order.paid_at:
        return [order]
    query = CafeOrder.query.options(
        joinedload(CafeOrder.table),
        joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
    ).filter(
        CafeOrder.status == "paid",
        CafeOrder.paid_at == order.paid_at,
        CafeOrder.table_id == order.table_id,
    )
    if order.payment_breakdown_json:
        query = query.filter(CafeOrder.payment_breakdown_json == order.payment_breakdown_json)
    else:
        query = query.filter(CafeOrder.id == order.id)
    related = query.order_by(CafeOrder.created_at.asc(), CafeOrder.id.asc()).all()
    return related or [order]


def _feedback_line_items_for_orders(receipt_orders: list[CafeOrder]) -> list[dict]:
    grouped = {}
    for order in receipt_orders:
        for oi in order.order_items:
            if (oi.approval_status or "pending") == "rejected":
                continue
            key = (
                int(oi.menu_item_id or 0),
                (oi.size_label or "").strip(),
                1 if oi.is_parcel else 0,
            )
            entry = grouped.get(key)
            if not entry:
                entry = {
                    "menu_item_id": oi.menu_item_id,
                    "order_item_id": oi.id,
                    "item_name": (oi.menu_item.name if oi.menu_item else "Item"),
                    "size_label": (oi.size_label or "").strip(),
                    "is_parcel": bool(oi.is_parcel),
                    "quantity": 0,
                }
                grouped[key] = entry
            entry["quantity"] += int(oi.quantity or 0)
    rows = list(grouped.values())
    rows.sort(key=lambda row: (row["item_name"].lower(), row["size_label"].lower(), row["is_parcel"]))
    for index, row in enumerate(rows, start=1):
        extras = []
        if row["size_label"]:
            extras.append(row["size_label"])
        if row["is_parcel"]:
            extras.append("Parcel")
        row["display_name"] = row["item_name"] + (f" ({' | '.join(extras)})" if extras else "")
        row["form_key"] = f"item_{index}"
    return rows


def _feedback_for_settlement(order: CafeOrder | None) -> CafeFeedback | None:
    if not order:
        return None
    receipt_orders = _receipt_orders_for_settlement(order)
    primary_order = receipt_orders[0]
    return (
        CafeFeedback.query.options(joinedload(CafeFeedback.items), joinedload(CafeFeedback.table))
        .filter(CafeFeedback.primary_order_id == primary_order.id)
        .order_by(CafeFeedback.submitted_at.desc().nullslast(), CafeFeedback.id.desc())
        .first()
    )


def _latest_paid_order_for_table(table_id: int | None) -> CafeOrder | None:
    if not table_id:
        return None
    today_start, today_end = _current_ist_day_bounds()
    return (
        CafeOrder.query.options(joinedload(CafeOrder.table), joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item))
        .filter(
            CafeOrder.table_id == table_id,
            CafeOrder.status == "paid",
            CafeOrder.paid_at.is_not(None),
            CafeOrder.paid_at >= today_start,
            CafeOrder.paid_at < today_end,
        )
        .order_by(CafeOrder.paid_at.desc(), CafeOrder.id.desc())
        .first()
    )


def _feedback_prompt_payload(order: CafeOrder | None) -> dict | None:
    if not order:
        return None
    receipt_orders = _receipt_orders_for_settlement(order)
    primary_order = receipt_orders[0]
    feedback = _feedback_for_settlement(primary_order)
    return {
        "primary_order_id": primary_order.id,
        "table_name": primary_order.table.name if primary_order.table else "-",
        "settlement_total": round(
            sum(float(settled_order.total_amount or 0) for settled_order in receipt_orders),
            2,
        ),
        "paid_at": _format_ist(primary_order.paid_at, "%d %b %Y, %I:%M %p") if primary_order.paid_at else "",
        "feedback_exists": bool(feedback),
        "feedback_source": (feedback.source if feedback else ""),
    }


def _upsert_feedback_for_settlement(
    order: CafeOrder,
    *,
    source: str,
    item_rows: list[dict],
    default_rating: int,
    submitted_by_user: User | None = None,
    submitted_by_name: str = "",
) -> tuple[CafeFeedback, bool]:
    receipt_orders = _receipt_orders_for_settlement(order)
    primary_order = receipt_orders[0]
    feedback = _feedback_for_settlement(primary_order)
    created = False
    if not feedback:
        feedback = CafeFeedback(
            table_id=primary_order.table_id,
            primary_order_id=primary_order.id,
        )
        db.session.add(feedback)
        created = True
    feedback.order_ids_json = json.dumps([int(settled_order.id) for settled_order in receipt_orders])
    feedback.source = source
    feedback.service_rating = _clamp_star_rating(request.form.get("service_rating"), default_rating)
    feedback.summary_text = (request.form.get("summary_text") or "").strip()
    feedback.submitted_by_user_id = submitted_by_user.id if submitted_by_user else None
    feedback.submitted_by_name = (
        submitted_by_name.strip()
        or (submitted_by_user.full_name if submitted_by_user else "")
        or feedback.submitted_by_name
    )
    feedback.submitted_at = datetime.utcnow()
    existing_items = {(item.menu_item_id, item.size_label or "", bool(item.is_parcel)): item for item in feedback.items}
    seen_keys = set()
    for row in item_rows:
        key = (row["menu_item_id"], row["size_label"] or "", bool(row["is_parcel"]))
        seen_keys.add(key)
        feedback_item = existing_items.get(key)
        if not feedback_item:
            feedback_item = CafeFeedbackItem(
                feedback=feedback,
                menu_item_id=row["menu_item_id"],
                order_item_id=row["order_item_id"],
                item_name=row["item_name"],
                size_label=row["size_label"] or None,
                is_parcel=bool(row["is_parcel"]),
            )
            db.session.add(feedback_item)
        feedback_item.item_name = row["item_name"]
        feedback_item.order_item_id = row["order_item_id"]
        feedback_item.rating = _clamp_star_rating(request.form.get(f'item_rating_{row["form_key"]}'), default_rating)
    for key, stale_item in existing_items.items():
        if key not in seen_keys:
            db.session.delete(stale_item)
    return feedback, created


def _has_valid_kiosk_access(token: str | None) -> bool:
    configured = _kiosk_token()
    supplied = (token or "").strip()
    return bool(configured and supplied and secrets.compare_digest(configured, supplied))


def _has_valid_reception_kiosk_access(token: str | None) -> bool:
    configured = _reception_kiosk_token()
    supplied = (token or "").strip()
    return bool(configured and supplied and secrets.compare_digest(configured, supplied))


def _kiosk_display_url(token: str, station: str = "kitchen") -> str:
    return url_for("cafe.kiosk_display", access_key=token, station=station)


def _reception_kiosk_url(token: str) -> str:
    return url_for("cafe.reception_kiosk", access_key=token)


def _reception_manifest_payload(access_key: str):
    scope = f"/cafe/reception/{access_key}"
    return {
        "name": "Brownberries Reception",
        "short_name": "Reception",
        "description": "Brownberries Cafe reception cashier kiosk.",
        "id": scope,
        "start_url": scope,
        "scope": scope,
        "display": "standalone",
        "orientation": "any",
        "background_color": "#fbf8f3",
        "theme_color": "#3a2419",
        "icons": [
            {
                "src": "/static/images/pwa-icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
            },
            {
                "src": "/static/images/pwa-icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
            },
            {
                "src": "/static/images/pwa-icon-512-maskable.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "maskable",
            },
        ],
    }


def _order_local_date(order: CafeOrder | None) -> date:
    local_dt = _ist_from_utc_naive(order.created_at if order else None)
    return local_dt.date() if local_dt else date.today()


def _current_ist_day_bounds():
    today_ist = datetime.now(IST_TZ).date()
    start_local = datetime.combine(today_ist, time.min).replace(tzinfo=IST_TZ)
    end_local = datetime.combine(today_ist, time.max).replace(tzinfo=IST_TZ)
    return (
        start_local.astimezone(UTC_TZ).replace(tzinfo=None),
        end_local.astimezone(UTC_TZ).replace(tzinfo=None),
    )


def _order_channel_code(order: CafeOrder | None) -> str:
    if order and order.is_delivery:
        return "O"
    return "D"


def _format_internal_order_code(order: CafeOrder | None, seq: int | None = None) -> str:
    if not order:
        return "-"
    local_day = _order_local_date(order)
    running_seq = int(seq if seq is not None else (order.daily_sequence or order.id or 0))
    return f"BB-{_order_channel_code(order)}-{local_day.strftime('%y%m%d')}-{running_seq:04d}"


def _format_pickup_number(order: CafeOrder | None, compact: bool = False) -> str:
    if not order:
        return "-"
    running_seq = int(order.daily_sequence or order.id or 0)
    if compact:
        return f"{running_seq}"
    return f"{_order_channel_code(order)}{running_seq:03d}"


def _normalize_single_order_code(order: CafeOrder, seq: int | None = None):
    if not order:
        return
    running_seq = int(seq if seq is not None else (order.daily_sequence or 0))
    if running_seq < 1:
        running_seq = int(order.id or 0)
    order.daily_sequence = running_seq
    order.display_code = _format_internal_order_code(order, running_seq)


def _backfill_order_codes():
    changed = False
    grouped: dict[date, list[CafeOrder]] = {}
    orders = CafeOrder.query.order_by(CafeOrder.created_at.asc(), CafeOrder.id.asc()).all()
    for order in orders:
        grouped.setdefault(_order_local_date(order), []).append(order)
    for _, day_orders in grouped.items():
        for idx, order in enumerate(day_orders, start=1):
            expected = _format_internal_order_code(order, idx)
            if order.daily_sequence != idx or (order.display_code or "") != expected:
                order.daily_sequence = idx
                order.display_code = expected
                changed = True
    if changed:
        db.session.commit()


def _backfill_paid_timestamps():
    changed = False
    paid_orders = CafeOrder.query.filter(
        CafeOrder.status == "paid",
        CafeOrder.paid_at.is_(None),
    ).all()
    for order in paid_orders:
        order.paid_at = order.created_at
        changed = True
    if changed:
        db.session.commit()


def _menu_form_state_from_request():
    form = request.form
    selected_category_ids = []
    for value in form.getlist("category_ids"):
        try:
            selected_category_ids.append(int(value))
        except (TypeError, ValueError):
            continue
    selected_category_ids = list(dict.fromkeys(selected_category_ids))
    size_rows = []
    idx = 1
    while True:
        size_name = form.get(f"size_name_{idx}")
        size_price = form.get(f"size_price_{idx}")
        if size_name is None and size_price is None:
            break
        size_rows.append(
            {
                "size": (size_name or "").strip(),
                "price": (size_price or "").strip(),
            }
        )
        idx += 1
    if not size_rows:
        size_rows = [{"size": "", "price": ""}, {"size": "", "price": ""}]
    return {
        "selected_category_ids": selected_category_ids,
        "menu_type_id": form.get("menu_type_id", "").strip(),
        "name": form.get("name", "").strip(),
        "prep_station": _normalize_prep_station(form.get("prep_station")),
        "chef_user_id": form.get("chef_user_id", "").strip(),
        "image_url": form.get("image_url", "").strip(),
        "short_description": form.get("short_description", "").strip(),
        "description": form.get("description", "").strip(),
        "calories": form.get("calories", "").strip(),
        "price": form.get("price", "").strip(),
        "has_size_variants": True if form.get("has_size_variants") else False,
        "size_rows": size_rows,
        "available": True if form.get("available") else False,
    }


def _default_menu_form_state():
    return {
        "selected_category_ids": [],
        "menu_type_id": "",
        "name": "",
        "prep_station": "",
        "chef_user_id": "",
        "image_url": "",
        "short_description": "",
        "description": "",
        "calories": "",
        "price": "",
        "has_size_variants": False,
        "size_rows": [{"size": "", "price": ""}, {"size": "", "price": ""}],
        "available": True,
    }


def _render_menu_page(active_menu_section: str = "catalog", add_form_state: dict | None = None):
    _ensure_workstations_seeded()
    items = MenuItem.query.filter(MenuItem.is_deleted.is_(False)).order_by(MenuItem.name).all()
    deleted_items = MenuItem.query.filter(MenuItem.is_deleted.is_(True)).order_by(MenuItem.updated_at.desc(), MenuItem.name.asc()).all()
    all_categories = MenuCategory.query.order_by(MenuCategory.name).all()
    workstation_options = _all_workstations(include_inactive=True)
    chef_options = _chef_options()
    workstation_name_map = {station.slug: station.name for station in workstation_options}
    chef_name_map = {chef.id: chef.full_name for chef in chef_options}
    item_category_map = {}
    item_size_map = {}
    for item in items + deleted_items:
        parsed = []
        if item.category_ids_json:
            try:
                raw = json.loads(item.category_ids_json)
                if isinstance(raw, list):
                    parsed = [int(x) for x in raw if str(x).isdigit()]
            except (TypeError, ValueError, json.JSONDecodeError):
                parsed = []
        if not parsed and item.category_id:
            parsed = [item.category_id]
        item_category_map[item.id] = parsed
        sizes = []
        if item.size_pricing_json:
            try:
                raw_sizes = json.loads(item.size_pricing_json)
                if isinstance(raw_sizes, list):
                    for row in raw_sizes:
                        if isinstance(row, dict) and row.get("size") and row.get("price") is not None:
                            sizes.append({"size": str(row["size"]), "price": float(row["price"])})
            except (TypeError, ValueError, json.JSONDecodeError):
                sizes = []
        item_size_map[item.id] = sizes

    selected_category_filter = request.args.get("category_filter", type=int)
    availability_items = _apply_category_filter(
        MenuItem.query.filter(MenuItem.is_deleted.is_(False)).order_by(MenuItem.name.asc()),
        selected_category_filter,
    ).all()
    if active_menu_section not in ["catalog", "add_item", "items", "availability", "deleted_items"]:
        active_menu_section = "catalog"
    return render_template(
        "cafe/menu.html",
        items=items,
        deleted_items=deleted_items,
        categories=all_categories,
        protected_category_ids=[c.id for c in all_categories if _is_protected_menu_category(c)],
        menu_types=MenuType.query.order_by(MenuType.name).all(),
        workstation_options=workstation_options,
        workstation_name_map=workstation_name_map,
        chef_options=chef_options,
        chef_name_map=chef_name_map,
        item_category_map=item_category_map,
        item_size_map=item_size_map,
        active_menu_section=active_menu_section,
        availability_items=availability_items,
        selected_category_filter=selected_category_filter,
        add_form=add_form_state or _default_menu_form_state(),
    )


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _parse_time_only(raw_value: str | None):
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def _parse_cutoff_time(raw_value: str | None):
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def _format_cutoff_value(raw_value: str | None) -> str:
    parsed = _parse_cutoff_time(raw_value)
    return parsed.strftime("%H:%M") if parsed else ""


def _order_cutoff_message(channel: str):
    cfg = load_deployment_config(current_app.instance_path)
    key = "QR_ORDER_CUTOFF_TIME" if channel == "qr" else "STAFF_ORDER_CUTOFF_TIME"
    label = "table QR ordering" if channel == "qr" else "staff assisted table ordering"
    parsed = _parse_cutoff_time(cfg.get(key))
    if not parsed:
        return None
    now_ist = datetime.now(IST_TZ).time().replace(second=0, microsecond=0)
    if now_ist <= parsed:
        return None
    return f"{label.title()} is closed for today after {parsed.strftime('%I:%M %p')}."


def _split_multiline(value: str | None) -> list[str]:
    if not value:
        return []
    return [line.strip() for line in str(value).splitlines() if line.strip()]


def _load_menu_item_size_variants(menu_item: MenuItem | None) -> list[dict]:
    if not menu_item or not menu_item.has_size_variants:
        return []
    try:
        raw = json.loads(menu_item.size_pricing_json or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        raw = []
    rows = []
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        label = str(entry.get("size") or "").strip()
        if not label:
            continue
        rows.append(
            {
                "size": label,
                "price": round(_safe_float(entry.get("price"), 0), 2),
            }
        )
    return rows


def _parse_recipe_size_notes_from_form() -> list[dict]:
    labels = request.form.getlist("size_sop_label[]")
    notes = request.form.getlist("size_sop_note[]")
    rows = []
    for idx, raw_label in enumerate(labels):
        label = (raw_label or "").strip()
        note = (notes[idx] if idx < len(notes) else "").strip()
        if not label:
            continue
        rows.append({"size": label, "note": note})
    return rows


def _recipe_size_note_map(recipe: InventoryRecipe | None) -> dict[str, str]:
    if not recipe or not recipe.size_sop_json:
        return {}
    try:
        rows = json.loads(recipe.size_sop_json or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        rows = []
    result = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        label = str(row.get("size") or "").strip()
        note = str(row.get("note") or "").strip()
        if label:
            result[label] = note
    return result


def _serialize_recipe_sop(recipe: InventoryRecipe | None, size_label: str | None = None) -> dict:
    recipe = recipe or None
    size_label = (size_label or "").strip()
    size_note_map = _recipe_size_note_map(recipe)
    ingredients = []
    if recipe:
        for ing in recipe.ingredients:
            if not ing.inventory_item:
                continue
            ingredients.append(
                {
                    "name": ing.inventory_item.name,
                    "qty": round(float(ing.qty_per_menu or 0), 3),
                    "unit": ing.unit or ing.inventory_item.unit or "",
                }
            )
    return {
        "prep_time_minutes": int(recipe.prep_time_minutes or 0) if recipe else 0,
        "yield_qty": round(float(recipe.yield_qty or 0), 3) if recipe else 0,
        "yield_unit": (recipe.yield_unit or "").strip() if recipe else "",
        "ingredients": ingredients,
        "ingredients_note": _split_multiline(recipe.ingredients_note if recipe else None),
        "preparation_steps": _split_multiline(recipe.preparation_steps if recipe else None),
        "plating_notes": _split_multiline(recipe.plating_notes if recipe else None),
        "quality_checks": _split_multiline(recipe.quality_checks if recipe else None),
        "allergy_alerts": _split_multiline(recipe.allergy_alerts if recipe else None),
        "training_notes": _split_multiline(recipe.training_notes if recipe else None),
        "photo_url": (recipe.sop_photo_url or "").strip() if recipe else "",
        "size_label": size_label,
        "size_note": size_note_map.get(size_label, "") if size_label else "",
        "size_notes": [{"size": key, "note": value} for key, value in size_note_map.items()],
    }


def _utc_naive_from_ist(dt_value: datetime) -> datetime:
    localized = dt_value.replace(tzinfo=IST_TZ) if dt_value.tzinfo is None else dt_value.astimezone(IST_TZ)
    return localized.astimezone(UTC_TZ).replace(tzinfo=None)


def _ist_from_utc_naive(dt_value: datetime | None):
    if not dt_value:
        return None
    aware_utc = dt_value.replace(tzinfo=UTC_TZ) if dt_value.tzinfo is None else dt_value.astimezone(UTC_TZ)
    return aware_utc.astimezone(IST_TZ)


def _format_ist(dt_value: datetime | None, fmt: str = "%Y-%m-%d %H:%M"):
    local = _ist_from_utc_naive(dt_value)
    return local.strftime(fmt) if local else "-"


def _emit_ops_notification(message: str, kind: str = "order", payload: dict | None = None):
    data = {"kind": kind, "message": message}
    if payload:
        data.update(payload)
    socketio.emit("ops_notification", data, namespace="/kitchen")
    socketio.emit("ops_notification", data, namespace="/table")


def _is_staff_user(user: User) -> bool:
    return bool(set(user.assigned_roles()) & set(_get_role_options())) and user.email != "qr.guest@brownberries.local"


def _ensure_staff_profile(user: User) -> StaffProfile:
    profile = user.staff_profile
    if not profile:
        profile = StaffProfile(user_id=user.id, archived=not user.active)
        db.session.add(profile)
        db.session.flush()
    return profile


def _is_protected_admin(user: User) -> bool:
    return user.email == "admin@brownberries.local" or (
        user_has_any_role(user, "admin") and user.full_name.strip().lower() == "cafe admin"
    )


def _get_role_options():
    names = {r for r in DEFAULT_ROLE_OPTIONS}
    names.update({(ut.name or "").strip() for ut in UserType.query.all() if (ut.name or "").strip()})
    names.update({(u.role or "").strip() for u in User.query.all() if (u.role or "").strip()})
    return tuple(sorted(names))


def _parse_roles_from_form():
    selected = []
    for value in request.form.getlist("roles"):
        role_name = (value or "").strip().lower()
        if role_name and role_name not in selected and role_name in {r.lower(): r for r in _get_role_options()}:
            selected.append(role_name)
    if not selected:
        fallback = (request.form.get("role") or "staff").strip().lower()
        if fallback:
            selected.append(fallback)
    return selected or ["staff"]


def _ensure_role_templates_exist():
    existing = {ut.name.lower(): ut for ut in UserType.query.all() if ut.name}
    changed = False
    for role_name in _get_role_options():
        if role_name.lower() not in existing:
            ut = UserType(name=role_name)
            ut.can_access_cafe = True
            ut.can_manage_orders = True
            ut.can_manage_kitchen = True
            if role_name.lower() == "delivery_partner":
                ut.can_view_delivery_locations = True
            db.session.add(ut)
            changed = True
    if changed:
        db.session.commit()


def _assign_user_type_from_role(user: User):
    primary_role = (user.role or "").strip()
    if not primary_role:
        return
    ut = UserType.query.filter(db.func.lower(UserType.name) == primary_role.lower()).first()
    user.user_type_id = ut.id if ut else None


def _save_uploaded_file(file_obj, subdir: str, prefix: str):
    if not file_obj or not file_obj.filename:
        return None
    root = current_app.config["UPLOADS_ROOT"]
    ext = os.path.splitext(secure_filename(file_obj.filename))[1].lower() or ".bin"
    filename = f"{prefix}-{uuid4().hex[:10]}{ext}"
    folder = os.path.join(root, subdir)
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, filename)
    file_obj.save(full_path)
    return os.path.relpath(full_path, root)


def _save_menu_image(file_obj):
    if not file_obj or not file_obj.filename:
        return None
    ext = os.path.splitext(secure_filename(file_obj.filename))[1].lower()
    filename = f"menu-{uuid4().hex[:12]}.webp"
    folder = os.path.join(current_app.static_folder, "uploads", "menu")
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, filename)
    try:
        img = Image.open(file_obj.stream).convert("RGB")
        img.thumbnail((900, 900))
        img.save(full_path, format="WEBP", quality=78, optimize=True, method=6)
    except Exception:
        return None
    return f"/static/uploads/menu/{filename}"


def _get_or_create_other_category():
    other = MenuCategory.query.filter(db.func.lower(MenuCategory.name) == "other").first()
    if other:
        return other
    other = MenuCategory(name="Other")
    db.session.add(other)
    db.session.flush()
    return other


def _get_or_create_utility_category():
    utility = MenuCategory.query.filter(db.func.lower(MenuCategory.name) == "utility").first()
    if utility:
        return utility
    utility = MenuCategory(name="Utility")
    db.session.add(utility)
    db.session.flush()
    return utility


def _ensure_protected_menu_categories():
    _get_or_create_other_category()
    _get_or_create_utility_category()
    db.session.flush()


def _is_protected_menu_category(category: MenuCategory | None) -> bool:
    return bool(category and (category.name or "").strip().lower() in PROTECTED_MENU_CATEGORY_NAMES)


def _public_menu_category_ids(item: MenuItem, category_name_by_id: dict[int, str]) -> list[int]:
    visible_ids: list[int] = []
    seen: set[int] = set()
    for cid in _menu_item_category_ids(item):
        cname = (category_name_by_id.get(cid) or "").strip().lower()
        if not cname or cname in PROTECTED_MENU_CATEGORY_NAMES or cid in seen:
            continue
        visible_ids.append(cid)
        seen.add(cid)
    return visible_ids


def _is_public_menu_item(item: MenuItem, category_name_by_id: dict[int, str]) -> bool:
    return len(_public_menu_category_ids(item, category_name_by_id)) > 0


def _menu_item_category_ids(item: MenuItem) -> list[int]:
    parsed: list[int] = []
    if item.category_ids_json:
        try:
            raw = json.loads(item.category_ids_json)
            if isinstance(raw, list):
                for value in raw:
                    try:
                        parsed.append(int(value))
                    except (TypeError, ValueError):
                        continue
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    if not parsed and item.category_id:
        parsed = [item.category_id]
    return list(dict.fromkeys(parsed))


def _get_item_category_names(item: MenuItem, category_name_by_id: dict[int, str], include_protected: bool = True) -> list[str]:
    names: list[str] = []
    names_seen: set[str] = set()
    for cid in _menu_item_category_ids(item):
        cname = category_name_by_id.get(cid)
        if not include_protected and cname and cname.strip().lower() in PROTECTED_MENU_CATEGORY_NAMES:
            continue
        if cname and cname.lower() not in names_seen:
            names.append(cname)
            names_seen.add(cname.lower())
    return names


def _visible_categories_for_available_menu(include_protected: bool = False) -> list[MenuCategory]:
    all_categories = MenuCategory.query.order_by(MenuCategory.name.asc()).all()
    category_name_by_id = {c.id: c.name for c in all_categories}
    categories = list(all_categories)
    if not include_protected:
        categories = [c for c in categories if (c.name or "").strip().lower() not in PROTECTED_MENU_CATEGORY_NAMES]
    available_items = MenuItem.query.filter_by(available=True, is_deleted=False).all()
    used_category_ids: set[int] = set()
    for item in available_items:
        category_ids = (
            _menu_item_category_ids(item)
            if include_protected
            else _public_menu_category_ids(item, category_name_by_id)
        )
        for cid in category_ids:
            used_category_ids.add(cid)
    return [
        c
        for c in categories
        if c.id in used_category_ids
    ]


def _build_staff_id_card(user: User, profile: StaffProfile):
    canvas = Image.new("RGB", (960, 540), "#f6eee5")
    draw = ImageDraw.Draw(canvas)
    title_font = ImageFont.load_default()
    text_font = ImageFont.load_default()
    draw.rectangle((24, 24, 936, 516), outline="#3f2b1d", width=4)
    logo_path = os.path.join(current_app.static_folder, "images", "cafe-logo.png")
    if os.path.exists(logo_path):
        logo = Image.open(logo_path).convert("RGBA").resize((140, 140))
        canvas.paste(logo, (50, 42), logo)
    draw.text((220, 58), "Brownberries Cafe Staff ID", fill="#2b1d13", font=title_font)
    draw.text((220, 95), f"Name: {user.full_name}", fill="#2b1d13", font=text_font)
    draw.text((220, 122), f"Role: {user.display_roles()}", fill="#2b1d13", font=text_font)
    draw.text((220, 149), f"Email: {user.email}", fill="#2b1d13", font=text_font)
    draw.text((220, 176), f"Phone: {profile.phone or '-'}", fill="#2b1d13", font=text_font)
    draw.text((220, 203), f"DOB: {profile.dob or '-'}", fill="#2b1d13", font=text_font)
    draw.text((220, 230), f"Gender: {profile.gender or '-'}", fill="#2b1d13", font=text_font)
    draw.text((220, 257), f"Govt ID: {(profile.govt_id_type or '-')} / {(profile.govt_id_number or '-')}", fill="#2b1d13", font=text_font)
    draw.text((220, 284), f"Joined: {profile.joining_date or '-'}", fill="#2b1d13", font=text_font)
    photo_box = (50, 220, 180, 380)
    draw.rectangle(photo_box, outline="#7d6654", width=2)
    if profile.photo_file_path:
        photo_path = os.path.join(current_app.config["UPLOADS_ROOT"], profile.photo_file_path)
        if os.path.exists(photo_path):
            photo = Image.open(photo_path).convert("RGB").resize((120, 150))
            canvas.paste(photo, (55, 225))
    out = BytesIO()
    canvas.save(out, format="PNG")
    out.seek(0)
    return out


def _normalized_item_prep_status(item: CafeOrderItem | None) -> str:
    raw = ((item.prep_status if item else "") or "").strip().lower()
    return raw if raw in ITEM_PREP_STATUSES else "pending"


def _refresh_order_status_from_items(order: CafeOrder | None) -> str:
    if not order:
        return "open"
    pending_left = any((item.approval_status or "pending") == "pending" for item in order.order_items)
    approved_items = [item for item in order.order_items if (item.approval_status or "pending") == "approved"]
    if not approved_items:
        order.status = "pending_approval" if pending_left else "cancelled"
        return order.status

    prep_states = [_normalized_item_prep_status(item) for item in approved_items]
    if all(state == "served" for state in prep_states):
        order.status = "served"
    elif any(state == "preparing" for state in prep_states):
        order.status = "preparing"
    elif any(state == "ready" for state in prep_states):
        order.status = "ready"
    else:
        order.status = "open"
    return order.status


def _serialize_order(order: CafeOrder):
    pending_count = sum(1 for item in order.order_items if (item.approval_status or "pending") == "pending")
    try:
        payment_breakdown = json.loads(order.payment_breakdown_json or "[]")
        if not isinstance(payment_breakdown, list):
            payment_breakdown = []
    except (TypeError, ValueError, json.JSONDecodeError):
        payment_breakdown = []
    return {
        "id": order.id,
        "order_code": order.display_code or _format_internal_order_code(order),
        "pickup_no": _format_pickup_number(order),
        "pickup_no_compact": _format_pickup_number(order, compact=True),
        "table_id": order.table_id,
        "table": "For Delivery" if order.is_delivery else (order.table.name if order.table else "-"),
        "status": order.status,
        "payment_type": order.payment_type or "-",
        "payment_breakdown": payment_breakdown,
        "total_amount": round(order.total_amount, 2),
        "is_delivery": bool(order.is_delivery),
        "delivery_customer_name": order.delivery_customer_name,
        "delivery_customer_mobile": order.delivery_customer_mobile,
        "delivery_address": order.delivery_address,
        "delivery_map_url": order.delivery_map_url,
        "packaging_charge": round(order.packaging_charge or 0, 2),
        "delivery_distance_km": round(order.delivery_distance_km or 0, 2),
        "delivery_charge": round(order.delivery_charge or 0, 2),
        "service_tax_amount": round(order.service_tax_amount or 0, 2),
        "gst_amount": 0.0,
        "cst_amount": 0.0,
        "tax_amount": 0.0,
        "created_at": _format_ist(order.created_at),
        "created_on_ist": _order_local_date(order).isoformat(),
        "pending_approval_count": pending_count,
        "items": [
            {
                "id": item.id,
                "name": item.menu_item.name if item.menu_item else "-",
                "qty": item.quantity,
                "size_label": item.size_label,
                "is_parcel": bool(item.is_parcel),
                "prep_station": item.menu_item.prep_station if item.menu_item else "kitchen",
                "approval_status": item.approval_status or "pending",
                "prep_status": _normalized_item_prep_status(item),
            }
            for item in order.order_items
        ],
    }


def _ensure_menu_types_seeded():
    existing_types = {t.name.lower(): t for t in MenuType.query.all()}
    item_types = (
        db.session.query(MenuItem.item_type)
        .filter(MenuItem.item_type.isnot(None), MenuItem.item_type != "")
        .distinct()
        .all()
    )
    created = False
    for row in item_types:
        name = row[0].strip()
        if name and name.lower() not in existing_types:
            db.session.add(MenuType(name=name))
            created = True
    if created:
        db.session.commit()


def _parse_category_ids_from_form():
    raw_ids = request.form.getlist("category_ids")
    parsed = []
    for value in raw_ids:
        try:
            cid = int(value)
        except (TypeError, ValueError):
            continue
        if cid not in parsed:
            parsed.append(cid)
    return parsed


def _parse_size_pricing_from_form():
    pairs = []
    indexes = []
    for key in request.form.keys():
        if key.startswith("size_name_"):
            suffix = key.removeprefix("size_name_")
            if suffix.isdigit():
                indexes.append(int(suffix))
    if not indexes:
        indexes = [1, 2]
    for i in sorted(set(indexes)):
        name = (request.form.get(f"size_name_{i}") or "").strip()
        price_raw = (request.form.get(f"size_price_{i}") or "").strip()
        if not name:
            continue
        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            continue
        if price < 0:
            continue
        pairs.append({"size": name, "price": round(price, 2)})
    return pairs


def _apply_category_filter(query, category_id: int | None):
    if not category_id:
        return query
    return query.filter(
        db.or_(
            MenuItem.category_id == category_id,
            MenuItem.category_ids_json.ilike(f"%{category_id}%"),
        )
    )


def _get_qr_guest_user_id() -> int:
    guest_email = "qr.guest@brownberries.local"
    guest = User.query.filter_by(email=guest_email).first()
    if guest:
        return guest.id
    guest = User(
        full_name="QR Guest",
        email=guest_email,
        password_hash=generate_password_hash("qr-guest"),
        role="staff",
        active=True,
    )
    db.session.add(guest)
    db.session.commit()
    return guest.id


def _parse_line_items_from_request():
    cart_payload = (request.form.get("cart_payload") or "").strip()
    cart_entries = []
    if cart_payload:
        try:
            parsed = json.loads(cart_payload)
            if isinstance(parsed, list):
                for row in parsed:
                    if not isinstance(row, dict):
                        continue
                    try:
                        menu_item_id = int(row.get("menu_item_id"))
                        quantity = int(row.get("quantity"))
                        is_parcel = True if row.get("is_parcel") else False
                        size_label = (row.get("size_label") or "").strip() or None
                        unit_price = float(row.get("unit_price")) if row.get("unit_price") is not None else None
                    except (TypeError, ValueError):
                        continue
                    if quantity > 0:
                        cart_entries.append((menu_item_id, quantity, is_parcel, size_label, unit_price))
        except (json.JSONDecodeError, TypeError):
            pass

    if not cart_entries:
        for value in request.form.getlist("menu_item_id"):
            try:
                menu_item_id = int(value)
                quantity = int(request.form.get(f"qty_{menu_item_id}", 1))
                if quantity > 0:
                    cart_entries.append((menu_item_id, quantity, False, None, None))
            except (TypeError, ValueError):
                continue

    if not cart_entries:
        return []

    merged = {}
    for menu_item_id, quantity, is_parcel, size_label, unit_price in cart_entries:
        key = (menu_item_id, bool(is_parcel), size_label or "", unit_price if unit_price is not None else -1.0)
        merged[key] = merged.get(key, 0) + quantity

    item_ids = list({key[0] for key in merged.keys()})
    items = MenuItem.query.filter(
        MenuItem.id.in_(item_ids),
        MenuItem.available.is_(True),
        MenuItem.is_deleted.is_(False),
    ).all()
    item_by_id = {item.id: item for item in items}
    line_items = []
    for (menu_item_id, is_parcel, size_label, unit_price_key), quantity in merged.items():
        menu_item = item_by_id.get(menu_item_id)
        if menu_item:
            line_items.append((menu_item, quantity, is_parcel, size_label or None, None if unit_price_key == -1.0 else unit_price_key))
    return line_items


def _find_default_water_menu_item():
    item = (
        MenuItem.query.filter(
            db.func.lower(MenuItem.name) == "water",
            MenuItem.is_deleted.is_(False),
        )
        .order_by(MenuItem.available.desc(), MenuItem.id.asc())
        .first()
    )
    if item:
        return item
    return (
        MenuItem.query.filter(
            db.func.lower(MenuItem.name).like("water%"),
            MenuItem.is_deleted.is_(False),
        )
        .order_by(MenuItem.available.desc(), MenuItem.id.asc())
        .first()
    )


def _append_auto_water_if_needed(table_id: int, line_items, force_first_cycle: bool = False):
    if not line_items:
        return line_items
    table = CafeTable.query.get(table_id)
    if not table or (table.name or "").strip().upper() == "DL":
        return line_items
    if not force_first_cycle:
        today_start, today_end = _current_ist_day_bounds()
        has_active_orders = db.session.query(CafeOrder.id).filter(
            CafeOrder.table_id == table_id,
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
        ).first()
        if has_active_orders:
            return line_items
    water_item = _find_default_water_menu_item()
    if not water_item:
        return line_items
    updated_items = list(line_items or [])
    if water_item.has_size_variants and water_item.size_pricing_json:
        try:
            variants = json.loads(water_item.size_pricing_json)
            if isinstance(variants, list) and variants:
                first = variants[0] if isinstance(variants[0], dict) else {}
                size_label = (first.get("size") or "").strip() or None
                unit_price = float(first.get("price")) if first.get("price") is not None else float(water_item.price)
                updated_items.append((water_item, 1, False, size_label, unit_price))
                return updated_items
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    updated_items.append((water_item, 1, False, None, float(water_item.price)))
    return updated_items


def _normalized_line_item_signature(line_items) -> list[tuple[int, int, int, str, float]]:
    signature = []
    for row in line_items or []:
        if len(row) == 5:
            menu_item, qty, is_parcel, size_label, unit_price = row
        elif len(row) == 3:
            menu_item, qty, is_parcel = row
            size_label = None
            unit_price = None
        else:
            menu_item, qty = row
            is_parcel = False
            size_label = None
            unit_price = None
        signature.append(
            (
                int(menu_item.id),
                int(qty or 0),
                1 if is_parcel else 0,
                (size_label or "").strip(),
                round(float(unit_price if unit_price is not None else menu_item.price or 0), 2),
            )
        )
    signature.sort()
    return signature


def _order_signature(order: CafeOrder) -> list[tuple[int, int, int, str, float]]:
    signature = []
    for item in order.order_items:
        signature.append(
            (
                int(item.menu_item_id),
                int(item.quantity or 0),
                1 if item.is_parcel else 0,
                (item.size_label or "").strip(),
                round(float(item.unit_price or 0), 2),
            )
        )
    signature.sort()
    return signature


def _cancel_orphan_active_orders(table_id: int):
    candidate_orders = (
        CafeOrder.query.options(joinedload(CafeOrder.order_items))
        .filter(
            CafeOrder.table_id == table_id,
            CafeOrder.status.notin_(["paid", "cancelled"]),
        )
        .all()
    )
    changed = False
    for order in candidate_orders:
        if order.order_items:
            continue
        order.status = "cancelled"
        changed = True
    if changed:
        db.session.commit()


def _find_recent_duplicate_order(table_id: int, ordered_by_user_id: int, status: str, allowed_signatures, window_seconds: int = 2):
    if not allowed_signatures:
        return None
    created_after = datetime.utcnow() - timedelta(seconds=window_seconds)
    candidates = (
        CafeOrder.query.options(joinedload(CafeOrder.order_items))
        .filter(
            CafeOrder.table_id == table_id,
            CafeOrder.ordered_by_user_id == ordered_by_user_id,
            CafeOrder.status == status,
            CafeOrder.created_at >= created_after,
        )
        .order_by(CafeOrder.created_at.desc())
        .all()
    )
    for order in candidates:
        if not order.order_items:
            order.status = "cancelled"
            db.session.commit()
            continue
        if tuple(_order_signature(order)) in allowed_signatures:
            return order
    return None


def _haversine_km(lat1, lng1, lat2, lng2):
    radius_km = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lng = math.radians(lng2 - lng1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lng / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius_km * c


def _create_order(table_id: int, ordered_by_user_id: int, status: str, payment_type, payment_reference):
    raw_line_items = _parse_line_items_from_request()
    if not raw_line_items:
        return None
    _cancel_orphan_active_orders(table_id)
    allowed_signatures = {
        tuple(_normalized_line_item_signature(raw_line_items)),
        tuple(_normalized_line_item_signature(_append_auto_water_if_needed(table_id, raw_line_items, force_first_cycle=True))),
    }
    duplicate_order = _find_recent_duplicate_order(table_id, ordered_by_user_id, status, allowed_signatures)
    if duplicate_order:
        return duplicate_order
    line_items = _append_auto_water_if_needed(table_id, raw_line_items)
    return create_cafe_order(
        table_id=table_id,
        ordered_by_user_id=ordered_by_user_id,
        line_items=line_items,
        status=status,
        payment_type=payment_type,
        payment_reference=payment_reference,
    )


def _recalculate_order_totals(order: CafeOrder):
    line_total = 0.0
    packaging_charge = 0.0
    order_items = (
        CafeOrderItem.query.filter_by(order_id=order.id)
        .order_by(CafeOrderItem.id.asc())
        .all()
    )
    for oi in order_items:
        if (oi.approval_status or "pending") == "rejected":
            continue
        line_total += float(oi.unit_price or 0) * int(oi.quantity or 0)
        if order.is_delivery:
            packaging_charge += 20.0 * int(oi.quantity or 0)
        elif oi.is_parcel:
            packaging_charge += 20.0 * int(oi.quantity or 0)
    order.packaging_charge = round(packaging_charge, 2)
    order.service_tax_amount = 0.0
    order.gst_amount = 0.0
    order.cst_amount = 0.0
    order.total_amount = round(line_total + order.packaging_charge + float(order.delivery_charge or 0), 2)


def _apply_recipe_inventory_deduction(order: CafeOrder):
    if not order or not order.order_items:
        return
    item_qty_map: dict[int, int] = {}
    for oi in order.order_items:
        item_qty_map[oi.menu_item_id] = item_qty_map.get(oi.menu_item_id, 0) + int(oi.quantity or 0)
    if not item_qty_map:
        return
    recipes = InventoryRecipe.query.filter(
        InventoryRecipe.menu_item_id.in_(list(item_qty_map.keys())),
        InventoryRecipe.active.is_(True),
    ).all()
    recipe_by_menu = {r.menu_item_id: r for r in recipes}
    for menu_item_id, qty in item_qty_map.items():
        recipe = recipe_by_menu.get(menu_item_id)
        if not recipe:
            continue
        for ing in recipe.ingredients:
            inv_item = ing.inventory_item
            if not inv_item:
                continue
            deduct_qty = float(ing.qty_per_menu or 0) * float(qty)
            inv_item.current_amount = round(max(0.0, float(inv_item.current_amount or 0) - deduct_qty), 3)


def _receipt_link_for_order(order: CafeOrder) -> str:
    base = (current_app.config.get("PUBLIC_BASE_URL") or request.host_url.rstrip("/")).rstrip("/")
    return f"{base}/cafe/receipt/{order.id}"


def _parse_split_payment_rows():
    rows = []
    row_count = int(request.form.get("payment_row_count") or 1)
    row_count = max(1, min(row_count, 12))
    for idx in range(1, row_count + 1):
        payment_type = (request.form.get(f"payment_type_{idx}") or "").strip()
        amount = _safe_float(request.form.get(f"payment_amount_{idx}"), 0)
        reference = (request.form.get(f"payment_reference_{idx}") or "").strip()
        if not payment_type and amount <= 0 and not reference:
            continue
        if not payment_type or amount <= 0:
            return None, "Each payment row must have a method and an amount."
        rows.append(
            {
                "method": payment_type,
                "amount": round(float(amount), 2),
                "reference": reference,
            }
        )
    if not rows:
        return None, "Add at least one payment entry."
    return rows, ""


def _send_receipt_sms(country_code: str, mobile: str, message: str):
    return send_sms_from_config(current_app.instance_path, country_code, mobile, message)


def _receipt_sms_message(order: CafeOrder) -> str:
    public_no = _format_pickup_number(order)
    internal_ref = order.display_code or _format_internal_order_code(order)
    paid_label = _format_ist(order.paid_at or datetime.utcnow(), "%d %b %Y, %I:%M %p")
    total_amount = round(float(order.total_amount or 0), 2)
    payment_label = (order.payment_type or "Payment").strip()
    table_label = "Delivery" if order.is_delivery else (order.table.name if order.table else "Cafe")
    receipt_link = _receipt_link_for_order(order)
    return (
        "Brownberries Cafe\n"
        f"Receipt: Order #{public_no}\n"
        f"Ref: {internal_ref}\n"
        f"Table: {table_label}\n"
        f"Total: Rs {total_amount:.2f}\n"
        f"Paid via: {payment_label}\n"
        f"Time: {paid_label}\n"
        f"Receipt Link: {receipt_link}"
    )


def create_cafe_order(
    table_id: int,
    ordered_by_user_id: int,
    line_items,
    status: str = "open",
    payment_type=None,
    payment_reference=None,
    is_delivery: bool = False,
    delivery_customer_name: str | None = None,
    delivery_customer_mobile: str | None = None,
    delivery_address: str | None = None,
    delivery_lat: float | None = None,
    delivery_lng: float | None = None,
    delivery_map_url: str | None = None,
):
    if not line_items:
        return None
    today = datetime.now(IST_TZ).date()
    day_start_utc = _utc_naive_from_ist(datetime.combine(today, time.min))
    day_end_utc = _utc_naive_from_ist(datetime.combine(today + timedelta(days=1), time.min))
    today_count = CafeOrder.query.filter(
        CafeOrder.created_at >= day_start_utc,
        CafeOrder.created_at < day_end_utc,
    ).count()
    daily_seq = today_count + 1

    order = CafeOrder(
        table_id=table_id,
        ordered_by_user_id=ordered_by_user_id,
        status=status,
        payment_type=payment_type,
        payment_reference=payment_reference,
        is_delivery=is_delivery,
        delivery_customer_name=delivery_customer_name,
        delivery_customer_mobile=delivery_customer_mobile,
        delivery_address=delivery_address,
        delivery_lat=delivery_lat,
        delivery_lng=delivery_lng,
        delivery_map_url=delivery_map_url,
        daily_sequence=daily_seq,
        display_code="",
    )
    db.session.add(order)
    db.session.flush()
    _normalize_single_order_code(order, daily_seq)

    for row in line_items:
        if len(row) == 5:
            menu_item, qty, is_parcel, size_label, unit_price = row
        elif len(row) == 3:
            menu_item, qty, is_parcel = row
            size_label = None
            unit_price = None
        else:
            menu_item, qty = row
            is_parcel = False
            size_label = None
            unit_price = None
        price_to_use = float(unit_price) if unit_price is not None else float(menu_item.price)
        db.session.add(
            CafeOrderItem(
                order_id=order.id,
                menu_item_id=menu_item.id,
                quantity=qty,
                unit_price=price_to_use,
                size_label=size_label,
                is_parcel=bool(is_parcel),
                approval_status="pending" if status == "pending_approval" else "approved",
                prep_status="pending",
            )
        )

    delivery_distance_km = 0.0
    delivery_charge = 0.0
    if is_delivery and delivery_lat is not None and delivery_lng is not None:
        cafe_lat = 25.207989477704068
        cafe_lng = 80.87374457551877
        direct_km = _haversine_km(cafe_lat, cafe_lng, float(delivery_lat), float(delivery_lng))
        delivery_distance_km = round(direct_km * 1.3, 2)
        delivery_charge = round(delivery_distance_km * 5.0, 2)

    order.delivery_distance_km = delivery_distance_km
    order.delivery_charge = delivery_charge
    _recalculate_order_totals(order)
    _refresh_order_status_from_items(order)
    _apply_recipe_inventory_deduction(order)
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_created", payload, namespace="/kitchen")
    socketio.emit("order_created", payload, namespace="/table")
    return order


@bp.route("/")
@login_required
def home():
    cfg = load_deployment_config(current_app.instance_path)
    tax_settings = _tax_settings()
    public_base = (current_app.config.get("PUBLIC_BASE_URL") or request.host_url.rstrip("/")).rstrip("/")
    attendance_settings = _attendance_settings()
    today_start, today_end = _current_ist_day_bounds()
    _ensure_workstations_seeded()
    workstation_options = _all_workstations()
    sms_enabled = str(cfg.get("SMS_ENABLED", "0")).strip() in ["1", "true", "True"]
    sms_ca_bundle = (cfg.get("SMS_CA_BUNDLE") or "").strip()
    sms_allow_insecure_ssl = str(cfg.get("SMS_ALLOW_INSECURE_SSL", "0")).strip() in ["1", "true", "True"]
    textbee_api_key = (cfg.get("TEXTBEE_API_KEY") or "").strip()
    textbee_device_id = (cfg.get("TEXTBEE_DEVICE_ID") or "").strip()
    textbee_sim_subscription_id = (cfg.get("TEXTBEE_SIM_SUBSCRIPTION_ID") or "").strip()
    textbee_base_url = (cfg.get("TEXTBEE_BASE_URL") or "https://api.textbee.dev/api/v1").strip()
    textbee_key_hint = f"{textbee_api_key[:4]}...{textbee_api_key[-4:]}" if len(textbee_api_key) >= 10 else ("Set" if textbee_api_key else "")
    qr_order_cutoff_time = _format_cutoff_value(cfg.get("QR_ORDER_CUTOFF_TIME"))
    staff_order_cutoff_time = _format_cutoff_value(cfg.get("STAFF_ORDER_CUTOFF_TIME"))
    kiosk_token = (cfg.get("KDS_KIOSK_TOKEN") or "").strip()
    reception_kiosk_token = (cfg.get("RECEPTION_KIOSK_TOKEN") or "").strip()
    return render_template(
        "cafe/home.html",
        tables=CafeTable.query.count(),
        open_orders=CafeOrder.query.filter(
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
        ).count(),
        low_stock=InventoryItem.query.filter(
            InventoryItem.current_amount <= InventoryItem.reorder_level
        ).count(),
        delivery_open_orders=CafeOrder.query.filter_by(is_delivery=True).filter(
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
        ).count(),
        upcoming_bookings=TableBooking.query.filter(
            TableBooking.booking_date >= date.today(),
            TableBooking.status == "booked",
        ).count(),
        public_notice_text=(cfg.get("PUBLIC_NOTICE_TEXT", "") or "").strip(),
        public_notice_enabled=(cfg.get("PUBLIC_NOTICE_ENABLED", "0") in [1, "1", True, "true", "True"]),
        sms_enabled=sms_enabled,
        sms_ca_bundle=sms_ca_bundle,
        sms_allow_insecure_ssl=sms_allow_insecure_ssl,
        textbee_key_hint=textbee_key_hint,
        textbee_device_id=textbee_device_id,
        textbee_sim_subscription_id=textbee_sim_subscription_id,
        textbee_base_url=textbee_base_url,
        qr_order_cutoff_time=qr_order_cutoff_time,
        staff_order_cutoff_time=staff_order_cutoff_time,
        service_charge_rate=tax_settings["service_charge_rate"],
        kiosk_token=kiosk_token,
        workstation_options=workstation_options,
        workstation_kiosk_urls=[
            {
                "slug": station.slug,
                "name": station.name,
                "url": f"{public_base}{_kiosk_display_url(kiosk_token, station.slug)}" if kiosk_token else "",
            }
            for station in workstation_options
        ],
        reception_kiosk_token=reception_kiosk_token,
        reception_kiosk_url=f"{public_base}{_reception_kiosk_url(reception_kiosk_token)}" if reception_kiosk_token else "",
        attendance_cafe_lat=attendance_settings["cafe_lat"],
        attendance_cafe_lng=attendance_settings["cafe_lng"],
        attendance_radius_m=attendance_settings["radius_m"],
        staff_attendance_qr_url=f"{public_base}{url_for('main.staff_attendance_check_in')}",
        staff_attendance_qr_png_url=url_for("cafe.staff_attendance_qr_png"),
    )


@bp.route("/public-notice", methods=["POST"])
@roles_required("admin", "manager")
def update_public_notice():
    text = (request.form.get("notice_text") or "").strip()
    enabled = True if request.form.get("notice_enabled") else False
    save_deployment_config(
        current_app.instance_path,
        {
            "PUBLIC_NOTICE_TEXT": text,
            "PUBLIC_NOTICE_ENABLED": "1" if (enabled and text) else "0",
        },
    )
    flash("Public notification updated.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/sms-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_sms_settings():
    cfg = load_deployment_config(current_app.instance_path)
    sms_enabled = True if request.form.get("sms_enabled") else False
    ca_bundle = (request.form.get("sms_ca_bundle") or "").strip()
    allow_insecure_ssl = True if request.form.get("sms_allow_insecure_ssl") else False
    textbee_api_key = (request.form.get("textbee_api_key") or "").strip()
    textbee_device_id = (request.form.get("textbee_device_id") or "").strip()
    textbee_sim_subscription_id = (request.form.get("textbee_sim_subscription_id") or "").strip()
    textbee_base_url = (request.form.get("textbee_base_url") or "").strip()

    updates = {
        "SMS_ENABLED": "1" if sms_enabled else "0",
        "SMS_CA_BUNDLE": ca_bundle or (cfg.get("SMS_CA_BUNDLE") or ""),
        "SMS_ALLOW_INSECURE_SSL": "1" if allow_insecure_ssl else "0",
        "TEXTBEE_API_KEY": textbee_api_key or (cfg.get("TEXTBEE_API_KEY") or ""),
        "TEXTBEE_DEVICE_ID": textbee_device_id or (cfg.get("TEXTBEE_DEVICE_ID") or ""),
        "TEXTBEE_SIM_SUBSCRIPTION_ID": textbee_sim_subscription_id or (cfg.get("TEXTBEE_SIM_SUBSCRIPTION_ID") or ""),
        "TEXTBEE_BASE_URL": textbee_base_url or (cfg.get("TEXTBEE_BASE_URL") or "https://api.textbee.dev/api/v1"),
    }
    save_deployment_config(current_app.instance_path, updates)
    flash("SMS gateway settings saved.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/order-cutoff-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_order_cutoff_settings():
    qr_cutoff = _format_cutoff_value(request.form.get("qr_order_cutoff_time"))
    staff_cutoff = _format_cutoff_value(request.form.get("staff_order_cutoff_time"))
    save_deployment_config(
        current_app.instance_path,
        {
            "QR_ORDER_CUTOFF_TIME": qr_cutoff,
            "STAFF_ORDER_CUTOFF_TIME": staff_cutoff,
        },
    )
    flash("Order cutoff settings saved.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/tax-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_tax_settings():
    try:
        service_charge_rate = round(float(request.form.get("service_charge_rate") or 0), 2)
    except (TypeError, ValueError):
        flash("Service charge must be a valid number.", "error")
        return redirect(url_for("cafe.home"))
    if service_charge_rate < 5 or service_charge_rate > 10:
        flash("Default service charge must be between 5% and 10%.", "error")
        return redirect(url_for("cafe.home"))
    save_deployment_config(
        current_app.instance_path,
        {
            "SERVICE_CHARGE_RATE": f"{service_charge_rate:.2f}",
        },
    )
    flash("Service charge setting saved.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/attendance-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_attendance_settings():
    raw_lat = (request.form.get("attendance_cafe_lat") or "").strip()
    raw_lng = (request.form.get("attendance_cafe_lng") or "").strip()
    raw_radius = (request.form.get("attendance_radius_m") or "").strip()
    try:
        cafe_lat = float(raw_lat)
        cafe_lng = float(raw_lng)
        radius_m = float(raw_radius)
    except (TypeError, ValueError):
        flash("Attendance geofence settings must be valid numbers.", "error")
        return redirect(url_for("cafe.home"))
    if radius_m < 20 or radius_m > 10000:
        flash("Attendance geofence radius must be between 20 and 10000 meters.", "error")
        return redirect(url_for("cafe.home"))
    save_deployment_config(
        current_app.instance_path,
        {
            "ATTENDANCE_CAFE_LAT": str(cafe_lat),
            "ATTENDANCE_CAFE_LNG": str(cafe_lng),
            "ATTENDANCE_RADIUS_METERS": str(radius_m),
        },
    )
    flash("Attendance geofence settings saved.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/attendance-checkin-qr.png")
@roles_required("admin", "manager")
def staff_attendance_qr_png():
    public_base = (current_app.config.get("PUBLIC_BASE_URL") or request.host_url.rstrip("/")).rstrip("/")
    qr_url = f"{public_base}{url_for('main.staff_attendance_check_in')}"
    qr = qrcode.QRCode(border=3, box_size=10)
    qr.add_data(qr_url)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return Response(
        output.getvalue(),
        headers={
            "Content-Type": "image/png",
            "Content-Disposition": 'inline; filename="brownberries-staff-attendance-qr.png"',
        },
    )


@bp.route("/kiosk-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_kiosk_settings():
    raw_token = (request.form.get("kiosk_token") or "").strip()
    if not raw_token:
        raw_token = secrets.token_urlsafe(18)
    save_deployment_config(
        current_app.instance_path,
        {
            "KDS_KIOSK_TOKEN": raw_token,
        },
    )
    flash("Prep display kiosk URLs updated.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/reception-kiosk-settings", methods=["POST"])
@roles_required("admin", "manager")
def update_reception_kiosk_settings():
    raw_token = (request.form.get("reception_kiosk_token") or "").strip()
    if not raw_token:
        raw_token = secrets.token_urlsafe(18)
    save_deployment_config(
        current_app.instance_path,
        {
            "RECEPTION_KIOSK_TOKEN": raw_token,
        },
    )
    flash("Reception kiosk URL updated.", "success")
    return redirect(url_for("cafe.home"))


@bp.route("/sms-settings/test", methods=["POST"])
@roles_required("admin", "manager")
def test_sms_settings():
    cc = (request.form.get("test_country_code") or "+91").strip()
    mobile = "".join(ch for ch in (request.form.get("test_mobile") or "") if ch.isdigit())
    message = (request.form.get("test_message") or "").strip() or "Brownberries Cafe SMS test message."
    if not mobile:
        flash("Please enter a valid test mobile number.", "error")
        return redirect(url_for("cafe.home"))
    ok, msg = _send_receipt_sms(cc, mobile, message)
    if ok:
        flash("Test SMS sent successfully.", "success")
    else:
        flash(f"Test SMS failed: {msg}", "error")
    return redirect(url_for("cafe.home"))


@bp.route("/bookings", methods=["GET", "POST"])
@roles_required("admin", "manager", "cashier")
def bookings():
    if request.method == "POST":
        booking = TableBooking.query.get_or_404(int(request.form["booking_id"]))
        new_status = (request.form.get("status") or "booked").strip().lower()
        if new_status not in ["booked", "confirmed", "arrived", "completed", "cancelled", "no_show"]:
            new_status = "booked"
        booking.status = new_status
        db.session.commit()
        flash("Booking status updated.", "success")
        return redirect(url_for("cafe.bookings", status=request.args.get("status", "")))

    status = (request.args.get("status") or "").strip().lower()
    bookings_query = TableBooking.query
    if status:
        bookings_query = bookings_query.filter(TableBooking.status == status)
    bookings_list = (
        bookings_query.order_by(TableBooking.booking_date.asc(), TableBooking.start_hour.asc(), TableBooking.created_at.asc())
        .all()
    )
    return render_template("cafe/bookings.html", bookings=bookings_list, selected_status=status)


@bp.route("/deliveries")
@roles_required("admin", "manager", "delivery_partner")
def delivery_locations():
    cafe_lat = 25.207989477704068
    cafe_lng = 80.87374457551877
    orders = (
        CafeOrder.query.options(joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item))
        .filter(CafeOrder.is_delivery.is_(True), CafeOrder.status.notin_(["paid", "cancelled"]))
        .order_by(CafeOrder.created_at.asc())
        .all()
    )
    return render_template(
        "cafe/deliveries.html",
        orders=orders,
        cafe_lat=cafe_lat,
        cafe_lng=cafe_lng,
    )


@bp.route("/tables", methods=["GET", "POST"])
@roles_required("admin", "manager")
def tables():
    if request.method == "POST":
        table = CafeTable(
            name=request.form["name"].strip(),
            seating_capacity=int(request.form["seating_capacity"]),
            metadata_note=request.form.get("metadata_note", "").strip() or None,
            qr_slug=f"{request.form['name'].strip().lower().replace(' ', '-')}-{uuid4().hex[:6]}",
            active=True if request.form.get("active") else False,
        )
        db.session.add(table)
        db.session.commit()
        flash("Table added.", "success")
        return redirect(url_for("cafe.tables"))
    return render_template("cafe/tables.html", tables=CafeTable.query.order_by(CafeTable.name).all())


@bp.route("/tables/<int:table_id>/delete", methods=["POST"])
@roles_required("admin", "manager")
def delete_table(table_id):
    table = CafeTable.query.get_or_404(table_id)
    db.session.delete(table)
    db.session.commit()
    flash("Table deleted.", "success")
    return redirect(url_for("cafe.tables"))


@bp.route("/tables/<int:table_id>/update", methods=["POST"])
@roles_required("admin", "manager")
def update_table(table_id):
    table = CafeTable.query.get_or_404(table_id)
    table.name = request.form["name"].strip()
    table.seating_capacity = int(request.form["seating_capacity"])
    table.metadata_note = request.form.get("metadata_note", "").strip() or None
    table.active = True if request.form.get("active") else False
    db.session.commit()
    flash("Table updated.", "success")
    return redirect(url_for("cafe.tables"))


@bp.route("/tables/<int:table_id>/qr")
@roles_required("admin", "manager")
def table_qr_download(table_id):
    table = CafeTable.query.get_or_404(table_id)
    base_url = (
        request.args.get("base_url")
        or current_app.config.get("PUBLIC_BASE_URL")
        or request.host_url.rstrip("/")
    ).strip()
    qr_url = f"{base_url}/table?slug={table.qr_slug}"

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=20,
        border=4,
    )
    qr.add_data(qr_url)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    output = BytesIO()
    image.save(output, format="PNG", optimize=True)
    output.seek(0)
    safe_table_name = table.name.lower().replace(" ", "-")
    filename = f"{safe_table_name}-qr.png"
    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "image/png",
        },
    )


@bp.route("/menu", methods=["GET", "POST"])
@roles_required("admin", "manager")
def menu():
    _ensure_protected_menu_categories()
    _ensure_menu_types_seeded()
    _ensure_workstations_seeded()
    stale_count = MenuItem.query.filter(MenuItem.subcategory_id.isnot(None)).count()
    if stale_count:
        MenuItem.query.filter(MenuItem.subcategory_id.isnot(None)).update(
            {MenuItem.subcategory_id: None}, synchronize_session=False
        )
        db.session.commit()
    if request.method == "POST":
        form_state = _menu_form_state_from_request()
        menu_type_id_raw = request.form.get("menu_type_id", "").strip()
        if not menu_type_id_raw.isdigit():
            flash("Please select a valid item type.", "error")
            return _render_menu_page("add_item", form_state)
        menu_type = MenuType.query.get(int(menu_type_id_raw))
        if not menu_type:
            flash("Please select a valid item type.", "error")
            return _render_menu_page("add_item", form_state)
        category_ids = _parse_category_ids_from_form()
        if not category_ids:
            flash("Please select at least one category.", "error")
            return _render_menu_page("add_item", form_state)
        chef_user = None
        chef_user_id_raw = request.form.get("chef_user_id", "").strip()
        if chef_user_id_raw:
            if not chef_user_id_raw.isdigit():
                flash("Please select a valid preparation responsibility.", "error")
                return _render_menu_page("add_item", form_state)
            chef_user = User.query.get(int(chef_user_id_raw))
            if not chef_user or not chef_user.active or not _is_preparation_responsibility_user(chef_user):
                flash("Please select a valid preparation responsibility.", "error")
                return _render_menu_page("add_item", form_state)
        name = request.form.get("name", "").strip()
        if not name:
            flash("Please enter the item name.", "error")
            return _render_menu_page("add_item", form_state)
        price_raw = request.form.get("price", "").strip()
        if not price_raw:
            flash("Please enter the item price.", "error")
            return _render_menu_page("add_item", form_state)
        try:
            price = float(price_raw)
        except ValueError:
            flash("Please enter a valid item price.", "error")
            return _render_menu_page("add_item", form_state)
        uploaded_image = _save_menu_image(request.files.get("image_file"))
        image_url = uploaded_image or (request.form.get("image_url", "").strip() or None)
        item = MenuItem(
            category_id=category_ids[0],
            subcategory_id=None,
            item_type=menu_type.name,
            category_ids_json=json.dumps(category_ids),
            name=name,
            image_url=image_url,
            short_description=request.form.get("short_description", "").strip() or None,
            description=request.form.get("description", "").strip() or None,
            calories=int(request.form["calories"]) if request.form.get("calories") else None,
            price=price,
            has_size_variants=True if request.form.get("has_size_variants") else False,
            size_pricing_json=None,
            prep_station=_normalize_prep_station(request.form.get("prep_station")),
            chef_user_id=chef_user.id if chef_user else None,
            available=True if request.form.get("available") else False,
            is_deleted=False,
        )
        if item.has_size_variants:
            size_pairs = _parse_size_pricing_from_form()
            if not size_pairs:
                flash("Please add at least one serving size with price.", "error")
                return _render_menu_page("add_item", form_state)
            item.size_pricing_json = json.dumps(size_pairs) if size_pairs else None
        db.session.add(item)
        db.session.commit()
        flash("Menu item added.", "success")
        return redirect(url_for("cafe.menu", section="items"))
    active_menu_section = (request.args.get("section") or "catalog").strip().lower()
    return _render_menu_page(active_menu_section)


@bp.route("/menu/availability", methods=["POST"])
@login_required
def update_menu_availability():
    category_filter = request.form.get("category_filter", "").strip()
    category_filter_id = int(category_filter) if category_filter.isdigit() else None
    scoped_items = _apply_category_filter(MenuItem.query.filter(MenuItem.is_deleted.is_(False)), category_filter_id).all()
    selected_ids = {
        int(x) for x in request.form.getlist("available_item_ids") if str(x).isdigit()
    }
    for item in scoped_items:
        item.available = item.id in selected_ids
    db.session.commit()
    flash("Menu item availability updated.", "success")
    next_url = request.form.get("next", "").strip()
    if next_url:
        return redirect(next_url)
    return redirect(url_for("cafe.items_availability", category_filter=category_filter or ""))


@bp.route("/items-availability")
@login_required
def items_availability():
    selected_category_filter = request.args.get("category_filter", type=int)
    categories = MenuCategory.query.order_by(MenuCategory.name.asc()).all()
    availability_items = _apply_category_filter(
        MenuItem.query.filter(MenuItem.is_deleted.is_(False)).order_by(MenuItem.name.asc()),
        selected_category_filter,
    ).all()
    return render_template(
        "cafe/items_availability.html",
        categories=categories,
        availability_items=availability_items,
        selected_category_filter=selected_category_filter,
    )


@bp.route("/menu/categories", methods=["POST"])
@roles_required("admin", "manager")
def add_category():
    name = request.form["name"].strip()
    if not name:
        flash("Category name is required.", "error")
        return redirect(url_for("cafe.menu"))
    if name.lower() in PROTECTED_MENU_CATEGORY_NAMES:
        flash("This category already exists as a protected system category.", "error")
        return redirect(url_for("cafe.menu"))
    existing = MenuCategory.query.filter(db.func.lower(MenuCategory.name) == name.lower()).first()
    if existing:
        flash("Category already exists.", "error")
        return redirect(url_for("cafe.menu"))
    db.session.add(MenuCategory(name=name))
    db.session.commit()
    flash("Category added.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/categories/<int:category_id>/update", methods=["POST"])
@roles_required("admin", "manager")
def update_category(category_id):
    category = MenuCategory.query.get_or_404(category_id)
    if _is_protected_menu_category(category):
        flash("Protected categories cannot be renamed.", "error")
        return redirect(url_for("cafe.menu"))
    name = request.form["name"].strip()
    if not name:
        flash("Category name is required.", "error")
        return redirect(url_for("cafe.menu"))
    if name.lower() in PROTECTED_MENU_CATEGORY_NAMES:
        flash("This category name is reserved for a protected system category.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate = MenuCategory.query.filter(
        db.func.lower(MenuCategory.name) == name.lower(), MenuCategory.id != category.id
    ).first()
    if duplicate:
        flash("Category with same name already exists.", "error")
        return redirect(url_for("cafe.menu"))
    category.name = name
    db.session.commit()
    flash("Category updated.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/categories/<int:category_id>/delete", methods=["POST"])
@roles_required("admin", "manager")
def delete_category(category_id):
    category = MenuCategory.query.get_or_404(category_id)
    other = _get_or_create_other_category()
    if _is_protected_menu_category(category) or category.id == other.id:
        flash("Protected categories cannot be deleted.", "error")
        return redirect(url_for("cafe.menu"))
    linked_items = MenuItem.query.filter(
        db.or_(
            MenuItem.category_id == category.id,
            MenuItem.category_ids_json.ilike(f"%{category.id}%"),
        )
    ).all()
    for item in linked_items:
        cat_ids = []
        if item.category_ids_json:
            try:
                raw = json.loads(item.category_ids_json)
                if isinstance(raw, list):
                    cat_ids = [int(x) for x in raw if str(x).isdigit()]
            except (TypeError, ValueError, json.JSONDecodeError):
                cat_ids = []
        if not cat_ids and item.category_id:
            cat_ids = [item.category_id]
        cat_ids = [cid for cid in cat_ids if cid != category.id]
        if not cat_ids:
            cat_ids = [other.id]
        item.category_id = cat_ids[0]
        item.category_ids_json = json.dumps(cat_ids)
    if category.subcategories:
        for sub in category.subcategories:
            db.session.delete(sub)
    db.session.delete(category)
    db.session.commit()
    flash("Category deleted. Linked items moved to Other.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/types", methods=["POST"])
@roles_required("admin", "manager")
def add_type():
    name = request.form["name"].strip()
    if not name:
        flash("Type name is required.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate = MenuType.query.filter(db.func.lower(MenuType.name) == name.lower()).first()
    if duplicate:
        flash("Type already exists.", "error")
        return redirect(url_for("cafe.menu"))
    db.session.add(MenuType(name=name))
    db.session.commit()
    flash("Type added.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/types/<int:type_id>/update", methods=["POST"])
@roles_required("admin", "manager")
def update_type(type_id):
    menu_type = MenuType.query.get_or_404(type_id)
    old_name = menu_type.name
    new_name = request.form["name"].strip()
    if not new_name:
        flash("Type name is required.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate = MenuType.query.filter(
        db.func.lower(MenuType.name) == new_name.lower(), MenuType.id != menu_type.id
    ).first()
    if duplicate:
        flash("Type already exists.", "error")
        return redirect(url_for("cafe.menu"))
    menu_type.name = new_name
    MenuItem.query.filter_by(item_type=old_name).update({"item_type": new_name})
    db.session.commit()
    flash("Type updated.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/types/<int:type_id>/delete", methods=["POST"])
@roles_required("admin", "manager")
def delete_type(type_id):
    menu_type = MenuType.query.get_or_404(type_id)
    if MenuItem.query.filter_by(item_type=menu_type.name).first():
        flash("Type cannot be deleted while linked menu items exist.", "error")
        return redirect(url_for("cafe.menu"))
    db.session.delete(menu_type)
    db.session.commit()
    flash("Type deleted.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/workstations", methods=["POST"])
@roles_required("admin", "manager")
def add_workstation():
    _ensure_workstations_seeded()
    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Workstation name is required.", "error")
        return redirect(url_for("cafe.menu"))
    slug = _slugify_workstation(request.form.get("slug") or name)
    if not slug:
        flash("Please enter a valid workstation name.", "error")
        return redirect(url_for("cafe.menu"))
    if Workstation.query.filter(db.func.lower(Workstation.slug) == slug.lower()).first():
        flash("Workstation slug already exists.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate_name = Workstation.query.filter(db.func.lower(Workstation.name) == name.lower()).first()
    if duplicate_name:
        flash("Workstation name already exists.", "error")
        return redirect(url_for("cafe.menu"))
    last_order = db.session.query(db.func.max(Workstation.display_order)).scalar() or 0
    db.session.add(
        Workstation(
            slug=slug,
            name=name,
            active=True,
            display_order=int(last_order) + 1,
        )
    )
    db.session.commit()
    flash("Workstation added.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/workstations/<int:workstation_id>/update", methods=["POST"])
@roles_required("admin", "manager")
def update_workstation(workstation_id):
    _ensure_workstations_seeded()
    workstation = Workstation.query.get_or_404(workstation_id)
    name = (request.form.get("name") or "").strip()
    slug = _slugify_workstation(request.form.get("slug") or workstation.slug)
    if not name or not slug:
        flash("Workstation name and slug are required.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate_slug = Workstation.query.filter(
        db.func.lower(Workstation.slug) == slug.lower(),
        Workstation.id != workstation.id,
    ).first()
    if duplicate_slug:
        flash("Another workstation already uses that slug.", "error")
        return redirect(url_for("cafe.menu"))
    duplicate_name = Workstation.query.filter(
        db.func.lower(Workstation.name) == name.lower(),
        Workstation.id != workstation.id,
    ).first()
    if duplicate_name:
        flash("Another workstation already uses that name.", "error")
        return redirect(url_for("cafe.menu"))
    old_slug = workstation.slug
    workstation.name = name
    workstation.slug = slug
    workstation.active = True if request.form.get("active") else False
    if old_slug != slug:
        MenuItem.query.filter_by(prep_station=old_slug).update({"prep_station": slug}, synchronize_session=False)
    db.session.commit()
    flash("Workstation updated.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/workstations/<int:workstation_id>/delete", methods=["POST"])
@roles_required("admin", "manager")
def delete_workstation(workstation_id):
    _ensure_workstations_seeded()
    workstation = Workstation.query.get_or_404(workstation_id)
    MenuItem.query.filter_by(prep_station=workstation.slug).update({"prep_station": ""}, synchronize_session=False)
    db.session.delete(workstation)
    db.session.commit()
    flash("Workstation deleted. Linked menu items are now unassigned.", "success")
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/<int:item_id>/availability", methods=["POST"])
@roles_required("admin", "manager")
def toggle_item(item_id):
    item = MenuItem.query.get_or_404(item_id)
    item.available = not item.available
    db.session.commit()
    return redirect(url_for("cafe.menu"))


@bp.route("/menu/items/<int:item_id>/update", methods=["POST"])
@roles_required("admin", "manager")
def update_menu_item(item_id):
    item = MenuItem.query.get_or_404(item_id)
    menu_type_id = request.form.get("menu_type_id", "").strip()
    if menu_type_id:
        menu_type = MenuType.query.get(int(menu_type_id))
        if not menu_type:
            flash("Please select a valid type.", "error")
            return redirect(url_for("cafe.menu", section="items"))
        item.item_type = menu_type.name

    category_ids = _parse_category_ids_from_form()
    if category_ids:
        item.category_id = category_ids[0]
        item.category_ids_json = json.dumps(category_ids)
    item.subcategory_id = None
    item_name = request.form.get("name", "").strip()
    if item_name:
        item.name = item_name

    uploaded_image = _save_menu_image(request.files.get("image_file"))
    if uploaded_image:
        item.image_url = uploaded_image
    else:
        image_url = request.form.get("image_url", "").strip()
        if image_url:
            item.image_url = image_url

    if "description" in request.form:
        description = request.form.get("description", "").strip()
        if description:
            item.description = description
    if "short_description" in request.form:
        short_description = request.form.get("short_description", "").strip()
        if short_description:
            item.short_description = short_description

    calories_raw = request.form.get("calories", "").strip()
    if calories_raw:
        item.calories = int(calories_raw)

    price_raw = request.form.get("price", "").strip()
    if price_raw:
        item.price = float(price_raw)

    has_size_variants = True if request.form.get("has_size_variants") else False
    if has_size_variants:
        item.has_size_variants = True
        size_pairs = _parse_size_pricing_from_form()
        if size_pairs:
            item.size_pricing_json = json.dumps(size_pairs)
    else:
        item.has_size_variants = False
        item.size_pricing_json = None

    if "prep_station" in request.form:
        item.prep_station = _normalize_prep_station(request.form.get("prep_station"))
    if "chef_user_id" in request.form:
        chef_user_id_raw = request.form.get("chef_user_id", "").strip()
        if not chef_user_id_raw:
            item.chef_user_id = None
        elif chef_user_id_raw.isdigit():
            chef_user = User.query.get(int(chef_user_id_raw))
            if chef_user and chef_user.active and _is_preparation_responsibility_user(chef_user):
                item.chef_user_id = chef_user.id
            else:
                flash("Please select a valid preparation responsibility.", "error")
                return redirect(url_for("cafe.menu", section="items"))
        else:
            flash("Please select a valid preparation responsibility.", "error")
            return redirect(url_for("cafe.menu", section="items"))

    item.available = True if request.form.get("available") else False
    db.session.commit()
    flash("Menu item updated.", "success")
    return redirect(url_for("cafe.menu", section="items"))


@bp.route("/menu/items/<int:item_id>/delete", methods=["POST"])
@roles_required("admin", "manager")
def delete_menu_item(item_id):
    item = MenuItem.query.get_or_404(item_id)
    item.is_deleted = True
    item.available = False
    db.session.commit()
    flash("Menu item moved to Deleted Items. You can restore it later.", "success")
    return redirect(url_for("cafe.menu", section="deleted_items"))


@bp.route("/menu/items/<int:item_id>/restore", methods=["POST"])
@roles_required("admin", "manager")
def restore_menu_item(item_id):
    item = MenuItem.query.get_or_404(item_id)
    item.is_deleted = False
    db.session.commit()
    flash("Menu item restored.", "success")
    return redirect(url_for("cafe.menu", section="deleted_items"))


@bp.route("/orders", methods=["GET", "POST"])
@login_required
def orders():
    return _render_orders_view(kiosk_mode=False)


def _render_orders_view(kiosk_mode: bool = False, access_key: str = ""):
    if kiosk_mode and not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    if request.method == "POST":
        selected_table_id = int(request.form["table_id"])
        cutoff_message = _order_cutoff_message("staff")
        if cutoff_message:
            flash(cutoff_message, "error")
            if kiosk_mode:
                return redirect(url_for("cafe.reception_kiosk_orders", access_key=access_key, table_id=selected_table_id))
            return redirect(url_for("cafe.orders", table_id=selected_table_id))
        order = _create_order(
            table_id=selected_table_id,
            ordered_by_user_id=g.current_user.id if getattr(g, "current_user", None) else _get_qr_guest_user_id(),
            status="pending_approval",
            payment_type=None,
            payment_reference=None,
        )
        if not order:
            flash("Please add at least one menu item in cart.", "error")
            if kiosk_mode:
                return redirect(url_for("cafe.reception_kiosk_orders", access_key=access_key, table_id=selected_table_id))
            return redirect(url_for("cafe.orders", table_id=selected_table_id))
        flash("Order created.", "success")
        if kiosk_mode:
            return redirect(url_for("cafe.reception_kiosk_orders", access_key=access_key, table_id=selected_table_id))
        return redirect(url_for("cafe.orders", table_id=selected_table_id))

    table_id = request.args.get("table_id", type=int)
    category_id = request.args.get("category_id", type=int)
    item_type = (request.args.get("item_type") or "").strip()
    today_start, today_end = _current_ist_day_bounds()
    tables = CafeTable.query.filter_by(active=True).order_by(CafeTable.name).all()
    if not table_id and tables:
        table_id = tables[0].id

    menu_query = MenuItem.query.filter_by(available=True, is_deleted=False)
    menu_query = _apply_category_filter(menu_query, category_id)
    if item_type:
        menu_query = menu_query.filter(MenuItem.item_type == item_type)
    filtered_items = menu_query.all()

    frequency_rows = (
        db.session.query(
            CafeOrderItem.menu_item_id,
            db.func.coalesce(db.func.sum(CafeOrderItem.quantity), 0).label("order_qty"),
        )
        .group_by(CafeOrderItem.menu_item_id)
        .all()
    )
    item_frequency = {row.menu_item_id: int(row.order_qty) for row in frequency_rows}
    menu_items = sorted(
        filtered_items,
        key=lambda item: (-item_frequency.get(item.id, 0), item.name.lower()),
    )
    item_size_map = {}
    for item in menu_items:
        sizes = []
        if item.size_pricing_json:
            try:
                raw = json.loads(item.size_pricing_json)
                if isinstance(raw, list):
                    for row in raw:
                        if isinstance(row, dict) and row.get("size") and row.get("price") is not None:
                            sizes.append({"size": str(row["size"]), "price": float(row["price"])})
            except (TypeError, ValueError, json.JSONDecodeError):
                sizes = []
        item_size_map[item.id] = sizes

    categories = _visible_categories_for_available_menu(include_protected=True)
    category_name_by_id = {c.id: c.name for c in categories}
    item_category_names_map = {
        item.id: _get_item_category_names(item, category_name_by_id) for item in menu_items
    }
    item_types = [
        row[0]
        for row in db.session.query(MenuItem.item_type)
        .filter(MenuItem.available.is_(True))
        .distinct()
        .order_by(MenuItem.item_type.asc())
        .all()
    ]

    return render_template(
        "cafe/orders.html",
        tables=tables,
        menu_items=menu_items,
        orders=(
            CafeOrder.query.filter(CafeOrder.created_at >= today_start, CafeOrder.created_at <= today_end)
            .order_by(CafeOrder.created_at.desc())
            .limit(50)
            .all()
        ),
        categories=categories,
        item_types=item_types,
        selected_category_id=category_id,
        selected_item_type=item_type,
        item_frequency=item_frequency,
        item_size_map=item_size_map,
        item_category_names_map=item_category_names_map,
        selected_table_id=table_id,
        table_orders=CafeOrder.query.filter(
            CafeOrder.table_id == table_id,
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
        )
        .order_by(CafeOrder.created_at.desc())
        .all()
        if table_id
        else [],
        hide_staff_nav=kiosk_mode,
        topbar_home_url=(
            url_for("cafe.reception_kiosk", access_key=access_key)
            if kiosk_mode
            else url_for("main.dashboard")
        ),
        manifest_url=(
            url_for("cafe.reception_kiosk_manifest", access_key=access_key)
            if kiosk_mode
            else url_for("static", filename="manifest.webmanifest")
        ),
        web_app_title="Brownberries Reception" if kiosk_mode else "Brownberries Café",
        kiosk_mode=kiosk_mode,
        kiosk_access_key=access_key,
    )


@bp.route("/reception/<string:access_key>/orders", methods=["GET", "POST"])
def reception_kiosk_orders(access_key):
    return _render_orders_view(kiosk_mode=True, access_key=access_key)


@bp.route("/orders/<int:order_id>/mark-paid", methods=["POST"])
@roles_required("admin", "manager", "cashier")
def mark_order_paid(order_id):
    order = CafeOrder.query.get(order_id)
    if not order:
        flash("Order not found. It may have already been updated on another screen.", "error")
        return redirect(request.form.get("next") or url_for("cafe.cashier"))
    payment_type = request.form.get("payment_type", "").strip() or order.payment_type or "Cash"
    payment_reference = request.form.get("payment_reference", "").strip() or order.payment_reference
    _apply_order_tax_breakdown(order, flags=_selected_tax_flags())
    order.status = "paid"
    order.paid_at = datetime.utcnow()
    order.payment_type = payment_type
    order.payment_reference = payment_reference
    sms_message = ""
    if request.form.get("send_receipt_sms"):
        cc = (request.form.get("country_code") or "+91").strip()
        mobile = "".join(ch for ch in (request.form.get("mobile") or "") if ch.isdigit())
        if mobile:
            msg = _receipt_sms_message(order)
            ok, sms_resp = _send_receipt_sms(cc, mobile, msg)
            sms_message = " SMS sent." if ok else f" SMS not sent: {sms_resp}"
        else:
            sms_message = " SMS not sent: mobile missing."
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    flash(f"Order #{_format_pickup_number(order)} marked as paid.{sms_message}", "success")
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    if next_url:
        return redirect(next_url)
    return redirect(url_for("cafe.cashier", table_id=order.table_id))


@bp.route("/orders/<int:order_id>/approve", methods=["POST"])
@roles_required("admin", "manager", "cashier", "librarian", "staff")
def approve_order(order_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _approve_order_impl(order_id, next_url)


def _approve_order_impl(order_id: int, next_url: str = ""):
    order = CafeOrder.query.get(order_id)
    if not order:
        flash("Order not found. It may have already been updated on another screen.", "error")
        return redirect(next_url or url_for("cafe.cashier"))
    decision = (request.form.get("decision") or "approve").strip().lower()
    if decision not in ["approve", "reject"]:
        decision = "approve"
    pending_items = [oi for oi in order.order_items if (oi.approval_status or "pending") == "pending"]
    if not pending_items:
        flash("No pending items left for action.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=order.table_id))
    selected_ids = {int(v) for v in request.form.getlist("pending_item_ids") if str(v).isdigit()}
    if not selected_ids:
        flash("Select at least one pending item.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=order.table_id))
    approved_names = []
    rejected_names = []
    for oi in pending_items:
        if oi.id not in selected_ids:
            continue
        if decision == "approve":
            oi.approval_status = "approved"
            oi.prep_status = "pending"
            approved_names.append(f"{oi.menu_item.name if oi.menu_item else 'Item'} x {oi.quantity}")
        else:
            oi.approval_status = "rejected"
            rejected_names.append(f"{oi.menu_item.name if oi.menu_item else 'Item'} x {oi.quantity}")

    _recalculate_order_totals(order)
    _refresh_order_status_from_items(order)
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    _emit_ops_notification(
        f"Order #{_format_pickup_number(order)} decision updated",
        kind="order_decision",
        payload={
            "order_id": order.display_code or _format_internal_order_code(order),
            "pickup_no": _format_pickup_number(order),
            "table_id": order.table_id,
            "approved_items": approved_names,
            "rejected_items": rejected_names,
        },
    )
    flash(f"Order #{_format_pickup_number(order)}: selected items {decision}d.", "success")
    return redirect(next_url or url_for("cafe.cashier", table_id=order.table_id))


@bp.route("/tables/<int:table_id>/clear-orders", methods=["POST"])
@roles_required("admin", "manager", "cashier")
def clear_table_orders(table_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _clear_table_orders_impl(table_id, next_url)


def _clear_table_orders_impl(table_id: int, next_url: str = ""):
    table = CafeTable.query.get_or_404(table_id)
    day_start, day_end = _current_ist_day_bounds()
    split_rows, split_error = _parse_split_payment_rows()
    if split_error:
        flash(split_error, "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=table.id, tab="running"))
    selected_order_ids = {
        int(value) for value in request.form.getlist("order_ids") if str(value).isdigit()
    }
    if not selected_order_ids:
        flash("Select at least one order to settle.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=table.id, tab="running"))
    orders = CafeOrder.query.filter(
        CafeOrder.table_id == table.id,
        CafeOrder.status.notin_(["paid", "cancelled"]),
        CafeOrder.created_at >= day_start,
        CafeOrder.created_at < day_end,
    ).all()
    orders = [order for order in orders if order.id in selected_order_ids]
    payable_orders = [
        order for order in orders
        if not any((oi.approval_status or "pending") == "pending" for oi in order.order_items)
    ]
    if not payable_orders:
        flash("Select at least one approved order to settle.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=table.id, tab="running"))
    tax_flags = _selected_tax_flags()
    selected_total = 0.0
    for order in payable_orders:
        breakdown = _order_tax_breakdown(order, flags=tax_flags)
        selected_total += breakdown["grand_total"]
    selected_total = round(selected_total, 2)
    split_total = round(sum(float(row["amount"]) for row in split_rows), 2)
    if abs(split_total - selected_total) > 0.01:
        flash(f"Payment split total ₹{split_total:.2f} must match selected orders total ₹{selected_total:.2f}.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=table.id, tab="running"))
    count = 0
    paid_now = datetime.utcnow()
    summary_label = split_rows[0]["method"] if len(split_rows) == 1 else "Split Payment"
    summary_ref = ", ".join(
        [f'{row["method"]}: ₹{row["amount"]:.2f}' + (f' ({row["reference"]})' if row["reference"] else "") for row in split_rows]
    )[:120] or None
    payment_breakdown_json = json.dumps(split_rows)
    for order in payable_orders:
        _apply_order_tax_breakdown(order, flags=tax_flags)
        order.status = "paid"
        order.paid_at = paid_now
        order.payment_type = summary_label
        order.payment_reference = summary_ref or order.payment_reference
        order.payment_breakdown_json = payment_breakdown_json
        payload = _serialize_order(order)
        socketio.emit("order_updated", payload, namespace="/kitchen")
        socketio.emit("order_updated", payload, namespace="/table")
        count += 1
    table.service_charge_opt_out_requested = False
    db.session.commit()
    sms_message = ""
    if request.form.get("send_receipt_sms"):
        cc = (request.form.get("receipt_country_code") or "+91").strip()
        mobile = "".join(ch for ch in (request.form.get("receipt_mobile") or "") if ch.isdigit())
        if mobile:
            msg = _receipt_sms_message(payable_orders[0])
            ok, sms_resp = _send_receipt_sms(cc, mobile, msg)
            sms_message = " Receipt SMS sent." if ok else f" Receipt SMS not sent: {sms_resp}"
        else:
            sms_message = " Receipt SMS not sent: mobile missing."
    flash(f"Settled {count} order(s) for {table.name}.{sms_message}", "success")
    if next_url:
        return redirect(next_url)
    return redirect(url_for("cafe.cashier", table_id=table.id))


@bp.route("/tables/<int:table_id>/move-orders", methods=["POST"])
@roles_required("admin", "manager", "cashier")
def move_table_orders(table_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _move_table_orders_impl(table_id, next_url)


def _move_table_orders_impl(table_id: int, next_url: str = ""):
    source_table = CafeTable.query.get_or_404(table_id)
    target_table_id = request.form.get("target_table_id", type=int)
    selected_order_ids = {
        int(value) for value in request.form.getlist("order_ids") if str(value).isdigit()
    }
    if not selected_order_ids:
        flash("Select at least one active order to move.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=source_table.id, tab="running"))
    if not target_table_id:
        flash("Select the destination table.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=source_table.id, tab="running"))
    if target_table_id == source_table.id:
        flash("Choose a different table to move the order(s).", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=source_table.id, tab="running"))
    target_table = CafeTable.query.filter_by(id=target_table_id, active=True).first()
    if not target_table:
        flash("Destination table was not found.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=source_table.id, tab="running"))

    today_start, today_end = _current_ist_day_bounds()
    orders = (
        CafeOrder.query.options(
            joinedload(CafeOrder.table),
            joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
        )
        .filter(
            CafeOrder.id.in_(selected_order_ids),
            CafeOrder.table_id == source_table.id,
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
        )
        .order_by(CafeOrder.created_at.asc())
        .all()
    )
    if not orders:
        flash("No movable active orders were found for the selected table.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=source_table.id, tab="running"))

    moved_refs = []
    for order in orders:
        order.table_id = target_table.id
        order.table = target_table
        moved_refs.append(order.display_code or _format_internal_order_code(order))
    db.session.commit()

    for order in orders:
        payload = _serialize_order(order)
        socketio.emit("order_updated", payload, namespace="/kitchen")
        socketio.emit("order_updated", payload, namespace="/table")

    _emit_ops_notification(
        f"Moved {len(orders)} order(s) from {source_table.name} to {target_table.name}",
        kind="order_move",
        payload={
            "source_table": source_table.name,
            "target_table": target_table.name,
            "table_id": target_table.id,
            "moved_orders": moved_refs,
        },
    )
    flash(
        f"Moved {len(orders)} active order(s) from {source_table.name} to {target_table.name}.",
        "success",
    )
    return redirect(next_url or url_for("cafe.cashier", table_id=target_table.id, tab="running"))


@bp.route("/cashier")
@roles_required("admin", "manager", "cashier")
def cashier():
    return _render_cashier_view(kiosk_mode=False)


def _render_cashier_view(kiosk_mode: bool = False, access_key: str = ""):
    tab = (request.args.get("tab") or "running").strip().lower()
    if tab not in ["running", "all_orders"]:
        tab = "running"
    table_id = request.args.get("table_id", type=int)
    today_ist = datetime.now(IST_TZ).date()
    sel_year = request.args.get("year", type=int) or today_ist.year
    sel_month = request.args.get("month", type=int) or today_ist.month
    sel_day = request.args.get("day", type=int) or today_ist.day
    try:
        selected_date = date(sel_year, sel_month, sel_day)
    except ValueError:
        selected_date = today_ist
        sel_year, sel_month, sel_day = selected_date.year, selected_date.month, selected_date.day
    day_start = _utc_naive_from_ist(datetime.combine(selected_date, time.min))
    day_end = _utc_naive_from_ist(datetime.combine(selected_date + timedelta(days=1), time.min))
    today_day_start = _utc_naive_from_ist(datetime.combine(today_ist, time.min))
    today_day_end = _utc_naive_from_ist(datetime.combine(today_ist + timedelta(days=1), time.min))
    tax_settings = _tax_settings()
    tables = CafeTable.query.filter_by(active=True).order_by(CafeTable.name).all()
    running_orders = (
        CafeOrder.query.options(
            joinedload(CafeOrder.table),
            joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
        )
        .filter(
            CafeOrder.status.notin_(["paid", "cancelled"]),
            CafeOrder.created_at >= day_start,
            CafeOrder.created_at < day_end,
        )
        .order_by(CafeOrder.created_at.asc())
        .all()
    )
    running_map = {}
    for order in running_orders:
        entry = running_map.setdefault(
            order.table_id,
            {
                "count": 0,
                "total": 0.0,
                "pending_approval_count": 0,
                "queued_count": 0,
                "preparing_count": 0,
                "ready_count": 0,
                "served_count": 0,
                "rejected_count": 0,
            },
        )
        entry["count"] += 1
        entry["total"] += float(order.total_amount or 0)
        for oi in order.order_items:
            approval = (oi.approval_status or "pending").strip().lower()
            if approval == "rejected":
                entry["rejected_count"] += int(oi.quantity or 0)
                continue
            if approval == "pending":
                entry["pending_approval_count"] += int(oi.quantity or 0)
                continue
            prep_status = _normalized_item_prep_status(oi)
            if prep_status == "preparing":
                entry["preparing_count"] += int(oi.quantity or 0)
            elif prep_status == "ready":
                entry["ready_count"] += int(oi.quantity or 0)
            elif prep_status == "served":
                entry["served_count"] += int(oi.quantity or 0)
            else:
                entry["queued_count"] += int(oi.quantity or 0)
    for running_table_id in list(running_map.keys()):
        running_map[running_table_id]["total"] = float(running_map[running_table_id]["total"])
    if not table_id and tables:
        running_table_ids = [t.id for t in tables if running_map.get(t.id, {}).get("count", 0) > 0]
        table_id = running_table_ids[0] if running_table_ids else tables[0].id
    selected_table = CafeTable.query.get(table_id) if table_id else None
    target_tables = [table for table in tables if not selected_table or table.id != selected_table.id]
    total_sale_today = (
        db.session.query(db.func.coalesce(db.func.sum(CafeOrder.total_amount), 0))
        .filter(
            CafeOrder.status == "paid",
            CafeOrder.paid_at.is_not(None),
            CafeOrder.paid_at >= today_day_start,
            CafeOrder.paid_at < today_day_end,
        )
        .scalar()
        or 0
    )
    unpaid_orders = []
    table_total = 0
    table_settle_total = 0
    payable_order_id_map = {}
    pending_order_item_map = {}
    latest_settled_order = None
    latest_settled_feedback = None
    if selected_table:
        unpaid_orders = (
            CafeOrder.query.options(
                joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
            ).filter(
                CafeOrder.table_id == selected_table.id,
                CafeOrder.status.notin_(["paid", "cancelled"]),
                CafeOrder.created_at >= day_start,
                CafeOrder.created_at < day_end,
            )
            .order_by(CafeOrder.created_at.asc())
            .all()
        )
        table_total = round(sum(o.total_amount for o in unpaid_orders), 2)
        payable_order_id_map = {
            o.id: (not any((oi.approval_status or "pending") == "pending" for oi in o.order_items))
            for o in unpaid_orders
        }
        pending_order_item_map = {
            o.id: [oi for oi in o.order_items if (oi.approval_status or "pending") == "pending"]
            for o in unpaid_orders
        }
        table_settle_total = round(
            sum(float(o.total_amount or 0) for o in unpaid_orders if payable_order_id_map.get(o.id)),
            2,
        )
        latest_settled_order = _latest_paid_order_for_table(selected_table.id)
        latest_settled_feedback = _feedback_for_settlement(latest_settled_order)
    running_order_time_map = {
        o.id: _format_ist(o.created_at, "%I:%M:%S %p")
        for o in unpaid_orders
    }
    all_orders = (
        CafeOrder.query.options(joinedload(CafeOrder.table), joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item))
        .filter(CafeOrder.created_at >= day_start, CafeOrder.created_at < day_end)
        .order_by(CafeOrder.created_at.desc())
        .all()
    )
    all_orders_payload = []
    public_base = (current_app.config.get("PUBLIC_BASE_URL") or request.host_url.rstrip("/")).rstrip("/")
    for o in all_orders:
        line_subtotal = round(
            sum(
                float(x.unit_price or 0) * int(x.quantity or 0)
                for x in o.order_items
                if (x.approval_status or "pending") != "rejected"
            ),
            2,
        )
        all_orders_payload.append(
            {
                "id": o.id,
                "code": o.display_code or _format_internal_order_code(o),
                "pickup_no": _format_pickup_number(o),
                "pickup_no_compact": _format_pickup_number(o, compact=True),
                "table": o.table.name if o.table else "-",
                "status": o.status,
                "payment_type": o.payment_type or "-",
                "payment_reference": o.payment_reference or "-",
                "payment_breakdown": (
                    json.loads(o.payment_breakdown_json)
                    if (o.payment_breakdown_json and str(o.payment_breakdown_json).strip().startswith("["))
                    else []
                ),
                "created_at": _format_ist(o.created_at, "%I:%M:%S %p"),
                "items": [
                    {
                        "name": (oi.menu_item.name if oi.menu_item else "Item"),
                        "qty": int(oi.quantity or 0),
                        "size_label": oi.size_label or "",
                        "is_parcel": bool(oi.is_parcel),
                        "unit_price": round(float(oi.unit_price or 0), 2),
                        "approval_status": (oi.approval_status or "pending"),
                        "prep_status": _normalized_item_prep_status(oi),
                    }
                    for oi in o.order_items
                ],
                "subtotal": line_subtotal,
                "packaging_charge": round(float(o.packaging_charge or 0), 2),
                "service_tax_amount": round(float(o.service_tax_amount or 0), 2),
                "gst_amount": 0.0,
                "cst_amount": 0.0,
                "tax_amount": 0.0,
                "total_amount": round(float(o.total_amount or 0), 2),
                "receipt_link": f"{public_base}/cafe/receipt/{o.id}",
            }
        )
    years = list(range(max(2024, date.today().year - 2), date.today().year + 3))
    current_cashier_url = request.url
    return render_template(
        "cafe/cashier.html",
        tab=tab,
        tables=tables,
        running_map=running_map,
        selected_table=selected_table,
        target_tables=target_tables,
        unpaid_orders=unpaid_orders,
        payable_order_id_map=payable_order_id_map,
        pending_order_item_map=pending_order_item_map,
        running_order_time_map=running_order_time_map,
        table_total=table_total,
        table_settle_total=table_settle_total,
        latest_settled_order=latest_settled_order,
        latest_settled_feedback=latest_settled_feedback,
        service_charge_rate=tax_settings["service_charge_rate"],
        total_sale_today=round(float(total_sale_today), 2),
        all_orders=all_orders,
        all_orders_payload=all_orders_payload,
        selected_date=selected_date,
        sel_year=sel_year,
        sel_month=sel_month,
        sel_day=sel_day,
        years=years,
        kiosk_mode=kiosk_mode,
        kiosk_access_key=access_key,
        current_cashier_url=current_cashier_url,
        hide_staff_nav=True,
        topbar_home_url=(
            url_for("cafe.reception_kiosk_orders", access_key=access_key)
            if kiosk_mode
            else url_for("main.dashboard")
        ),
        manifest_url=(
            url_for("cafe.reception_kiosk_manifest", access_key=access_key)
            if kiosk_mode
            else url_for("static", filename="manifest.webmanifest")
        ),
        web_app_title="Brownberries Reception" if kiosk_mode else "Brownberries Café",
    )


@bp.route("/reception/<string:access_key>")
def reception_kiosk(access_key):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    return _render_cashier_view(kiosk_mode=True, access_key=access_key)


@bp.route("/reception/<string:access_key>/manifest.webmanifest")
def reception_kiosk_manifest(access_key):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    return Response(
        json.dumps(_reception_manifest_payload(access_key)),
        mimetype="application/manifest+json",
    )


@bp.route("/receipt/<int:order_id>")
def public_receipt(order_id):
    order = CafeOrder.query.options(
        joinedload(CafeOrder.table),
        joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
    ).get_or_404(order_id)
    receipt_orders = _receipt_orders_for_settlement(order)
    primary_order = receipt_orders[0]
    payment_breakdown = _payment_breakdown_rows(primary_order)
    subtotal = round(
        sum(
            float(oi.unit_price or 0) * int(oi.quantity or 0)
            for settled_order in receipt_orders
            for oi in settled_order.order_items
            if (oi.approval_status or "pending") != "rejected"
        ),
        2,
    )
    packaging_charge = round(sum(float(settled_order.packaging_charge or 0) for settled_order in receipt_orders), 2)
    delivery_charge = round(sum(float(settled_order.delivery_charge or 0) for settled_order in receipt_orders), 2)
    service_tax_amount = round(sum(float(settled_order.service_tax_amount or 0) for settled_order in receipt_orders), 2)
    gst_amount = 0.0
    cst_amount = 0.0
    tax_amount = 0.0
    settlement_total = round(sum(float(settled_order.total_amount or 0) for settled_order in receipt_orders), 2)
    receipt_lines = []
    for settled_order in receipt_orders:
        pickup_no = _format_pickup_number(settled_order)
        for oi in settled_order.order_items:
            if (oi.approval_status or "pending") == "rejected":
                continue
            receipt_lines.append(
                {
                    "order_public_id": f"#{pickup_no}",
                    "order_internal_ref": settled_order.display_code or str(settled_order.id),
                    "item_name": oi.menu_item.name if oi.menu_item else "Item",
                    "size_label": oi.size_label or "",
                    "is_parcel": bool(oi.is_parcel),
                    "quantity": int(oi.quantity or 0),
                    "unit_price": round(float(oi.unit_price or 0), 2),
                    "line_total": round(float(oi.unit_price or 0) * int(oi.quantity or 0), 2),
                }
            )
    return render_template(
        "cafe/receipt_public.html",
        order=primary_order,
        receipt_orders=receipt_orders,
        receipt_lines=receipt_lines,
        payment_breakdown=payment_breakdown,
        subtotal=subtotal,
        packaging_charge=packaging_charge,
        delivery_charge=delivery_charge,
        service_tax_amount=service_tax_amount,
        gst_amount=gst_amount,
        cst_amount=cst_amount,
        tax_amount=tax_amount,
        settlement_total=settlement_total,
        paid_at_display=_format_ist(primary_order.paid_at, "%d %b %Y, %I:%M %p") if primary_order.paid_at else "",
        receipt_location=_receipt_location_text(),
        review_link=_google_review_link(),
        feedback_link=url_for("cafe.public_settlement_feedback", order_id=primary_order.id),
    )


def _render_feedback_form(
    order: CafeOrder,
    *,
    source: str,
    kiosk_mode: bool = False,
    kiosk_access_key: str = "",
):
    receipt_orders = _receipt_orders_for_settlement(order)
    primary_order = receipt_orders[0]
    feedback = _feedback_for_settlement(primary_order)
    item_rows = _feedback_line_items_for_orders(receipt_orders)
    existing_item_map = {}
    if feedback:
        existing_item_map = {
            (item.menu_item_id, item.size_label or "", bool(item.is_parcel)): item.rating
            for item in feedback.items
        }
    default_rating = 3 if source == "offline" else 5
    for row in item_rows:
        row["current_rating"] = existing_item_map.get(
            (row["menu_item_id"], row["size_label"] or "", bool(row["is_parcel"])),
            default_rating,
        )
    feedback_locked = bool(source == "offline" and feedback and feedback.source == "online")
    if request.method == "POST":
        if feedback_locked:
            flash("Customer feedback is already submitted for this settlement.", "error")
        else:
            saved_feedback, _ = _upsert_feedback_for_settlement(
                primary_order,
                source=source,
                item_rows=item_rows,
                default_rating=default_rating,
                submitted_by_user=g.current_user if source == "offline" else None,
                submitted_by_name=("Guest Feedback" if source == "online" else (g.current_user.full_name if g.current_user else "Cashier Feedback")),
            )
            db.session.commit()
            flash("Feedback saved. Thank you.", "success")
            feedback = saved_feedback
            feedback_locked = bool(source == "offline" and feedback.source == "online")
            existing_item_map = {
                (item.menu_item_id, item.size_label or "", bool(item.is_parcel)): item.rating
                for item in feedback.items
            }
            for row in item_rows:
                row["current_rating"] = existing_item_map.get(
                    (row["menu_item_id"], row["size_label"] or "", bool(row["is_parcel"])),
                    default_rating,
                )
    pickup_numbers = [f"#{_format_pickup_number(settled_order)}" for settled_order in receipt_orders]
    settlement_total = round(sum(float(settled_order.total_amount or 0) for settled_order in receipt_orders), 2)
    back_url = (request.args.get("next") or "").strip()
    if not back_url:
        if source == "online":
            back_url = url_for("cafe.public_receipt", order_id=primary_order.id)
        elif kiosk_mode:
            back_url = url_for("cafe.reception_kiosk", access_key=kiosk_access_key, table_id=primary_order.table_id, tab="running")
        else:
            back_url = url_for("cafe.cashier", table_id=primary_order.table_id, tab="running")
    return render_template(
        "cafe/feedback_form.html",
        order=primary_order,
        receipt_orders=receipt_orders,
        pickup_numbers=pickup_numbers,
        settlement_total=settlement_total,
        paid_at_display=_format_ist(primary_order.paid_at, "%d %b %Y, %I:%M %p") if primary_order.paid_at else "",
        item_rows=item_rows,
        feedback=feedback,
        existing_item_map=existing_item_map,
        default_rating=default_rating,
        feedback_locked=feedback_locked,
        source=source,
        back_url=back_url,
        review_link=_google_review_link(),
        receipt_location=_receipt_location_text(),
        hide_staff_nav=(source == "online" or kiosk_mode),
        topbar_home_url=(
            url_for("main.public_home")
            if source == "online"
            else (url_for("cafe.reception_kiosk_orders", access_key=kiosk_access_key) if kiosk_mode else url_for("main.dashboard"))
        ),
    )


@bp.route("/receipt/<int:order_id>/feedback", methods=["GET", "POST"])
def public_settlement_feedback(order_id):
    order = CafeOrder.query.options(
        joinedload(CafeOrder.table),
        joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
    ).get_or_404(order_id)
    if order.status != "paid":
        flash("Feedback is available after settlement.", "error")
        return redirect(url_for("main.table_qr_page", slug=order.table.qr_slug) if order.table else url_for("main.public_home"))
    return _render_feedback_form(order, source="online")


@bp.route("/orders/<int:order_id>/feedback", methods=["GET", "POST"])
@roles_required("admin", "manager", "cashier")
def settlement_feedback(order_id):
    order = CafeOrder.query.options(
        joinedload(CafeOrder.table),
        joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
    ).get_or_404(order_id)
    return _render_feedback_form(order, source="offline")


@bp.route("/orders/<int:order_id>/send-receipt-sms", methods=["POST"])
@roles_required("admin", "manager", "cashier")
def send_order_receipt_sms(order_id):
    return _send_order_receipt_sms_impl(order_id)


def _send_order_receipt_sms_impl(order_id: int):
    order = CafeOrder.query.options(joinedload(CafeOrder.table)).get_or_404(order_id)
    country_code = (request.form.get("country_code") or "+91").strip()
    mobile = (request.form.get("mobile") or "").strip()
    if not mobile or not any(ch.isdigit() for ch in mobile):
        return jsonify({"ok": False, "message": "Valid mobile number is required."}), 400
    mobile_digits = "".join(ch for ch in mobile if ch.isdigit())
    message = _receipt_sms_message(order)
    ok, msg = _send_receipt_sms(country_code, mobile_digits, message)
    if ok:
        return jsonify({"ok": True, "message": "Receipt SMS sent."})
    return jsonify({"ok": False, "message": msg}), 400


@bp.route("/orders/<int:order_id>/edit-items", methods=["POST"])
@roles_required("admin", "manager", "cashier")
def edit_order_items(order_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _edit_order_items_impl(order_id, next_url)


def _edit_order_items_impl(order_id: int, next_url: str = ""):
    order = CafeOrder.query.get_or_404(order_id)
    if order.status in ["paid", "cancelled"]:
        flash("Paid/cancelled orders cannot be edited.", "error")
        return redirect(next_url or url_for("cafe.cashier", table_id=order.table_id, tab="running"))
    rows = list(order.order_items)
    kept = 0
    for oi in rows:
        keep_values = request.form.getlist(f"keep_{oi.id}")
        include_flag = "1" in keep_values
        qty_raw = (request.form.get(f"qty_{oi.id}") or "").strip()
        qty = int(qty_raw) if qty_raw.isdigit() else 0
        if not include_flag or qty <= 0:
            db.session.delete(oi)
            continue
        oi.quantity = qty
        oi.is_parcel = True if request.form.get(f"is_parcel_{oi.id}") else False
        kept += 1
    db.session.flush()
    order = CafeOrder.query.options(joinedload(CafeOrder.order_items)).get(order_id)
    if kept == 0:
        order.status = "cancelled"
    else:
        _refresh_order_status_from_items(order)
    _recalculate_order_totals(order)
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    flash("Order updated.", "success")
    return redirect(next_url or url_for("cafe.cashier", table_id=order.table_id, tab="running"))


@bp.route("/table-order", methods=["POST"])
def table_order():
    slug = (request.form.get("slug") or "").strip()
    table = CafeTable.query.filter_by(qr_slug=slug, active=True).first()
    if not table:
        flash("Invalid table QR.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    cutoff_message = _order_cutoff_message("qr")
    if cutoff_message:
        flash(cutoff_message, "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    order = _create_order(
        table_id=table.id,
        ordered_by_user_id=g.current_user.id if getattr(g, "current_user", None) else _get_qr_guest_user_id(),
        status="pending_approval",
        payment_type=None,
        payment_reference=None,
    )
    if not order:
        flash("Please add at least one menu item in cart.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    session["qr_success_toast"] = "Order placed successfully"
    return redirect(url_for("main.table_qr_page", slug=slug))


@bp.route("/table-order/<int:order_id>/delete", methods=["POST"])
def delete_pending_table_order(order_id):
    slug = (request.form.get("slug") or "").strip()
    table = CafeTable.query.filter_by(qr_slug=slug, active=True).first()
    if not table:
        flash("Invalid table QR.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    order = CafeOrder.query.get_or_404(order_id)
    if order.table_id != table.id or order.status != "pending_approval":
        flash("Only pending approval orders can be deleted.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    for oi in list(order.order_items):
        db.session.delete(oi)
    db.session.delete(order)
    db.session.commit()
    flash("Pending order deleted.", "success")
    return redirect(url_for("main.table_qr_page", slug=slug))


@bp.route("/table-order/<int:order_id>/edit", methods=["POST"])
def edit_pending_table_order(order_id):
    slug = (request.form.get("slug") or "").strip()
    table = CafeTable.query.filter_by(qr_slug=slug, active=True).first()
    if not table:
        flash("Invalid table QR.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    order = CafeOrder.query.get_or_404(order_id)
    if order.table_id != table.id or order.status != "pending_approval":
        flash("Only pending approval orders can be edited.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    line_items = _parse_line_items_from_request()
    if not line_items:
        flash("Please add at least one menu item in cart.", "error")
        return redirect(url_for("main.table_qr_page", slug=slug))
    for oi in list(order.order_items):
        db.session.delete(oi)
    for row in line_items:
        menu_item, qty, is_parcel, size_label, unit_price = row if len(row) == 5 else (*row, None, None)  # type: ignore
        price_to_use = float(unit_price) if unit_price is not None else float(menu_item.price)
        db.session.add(
            CafeOrderItem(
                order_id=order.id,
                menu_item_id=menu_item.id,
                quantity=qty,
                unit_price=price_to_use,
                size_label=size_label,
                is_parcel=bool(is_parcel),
                approval_status="pending",
                prep_status="pending",
            )
        )
    db.session.flush()
    _recalculate_order_totals(order)
    _refresh_order_status_from_items(order)
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    flash("Pending order updated.", "success")
    return redirect(url_for("main.table_qr_page", slug=slug))


@bp.route("/table/call-staff", methods=["POST"])
def call_staff_for_table():
    slug = (request.form.get("slug") or "").strip()
    table = CafeTable.query.filter_by(qr_slug=slug, active=True).first()
    if not table:
        return jsonify({"ok": False, "message": "Table not found."}), 404
    cooldown_remaining = _staff_call_cooldown_remaining_seconds(table)
    if cooldown_remaining > 0:
        minutes = cooldown_remaining // 60
        seconds = cooldown_remaining % 60
        return jsonify(
            {
                "ok": False,
                "message": f"Please wait {minutes}:{seconds:02d} before calling staff again.",
                "cooldown_remaining": cooldown_remaining,
            }
        ), 429
    table.last_staff_call_at = datetime.utcnow()
    db.session.commit()
    _emit_ops_notification(
        f"Staff requested at {table.name}",
        kind="staff_call",
        payload={"table_id": table.id, "table_name": table.name},
    )
    return jsonify(
        {
            "ok": True,
            "message": f"Staff has been notified for {table.name}.",
            "cooldown_remaining": STAFF_CALL_COOLDOWN_SECONDS,
        }
    )


@bp.route("/reception/<string:access_key>/orders/<int:order_id>/approve", methods=["POST"])
def reception_kiosk_approve_order(access_key, order_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _reception_kiosk_url(access_key)
    return _approve_order_impl(order_id, next_url)


@bp.route("/reception/<string:access_key>/tables/<int:table_id>/clear-orders", methods=["POST"])
def reception_kiosk_clear_table_orders(access_key, table_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _reception_kiosk_url(access_key)
    return _clear_table_orders_impl(table_id, next_url)


@bp.route("/reception/<string:access_key>/tables/<int:table_id>/move-orders", methods=["POST"])
def reception_kiosk_move_table_orders(access_key, table_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _reception_kiosk_url(access_key)
    return _move_table_orders_impl(table_id, next_url)


@bp.route("/reception/<string:access_key>/orders/<int:order_id>/edit-items", methods=["POST"])
def reception_kiosk_edit_order_items(access_key, order_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _reception_kiosk_url(access_key)
    return _edit_order_items_impl(order_id, next_url)


@bp.route("/reception/<string:access_key>/orders/<int:order_id>/send-receipt-sms", methods=["POST"])
def reception_kiosk_send_order_receipt_sms(access_key, order_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return jsonify({"ok": False, "message": "Invalid reception kiosk access key."}), 403
    return _send_order_receipt_sms_impl(order_id)


@bp.route("/reception/<string:access_key>/orders/<int:order_id>/feedback", methods=["GET", "POST"])
def reception_kiosk_settlement_feedback(access_key, order_id):
    if not _has_valid_reception_kiosk_access(access_key):
        return Response("Invalid reception kiosk access key.", status=403)
    order = CafeOrder.query.options(
        joinedload(CafeOrder.table),
        joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
    ).get_or_404(order_id)
    return _render_feedback_form(order, source="offline", kiosk_mode=True, kiosk_access_key=access_key)


def _render_kitchen_display(station: str = "kitchen", kiosk_mode: bool = False, access_key: str = ""):
    _ensure_workstations_seeded()
    workstation_options = _all_workstations()
    station_lookup = {ws.slug: ws for ws in workstation_options}
    fallback_station = workstation_options[0].slug if workstation_options else "kitchen"
    station = (request.args.get("station") or station or fallback_station).strip().lower()
    if station not in station_lookup:
        station = fallback_station
    station_name = station_lookup.get(station).name if station in station_lookup else _workstation_display_name(station)
    today_start, today_end = _current_ist_day_bounds()
    orders = (
        CafeOrder.query.join(CafeOrderItem, CafeOrderItem.order_id == CafeOrder.id)
        .join(MenuItem, MenuItem.id == CafeOrderItem.menu_item_id)
        .filter(
            CafeOrder.status.in_(["pending_approval", "open", "preparing", "ready", "served"]),
            CafeOrder.created_at >= today_start,
            CafeOrder.created_at <= today_end,
            MenuItem.prep_station == station,
        )
        .options(
            joinedload(CafeOrder.table),
            joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
        )
        .distinct()
        .order_by(CafeOrder.created_at.asc())
        .all()
    )
    recipes = (
        InventoryRecipe.query.join(MenuItem, MenuItem.id == InventoryRecipe.menu_item_id)
        .filter(InventoryRecipe.active.is_(True), MenuItem.prep_station == station)
        .options(joinedload(InventoryRecipe.menu_item), joinedload(InventoryRecipe.ingredients).joinedload(InventoryRecipeItem.inventory_item))
        .all()
    )
    recipe_map = {recipe.menu_item_id: recipe for recipe in recipes}
    table_cards_map = {}
    prep_minutes = []
    status_changed = False
    for order in orders:
        previous_status = (order.status or "").strip().lower()
        derived_status = _refresh_order_status_from_items(order)
        if derived_status != previous_status:
            status_changed = True
        order_expected = 0
        created_local = _ist_from_utc_naive(order.created_at)
        elapsed_minutes = max(0, int(((datetime.now(IST_TZ) - created_local).total_seconds() // 60) if created_local else 0))
        prep_minutes.append(elapsed_minutes)
        table_key = f"table-{order.table_id or 0}"
        table_label = order.table.name if order.table else "-"
        card = table_cards_map.get(table_key)
        if not card:
            card = {
                "id": table_key,
                "table_id": order.table_id or 0,
                "table_name": table_label,
                "status": "open",
                "order_count": 0,
                "created_at": _format_ist(order.created_at, "%I:%M %p"),
                "created_at_iso": created_local.isoformat() if created_local else "",
                "expected_minutes": 0,
                "items": [],
                "order_refs": [],
            }
            table_cards_map[table_key] = card
        card["order_count"] += 1
        card["expected_minutes"] = max(int(card["expected_minutes"] or 0), order_expected)
        if created_local and (not card["created_at_iso"] or created_local.isoformat() < card["created_at_iso"]):
            card["created_at_iso"] = created_local.isoformat()
            card["created_at"] = _format_ist(order.created_at, "%I:%M %p")
        card["order_refs"].append(
            {
                "pickup_no": _format_pickup_number(order),
                "display_code": order.display_code or _format_internal_order_code(order),
                "created_at": _format_ist(order.created_at, "%I:%M:%S %p"),
            }
        )
        for oi in sorted(order.order_items, key=lambda row: (row.created_at or datetime.min, row.id or 0)):
            if not oi.menu_item or (oi.menu_item.prep_station or "").strip().lower() != station:
                continue
            if (oi.approval_status or "pending") == "rejected":
                continue
            recipe = recipe_map.get(oi.menu_item_id)
            sop = _serialize_recipe_sop(recipe, oi.size_label)
            expected_minutes = int(sop.get("prep_time_minutes") or 0)
            if expected_minutes > 0:
                order_expected = max(order_expected, expected_minutes)
            card["items"].append(
                {
                    "id": oi.id,
                    "name": oi.menu_item.name,
                    "qty": int(oi.quantity or 0),
                    "size_label": oi.size_label or "",
                    "is_parcel": bool(oi.is_parcel),
                    "approval_status": oi.approval_status or "pending",
                    "prep_status": _normalized_item_prep_status(oi),
                    "order_pickup_no": _format_pickup_number(order),
                    "order_display_code": order.display_code or _format_internal_order_code(order),
                    "ordered_at": _format_ist(order.created_at, "%I:%M:%S %p"),
                    "sop": sop,
                }
            )
        card["expected_minutes"] = max(int(card["expected_minutes"] or 0), order_expected)
        if order.status == "preparing":
            card["status"] = "preparing"
        elif order.status == "ready" and card["status"] not in ["preparing"]:
            card["status"] = "ready"
        elif order.status == "served" and card["status"] not in ["preparing", "ready"]:
            card["status"] = "served"
        elif order.status == "pending_approval" and card["status"] == "open":
            card["status"] = "pending_approval"
    if status_changed:
        db.session.commit()
    order_cards = []
    for card in table_cards_map.values():
        if not card["items"]:
            continue
        card["items"] = sorted(
            card["items"],
            key=lambda row: (
                row["approval_status"] != "pending",
                row["prep_status"] == "served",
                row["ordered_at"],
                row["id"],
            ),
        )
        pending_approval = any((item.get("approval_status") or "pending") == "pending" for item in card["items"])
        approved_items = [item for item in card["items"] if (item.get("approval_status") or "pending") == "approved"]
        if approved_items:
            prep_states = [item.get("prep_status") or "pending" for item in approved_items]
            if all(state == "served" for state in prep_states):
                card["status"] = "served"
            elif any(state == "preparing" for state in prep_states):
                card["status"] = "preparing"
            elif any(state == "ready" for state in prep_states):
                card["status"] = "ready"
            elif pending_approval:
                card["status"] = "pending_approval"
            else:
                card["status"] = "open"
        else:
            card["status"] = "pending_approval" if pending_approval else "open"
        card["order_refs"] = sorted(card["order_refs"], key=lambda row: row["created_at"])
        order_cards.append(card)
    order_cards.sort(key=lambda row: row["created_at_iso"] or "")
    sop_library = [
        {
            "menu_item_id": recipe.menu_item_id,
            "name": recipe.menu_item.name if recipe.menu_item else "Menu Item",
            "prep_station": recipe.menu_item.prep_station if recipe.menu_item else station,
            "sizes": _load_menu_item_size_variants(recipe.menu_item),
            "sop": _serialize_recipe_sop(recipe),
        }
        for recipe in sorted(recipes, key=lambda r: (r.menu_item.name.lower() if r.menu_item else ""))
    ]
    avg_ticket_minutes = round(sum(prep_minutes) / len(prep_minutes), 1) if prep_minutes else 0
    return render_template(
        "cafe/kitchen_display.html",
        orders=orders,
        order_cards=order_cards,
        station=station,
        station_name=station_name,
        station_action_label=_station_action_label(station, station_name),
        workstation_options=workstation_options,
        active_orders=len(order_cards),
        avg_ticket_minutes=avg_ticket_minutes,
        sop_library=sop_library,
        kiosk_mode=kiosk_mode,
        kiosk_access_key=access_key,
        current_page_url=_kiosk_display_url(access_key, station) if kiosk_mode and access_key else url_for("cafe.kitchen_display", station=station),
        status_update_url_template=(
            url_for("cafe.kiosk_update_order_status", access_key=access_key, order_id=0)
            if kiosk_mode and access_key
            else url_for("cafe.update_order_status", order_id=0)
        ),
        item_status_update_url_template=(
            url_for("cafe.kiosk_update_order_item_status", access_key=access_key, item_id=0)
            if kiosk_mode and access_key
            else url_for("cafe.update_order_item_status", item_id=0)
        ),
        hide_staff_nav=kiosk_mode,
    )


@bp.route("/kitchen")
@login_required
def kitchen_display():
    return _render_kitchen_display("kitchen", kiosk_mode=False)


@bp.route("/display/<string:access_key>")
@bp.route("/display/<string:access_key>/<string:station>")
def kiosk_display(access_key, station="kitchen"):
    if not _has_valid_kiosk_access(access_key):
        return Response("Invalid kiosk access key.", status=403)
    return _render_kitchen_display(station=station, kiosk_mode=True, access_key=access_key)


@bp.route("/barista")
@login_required
def barista_display():
    return redirect(url_for("cafe.kitchen_display", station="barista"))


@bp.route("/orders/<int:order_id>/status", methods=["POST"])
@roles_required("admin", "manager", "staff", "server", "barista", "chef", "cashier")
def update_order_status(order_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _update_order_status_impl(order_id, next_url)


def _update_order_status_impl(order_id: int, next_url: str = ""):
    order = CafeOrder.query.get_or_404(order_id)
    new_status = request.form["status"]
    if new_status not in ["open", "preparing", "ready", "served", "cancelled", "paid"]:
        flash("Invalid status.", "error")
        return redirect(next_url or url_for("cafe.kitchen_display"))
    if new_status in ["open", "preparing", "ready", "served"]:
        for item in order.order_items:
            if (item.approval_status or "pending") == "approved":
                item.prep_status = "pending" if new_status == "open" else new_status
    order.status = new_status
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    flash("Order status updated.", "success")
    return redirect(next_url or url_for("cafe.kitchen_display"))


@bp.route("/display/<string:access_key>/orders/<int:order_id>/status", methods=["POST"])
def kiosk_update_order_status(access_key, order_id):
    if not _has_valid_kiosk_access(access_key):
        return Response("Invalid kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _kiosk_display_url(access_key)
    return _update_order_status_impl(order_id, next_url)


@bp.route("/order-items/<int:item_id>/status", methods=["POST"])
@roles_required("admin", "manager", "staff", "server", "barista", "chef", "cashier")
def update_order_item_status(item_id):
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip()
    return _update_order_item_status_impl(item_id, next_url)


def _update_order_item_status_impl(item_id: int, next_url: str = ""):
    item = CafeOrderItem.query.options(joinedload(CafeOrderItem.order).joinedload(CafeOrder.order_items)).get_or_404(item_id)
    order = item.order
    new_status = (request.form.get("status") or "").strip().lower()
    if new_status not in ITEM_PREP_STATUSES:
        flash("Invalid item status.", "error")
        return redirect(next_url or url_for("cafe.kitchen_display"))
    if (item.approval_status or "pending") != "approved":
        flash("Only approved items can be updated from the display.", "error")
        return redirect(next_url or url_for("cafe.kitchen_display"))
    item.prep_status = new_status
    _refresh_order_status_from_items(order)
    db.session.commit()
    payload = _serialize_order(order)
    socketio.emit("order_updated", payload, namespace="/kitchen")
    socketio.emit("order_updated", payload, namespace="/table")
    flash("Item status updated.", "success")
    return redirect(next_url or url_for("cafe.kitchen_display"))


@bp.route("/display/<string:access_key>/order-items/<int:item_id>/status", methods=["POST"])
def kiosk_update_order_item_status(access_key, item_id):
    if not _has_valid_kiosk_access(access_key):
        return Response("Invalid kiosk access key.", status=403)
    next_url = request.form.get("next", "").strip() or request.args.get("next", "").strip() or _kiosk_display_url(access_key)
    return _update_order_item_status_impl(item_id, next_url)


def _inventory_item_status(item: InventoryItem):
    current_amount = float(item.current_amount or 0)
    reorder_level = float(item.reorder_level or 0)
    required_amount = float(item.required_amount or 0)
    if current_amount <= reorder_level:
        return "low"
    if required_amount > 0 and current_amount >= required_amount * 1.35:
        return "overstock"
    return "healthy"


def _inventory_filtered_items(items, search_text: str, area_filter: str, category_filter: str, status_filter: str):
    filtered = []
    q = search_text.strip().lower()
    for item in items:
        status_key = _inventory_item_status(item)
        if q:
            hay = " ".join(
                [
                    item.name or "",
                    item.item_code or "",
                    item.category_name or "",
                    item.subcategory_name or "",
                    item.storage_location or "",
                    item.selling_relation or "",
                ]
            ).lower()
            if q not in hay:
                continue
        if area_filter != "all" and (item.area or "").lower() != area_filter:
            continue
        if category_filter != "all" and (item.category_name or "").strip().lower() != category_filter:
            continue
        if status_filter != "all" and status_key != status_filter:
            continue
        filtered.append(item)
    return filtered


def _inventory_period_bounds(period_key: str, today: date | None = None) -> tuple[date, date]:
    today = today or date.today()
    period = (period_key or "today").strip().lower()
    if period == "yesterday":
        target = today - timedelta(days=1)
        return target, target
    if period == "week":
        start = today - timedelta(days=today.weekday())
        return start, today
    if period == "last_week":
        end = today - timedelta(days=today.weekday() + 1)
        start = end - timedelta(days=6)
        return start, end
    if period == "month":
        start = today.replace(day=1)
        return start, today
    if period == "last_month":
        this_month_start = today.replace(day=1)
        end = this_month_start - timedelta(days=1)
        start = end.replace(day=1)
        return start, end
    if period == "quarter":
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        start = date(today.year, quarter_start_month, 1)
        return start, today
    if period == "last_quarter":
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        this_quarter_start = date(today.year, quarter_start_month, 1)
        prev_quarter_end = this_quarter_start - timedelta(days=1)
        prev_quarter_start_month = ((prev_quarter_end.month - 1) // 3) * 3 + 1
        start = date(prev_quarter_end.year, prev_quarter_start_month, 1)
        return start, prev_quarter_end
    if period == "year":
        start = date(today.year, 1, 1)
        return start, today
    return today, today


def _inventory_period_options() -> list[tuple[str, str]]:
    return [
        ("today", "Today"),
        ("yesterday", "Yesterday"),
        ("week", "This Week"),
        ("last_week", "Last Week"),
        ("month", "This Month"),
        ("last_month", "Last Month"),
        ("quarter", "This Quarter"),
        ("last_quarter", "Last Quarter"),
        ("year", "This Year"),
    ]


def _purchase_todo_payloads():
    active_rows = (
        InventoryToPurchase.query.options(
            joinedload(InventoryToPurchase.category),
            joinedload(InventoryToPurchase.created_by),
            joinedload(InventoryToPurchase.completed_by),
            joinedload(InventoryToPurchase.closed_by),
        )
        .filter_by(active=True)
        .order_by(
            db.case((InventoryToPurchase.status == "purchased", 1), else_=0).asc(),
            InventoryToPurchase.created_at.asc(),
        )
        .all()
    )
    history_rows = (
        InventoryToPurchase.query.options(
            joinedload(InventoryToPurchase.category),
            joinedload(InventoryToPurchase.created_by),
            joinedload(InventoryToPurchase.completed_by),
            joinedload(InventoryToPurchase.closed_by),
        )
        .filter_by(active=False)
        .order_by(InventoryToPurchase.closed_at.desc(), InventoryToPurchase.updated_at.desc())
        .limit(30)
        .all()
    )
    active_payload = [
        {
            "id": row.id,
            "item_name": row.item_name,
            "category_name": row.category.name if row.category else "",
            "quantity_note": row.quantity_note or "",
            "note": row.note or "",
            "status": row.status,
            "created_by_name": row.created_by.full_name if row.created_by else "-",
            "created_at_ist": _format_ist(row.created_at, "%d %b %Y, %I:%M %p"),
            "completed_by_name": row.completed_by.full_name if row.completed_by else "",
            "completed_at_ist": _format_ist(row.completed_at, "%d %b %Y, %I:%M %p") if row.completed_at else "",
        }
        for row in active_rows
    ]
    history_payload = [
        {
            "item_name": row.item_name,
            "category_name": row.category.name if row.category else "-",
            "quantity_note": row.quantity_note or "",
            "status": row.status.replace("_", " ").title(),
            "created_by_name": row.created_by.full_name if row.created_by else "-",
            "closed_by_name": row.closed_by.full_name if row.closed_by else "-",
            "closed_at_ist": _format_ist(row.closed_at, "%d %b %Y, %I:%M %p") if row.closed_at else "-",
        }
        for row in history_rows
    ]
    return active_payload, history_payload


def _dedupe_active_purchase_todos():
    rows = (
        InventoryToPurchase.query.filter_by(active=True)
        .order_by(InventoryToPurchase.created_at.asc(), InventoryToPurchase.id.asc())
        .all()
    )
    seen: set[tuple[str, str]] = set()
    changed = False
    for row in rows:
        key = (
            (row.item_name or "").strip().lower(),
            (row.quantity_note or "").strip().lower(),
        )
        if not key[0]:
            continue
        if key in seen:
            row.active = False
            row.status = "removed"
            row.closed_at = row.closed_at or datetime.utcnow()
            changed = True
            continue
        seen.add(key)
    if changed:
        db.session.commit()


@bp.route("/to-purchase", methods=["GET", "POST"])
@roles_required("owner", "admin", "manager", "accountant", "cashier", "staff", "server", "barista", "chef", "inventory_manager", "cleaner", "librarian", "delivery_partner")
def to_purchase():
    if request.method == "POST":
        action = (request.form.get("action") or "add_batch_to_purchase").strip().lower()

        if action == "add_batch_to_purchase":
            item_names = request.form.getlist("item_name[]")
            quantities = request.form.getlist("quantity_note[]")
            added = 0
            for idx, raw_name in enumerate(item_names):
                item_name = (raw_name or "").strip()
                quantity_note = (quantities[idx] if idx < len(quantities) else "").strip()
                if not item_name or not quantity_note:
                    continue
                existing = InventoryToPurchase.query.filter(
                    InventoryToPurchase.active.is_(True),
                    db.func.lower(db.func.trim(InventoryToPurchase.item_name)) == item_name.lower(),
                    db.func.lower(db.func.trim(db.func.coalesce(InventoryToPurchase.quantity_note, ""))) == quantity_note.lower(),
                ).first()
                if existing:
                    continue
                db.session.add(
                    InventoryToPurchase(
                        item_name=item_name,
                        quantity_note=quantity_note,
                        status="open",
                        active=True,
                        created_by_user_id=g.current_user.id if g.current_user else None,
                    )
                )
                added += 1
            if added == 0:
                flash("Please enter at least one item with quantity.", "error")
                return redirect(url_for("cafe.to_purchase"))
            db.session.commit()
            flash(f"Added {added} item(s) to purchase list.", "success")
            return redirect(url_for("cafe.to_purchase"))

        if action == "toggle_to_purchase":
            row = InventoryToPurchase.query.get_or_404(_safe_int(request.form.get("todo_id"), 0))
            if row.status == "purchased":
                row.status = "open"
                row.completed_at = None
                row.completed_by_user_id = None
            else:
                row.status = "purchased"
                row.completed_at = datetime.utcnow()
                row.completed_by_user_id = g.current_user.id if g.current_user else None
            db.session.commit()
            flash("Purchase item updated.", "success")
            return redirect(url_for("cafe.to_purchase"))

        if action == "remove_to_purchase":
            row = InventoryToPurchase.query.get_or_404(_safe_int(request.form.get("todo_id"), 0))
            row.active = False
            row.status = "removed"
            row.closed_at = datetime.utcnow()
            row.closed_by_user_id = g.current_user.id if g.current_user else None
            db.session.commit()
            flash("Purchase item removed.", "success")
            return redirect(url_for("cafe.to_purchase"))

        if action == "clear_to_purchase":
            active_rows = InventoryToPurchase.query.filter_by(active=True).all()
            now_utc = datetime.utcnow()
            changed = 0
            for row in active_rows:
                row.active = False
                if row.status not in {"purchased", "removed"}:
                    row.status = "cleared"
                row.closed_at = now_utc
                row.closed_by_user_id = g.current_user.id if g.current_user else None
                changed += 1
            db.session.commit()
            flash(f"Cleared {changed} purchase item(s).", "success")
            return redirect(url_for("cafe.to_purchase"))

    _dedupe_active_purchase_todos()
    purchase_todos_active, purchase_todos_history = _purchase_todo_payloads()
    return render_template(
        "cafe/to_purchase.html",
        purchase_todos_active=purchase_todos_active,
        purchase_todos_history=purchase_todos_history,
    )


@bp.route("/inventory", methods=["GET", "POST"])
@roles_required("owner", "admin", "manager", "accountant", "barista", "inventory_manager")
def inventory():
    section = (request.args.get("section") or "dashboard").strip().lower()
    allowed_sections = {
        "dashboard", "daily_closing", "stock_levels", "purchases", "vendors",
        "recipes", "wastage", "analytics", "categories", "settings", "to_purchase"
    }
    if section not in allowed_sections:
        section = "dashboard"

    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower()
        if action == "log_expense":
            category_id = _safe_int(request.form.get("category_id"), 0)
            vendor_id = _safe_int(request.form.get("vendor_id"), 0) or None
            amount = _safe_float(request.form.get("amount"), 0)
            transaction_mode = (request.form.get("transaction_mode") or "qr").strip().lower()
            if transaction_mode not in {"cash", "card", "qr"}:
                transaction_mode = "qr"
            entry_date_raw = (request.form.get("entry_date") or "").strip()
            try:
                entry_date = date.fromisoformat(entry_date_raw) if entry_date_raw else date.today()
            except ValueError:
                entry_date = date.today()
            if category_id <= 0:
                flash("Please select an inventory category.", "error")
                return redirect(url_for("cafe.inventory", section="dashboard"))
            if amount <= 0:
                flash("Please enter a valid inventory amount.", "error")
                return redirect(url_for("cafe.inventory", section="dashboard"))
            log_row = InventoryExpenseLog(
                entry_date=entry_date,
                category_id=category_id,
                vendor_id=vendor_id,
                amount=round(amount, 2),
                transaction_mode=transaction_mode,
                note=(request.form.get("note") or "").strip() or None,
                created_by_user_id=g.current_user.id if g.current_user else None,
            )
            db.session.add(log_row)
            db.session.commit()
            flash("Inventory expense logged.", "success")
            return redirect(url_for("cafe.inventory", section="dashboard", period=request.args.get("period") or "today"))

        if action == "add_item":
            item = InventoryItem(
                item_code=(request.form.get("item_code") or "").strip() or None,
                area=_normalize_inventory_area(request.form.get("area")),
                name=request.form.get("name", "").strip(),
                category_name=(request.form.get("category_name") or "").strip() or None,
                subcategory_name=(request.form.get("subcategory_name") or "").strip() or None,
                unit=(request.form.get("unit") or "pcs").strip(),
                current_amount=_safe_float(request.form.get("current_amount"), 0),
                required_amount=_safe_float(request.form.get("required_amount"), 0),
                reorder_level=_safe_float(request.form.get("reorder_level"), 0),
                average_daily_usage=_safe_float(request.form.get("average_daily_usage"), 0),
                purchase_price=_safe_float(request.form.get("purchase_price"), 0),
                selling_relation=(request.form.get("selling_relation") or "").strip() or None,
                shelf_life_days=_safe_int(request.form.get("shelf_life_days"), 0) if request.form.get("shelf_life_days") else None,
                expiry_tracking=True if request.form.get("expiry_tracking") else False,
                storage_location=(request.form.get("storage_location") or "").strip() or None,
                vendor_id=int(request.form.get("vendor_id")) if request.form.get("vendor_id") else None,
                note=(request.form.get("note") or "").strip() or None,
            )
            if not item.name:
                flash("Item name is required.", "error")
                return redirect(url_for("cafe.inventory", section="stock_levels"))
            db.session.add(item)
            db.session.commit()
            flash("Inventory item added.", "success")
            return redirect(url_for("cafe.inventory", section="stock_levels"))

        if action == "update_item":
            item = InventoryItem.query.get_or_404(_safe_int(request.form.get("item_id"), 0))
            item.item_code = (request.form.get("item_code") or "").strip() or None
            item.name = (request.form.get("name") or "").strip()
            item.category_name = (request.form.get("category_name") or "").strip() or None
            item.subcategory_name = (request.form.get("subcategory_name") or "").strip() or None
            item.area = _normalize_inventory_area(request.form.get("area"), item.area)
            item.unit = (request.form.get("unit") or item.unit or "pcs").strip()
            item.current_amount = _safe_float(request.form.get("current_amount"), item.current_amount or 0)
            item.reorder_level = _safe_float(request.form.get("reorder_level"), item.reorder_level or 0)
            item.required_amount = _safe_float(request.form.get("required_amount"), item.required_amount or 0)
            item.average_daily_usage = _safe_float(request.form.get("average_daily_usage"), item.average_daily_usage or 0)
            item.purchase_price = _safe_float(request.form.get("purchase_price"), item.purchase_price or 0)
            item.shelf_life_days = _safe_int(request.form.get("shelf_life_days"), 0) if request.form.get("shelf_life_days") else None
            item.expiry_tracking = True if request.form.get("expiry_tracking") else False
            item.vendor_id = _safe_int(request.form.get("vendor_id"), 0) or None
            item.storage_location = (request.form.get("storage_location") or "").strip() or None
            item.selling_relation = (request.form.get("selling_relation") or "").strip() or None
            item.note = (request.form.get("note") or "").strip() or None
            if not item.name:
                flash("Item name is required.", "error")
                return redirect(url_for("cafe.inventory", section="stock_levels", edit_item_id=item.id))
            db.session.commit()
            flash("Inventory item updated.", "success")
            return redirect(url_for("cafe.inventory", section="stock_levels", edit_item_id=item.id))

        if action == "adjust_stock":
            item = InventoryItem.query.get_or_404(_safe_int(request.form.get("item_id"), 0))
            adjustment = _safe_float(request.form.get("adjustment"), 0)
            if adjustment == 0:
                flash("Enter a stock adjustment value.", "error")
                return redirect(url_for("cafe.inventory", section="stock_levels"))
            item.current_amount = round(max(0.0, float(item.current_amount or 0) + adjustment), 3)
            db.session.commit()
            flash("Stock adjusted.", "success")
            return redirect(url_for("cafe.inventory", section="stock_levels", edit_item_id=item.id))

        if action == "daily_closing_save":
            closing_date = date.fromisoformat(request.form.get("closing_date") or date.today().isoformat())
            item_ids = [int(v) for v in request.form.getlist("item_id") if str(v).isdigit()]
            for item_id in item_ids:
                item = InventoryItem.query.get(item_id)
                if not item:
                    continue
                closing_raw = (request.form.get(f"closing_stock_{item_id}") or "").strip()
                if closing_raw == "":
                    continue
                closing_stock = _safe_float(closing_raw, 0)
                prev_row = (
                    InventoryDailyClosing.query.filter(
                        InventoryDailyClosing.item_id == item_id,
                        InventoryDailyClosing.closing_date < closing_date,
                    )
                    .order_by(InventoryDailyClosing.closing_date.desc())
                    .first()
                )
                opening_stock = float(prev_row.closing_stock) if prev_row else float(item.current_amount or 0)
                consumed = round(opening_stock - closing_stock, 3)
                variance = round(consumed - float(item.average_daily_usage or 0), 3)
                row = InventoryDailyClosing.query.filter_by(item_id=item_id, closing_date=closing_date).first()
                if not row:
                    row = InventoryDailyClosing(item_id=item_id, closing_date=closing_date)
                    db.session.add(row)
                row.opening_stock = opening_stock
                row.closing_stock = closing_stock
                row.consumed_amount = consumed
                row.variance_amount = variance
                row.note = (request.form.get(f"closing_note_{item_id}") or "").strip() or None
                item.current_amount = closing_stock
            db.session.commit()
            flash("Daily closing saved.", "success")
            return redirect(url_for("cafe.inventory", section="daily_closing", closing_date=closing_date.isoformat()))

        if action == "add_vendor":
            vendor = InventoryVendor(
                name=(request.form.get("name") or "").strip(),
                vendor_category=(request.form.get("vendor_category") or "").strip() or None,
                contact_person=(request.form.get("contact_person") or "").strip() or None,
                phone=(request.form.get("phone") or "").strip() or None,
                email=(request.form.get("email") or "").strip() or None,
                gst_number=(request.form.get("gst_number") or "").strip() or None,
                payment_terms=(request.form.get("payment_terms") or "").strip() or None,
                outstanding_balance=_safe_float(request.form.get("outstanding_balance"), 0),
                average_rate_note=(request.form.get("average_rate_note") or "").strip() or None,
                note=(request.form.get("note") or "").strip() or None,
                active=True,
            )
            if not vendor.name:
                flash("Vendor name is required.", "error")
                return redirect(url_for("cafe.inventory", section="vendors"))
            db.session.add(vendor)
            db.session.commit()
            flash("Vendor added.", "success")
            return redirect(url_for("cafe.inventory", section="vendors"))

        if action == "update_vendor":
            vendor = InventoryVendor.query.get_or_404(_safe_int(request.form.get("vendor_id"), 0))
            vendor.name = (request.form.get("name") or "").strip() or vendor.name
            vendor.vendor_category = (request.form.get("vendor_category") or "").strip() or None
            vendor.contact_person = (request.form.get("contact_person") or "").strip() or None
            vendor.phone = (request.form.get("phone") or "").strip() or None
            vendor.email = (request.form.get("email") or "").strip() or None
            vendor.gst_number = (request.form.get("gst_number") or "").strip() or None
            vendor.payment_terms = (request.form.get("payment_terms") or "").strip() or None
            vendor.outstanding_balance = _safe_float(request.form.get("outstanding_balance"), vendor.outstanding_balance or 0)
            vendor.average_rate_note = (request.form.get("average_rate_note") or "").strip() or None
            vendor.note = (request.form.get("note") or "").strip() or None
            db.session.commit()
            flash("Vendor updated.", "success")
            return redirect(url_for("cafe.inventory", section="vendors", edit_vendor_id=vendor.id))

        if action == "add_to_purchase":
            item_name = (request.form.get("item_name") or "").strip()
            if not item_name:
                flash("Please enter an item to purchase.", "error")
                return redirect(url_for("cafe.inventory", section="to_purchase"))
            row = InventoryToPurchase(
                item_name=item_name,
                category_id=_safe_int(request.form.get("category_id"), 0) or None,
                quantity_note=(request.form.get("quantity_note") or "").strip() or None,
                note=(request.form.get("note") or "").strip() or None,
                status="open",
                active=True,
                created_by_user_id=g.current_user.id if g.current_user else None,
            )
            db.session.add(row)
            db.session.commit()
            flash("Added to purchase list.", "success")
            return redirect(url_for("cafe.inventory", section="to_purchase"))

        if action == "toggle_to_purchase":
            row = InventoryToPurchase.query.get_or_404(_safe_int(request.form.get("todo_id"), 0))
            if row.status == "purchased":
                row.status = "open"
                row.completed_at = None
                row.completed_by_user_id = None
            else:
                row.status = "purchased"
                row.completed_at = datetime.utcnow()
                row.completed_by_user_id = g.current_user.id if g.current_user else None
            db.session.commit()
            flash("Purchase item updated.", "success")
            return redirect(url_for("cafe.inventory", section="to_purchase"))

        if action == "remove_to_purchase":
            row = InventoryToPurchase.query.get_or_404(_safe_int(request.form.get("todo_id"), 0))
            row.active = False
            row.status = "removed"
            row.closed_at = datetime.utcnow()
            row.closed_by_user_id = g.current_user.id if g.current_user else None
            db.session.commit()
            flash("Purchase item removed.", "success")
            return redirect(url_for("cafe.inventory", section="to_purchase"))

        if action == "clear_to_purchase":
            active_rows = InventoryToPurchase.query.filter_by(active=True).all()
            now_utc = datetime.utcnow()
            changed = 0
            for row in active_rows:
                row.active = False
                if row.status not in {"purchased", "removed"}:
                    row.status = "cleared"
                row.closed_at = now_utc
                row.closed_by_user_id = g.current_user.id if g.current_user else None
                changed += 1
            db.session.commit()
            flash(f"Cleared {changed} purchase item(s).", "success")
            return redirect(url_for("cafe.inventory", section="to_purchase"))

        if action == "add_purchase":
            purchase = InventoryPurchase(
                purchase_date=date.fromisoformat(request.form.get("purchase_date") or date.today().isoformat()),
                vendor_id=int(request.form.get("vendor_id")) if request.form.get("vendor_id") else None,
                invoice_number=(request.form.get("invoice_number") or "").strip() or None,
                tax_amount=_safe_float(request.form.get("tax_amount"), 0),
                payment_status=(request.form.get("payment_status") or "pending").strip(),
                note=(request.form.get("note") or "").strip() or None,
            )
            db.session.add(purchase)
            db.session.flush()
            subtotal = 0.0
            item_ids = [int(v) for v in request.form.getlist("purchase_item_id") if str(v).isdigit()]
            for item_id in item_ids:
                qty = _safe_float(request.form.get(f"purchase_qty_{item_id}"), 0)
                unit_price = _safe_float(request.form.get(f"purchase_price_{item_id}"), 0)
                if qty <= 0:
                    continue
                line_total = round(qty * unit_price, 2)
                db.session.add(
                    InventoryPurchaseLine(
                        purchase_id=purchase.id,
                        item_id=item_id,
                        quantity=qty,
                        unit_price=unit_price,
                        line_total=line_total,
                    )
                )
                item = InventoryItem.query.get(item_id)
                if item:
                    item.current_amount = round(float(item.current_amount or 0) + qty, 3)
                    if unit_price > 0:
                        item.purchase_price = unit_price
                subtotal += line_total
            purchase.subtotal = round(subtotal, 2)
            purchase.total_amount = round(subtotal + float(purchase.tax_amount or 0), 2)
            db.session.commit()
            flash("Purchase logged and stock updated.", "success")
            return redirect(url_for("cafe.inventory", section="purchases"))

        if action == "save_recipe":
            menu_item_id = int(request.form.get("menu_item_id") or 0)
            if menu_item_id <= 0:
                flash("Please select a menu item.", "error")
                return redirect(url_for("cafe.inventory", section="recipes"))
            recipe = InventoryRecipe.query.filter_by(menu_item_id=menu_item_id).first()
            if not recipe:
                recipe = InventoryRecipe(menu_item_id=menu_item_id, yield_qty=1, active=True)
                db.session.add(recipe)
                db.session.flush()
            recipe.yield_qty = _safe_float(request.form.get("yield_qty"), 1)
            recipe.yield_unit = (request.form.get("yield_unit") or "").strip() or None
            recipe.prep_time_minutes = _safe_int(request.form.get("prep_time_minutes"), 0) or None
            recipe.ingredients_note = (request.form.get("ingredients_note") or "").strip() or None
            recipe.preparation_steps = (request.form.get("preparation_steps") or "").strip() or None
            recipe.plating_notes = (request.form.get("plating_notes") or "").strip() or None
            recipe.quality_checks = (request.form.get("quality_checks") or "").strip() or None
            recipe.allergy_alerts = (request.form.get("allergy_alerts") or "").strip() or None
            recipe.training_notes = (request.form.get("training_notes") or "").strip() or None
            recipe.sop_photo_url = (request.form.get("sop_photo_url") or "").strip() or None
            size_notes = _parse_recipe_size_notes_from_form()
            recipe.size_sop_json = json.dumps(size_notes) if size_notes else None
            for old in list(recipe.ingredients):
                db.session.delete(old)
            ingredient_ids = [int(v) for v in request.form.getlist("recipe_item_id") if str(v).isdigit()]
            for inv_id in ingredient_ids:
                qty = _safe_float(request.form.get(f"recipe_qty_{inv_id}"), 0)
                unit = (request.form.get(f"recipe_unit_{inv_id}") or "pcs").strip()
                if qty <= 0:
                    continue
                db.session.add(
                    InventoryRecipeItem(
                        recipe_id=recipe.id,
                        inventory_item_id=inv_id,
                        qty_per_menu=qty,
                        unit=unit,
                    )
                )
            db.session.commit()
            flash("Recipe and SOP saved.", "success")
            return redirect(url_for("cafe.inventory", section="recipes"))

        if action == "add_wastage":
            item_id = int(request.form.get("item_id") or 0)
            qty = _safe_float(request.form.get("quantity"), 0)
            if item_id <= 0 or qty <= 0:
                flash("Select item and quantity.", "error")
                return redirect(url_for("cafe.inventory", section="wastage"))
            wastage = InventoryWastage(
                wastage_date=date.fromisoformat(request.form.get("wastage_date") or date.today().isoformat()),
                item_id=item_id,
                quantity=qty,
                reason=(request.form.get("reason") or "").strip() or None,
            )
            db.session.add(wastage)
            item = InventoryItem.query.get(item_id)
            if item:
                item.current_amount = round(max(0.0, float(item.current_amount or 0) - qty), 3)
            db.session.commit()
            flash("Wastage logged.", "success")
            return redirect(url_for("cafe.inventory", section="wastage"))

        if action == "add_category":
            name = (request.form.get("name") or "").strip()
            if not name:
                flash("Category name is required.", "error")
                return redirect(url_for("cafe.inventory", section="categories"))
            if InventoryCategory.query.filter(db.func.lower(InventoryCategory.name) == name.lower()).first():
                flash("Category already exists.", "error")
                return redirect(url_for("cafe.inventory", section="categories"))
            db.session.add(
                InventoryCategory(
                    name=name,
                    icon=(request.form.get("icon") or "").strip() or None,
                    color=(request.form.get("color") or "").strip() or None,
                    active=True,
                )
            )
            db.session.commit()
            flash("Category added.", "success")
            return redirect(url_for("cafe.inventory", section="categories"))

        if action == "update_category":
            category = InventoryCategory.query.get_or_404(_safe_int(request.form.get("category_id"), 0))
            category.name = (request.form.get("name") or "").strip() or category.name
            category.icon = (request.form.get("icon") or "").strip() or None
            category.color = (request.form.get("color") or "").strip() or None
            category.active = True if request.form.get("active") else False
            db.session.commit()
            flash("Category updated.", "success")
            return redirect(url_for("cafe.inventory", section="categories", edit_category_id=category.id))

    closing_date = date.fromisoformat(request.args.get("closing_date") or date.today().isoformat())
    inventory_search = (request.args.get("q") or "").strip()
    inventory_period = (request.args.get("period") or "today").strip().lower()
    if inventory_period not in {key for key, _ in _inventory_period_options()}:
        inventory_period = "today"
    inventory_period_start, inventory_period_end = _inventory_period_bounds(inventory_period)
    inventory_tracking_category_id = request.args.get("track_category_id", type=int) or 0
    inventory_area_options = _inventory_area_options(include_all=True)
    inventory_area_name_map = _inventory_area_name_map()
    inventory_area = (request.args.get("area") or "all").strip().lower()
    if inventory_area not in {row["value"] for row in inventory_area_options}:
        inventory_area = "all"
    inventory_category_filter = (request.args.get("category") or "all").strip().lower()
    inventory_status_filter = (request.args.get("status") or "all").strip().lower()
    if inventory_status_filter not in ["all", "healthy", "low", "overstock"]:
        inventory_status_filter = "all"
    categories = InventoryCategory.query.filter_by(active=True).order_by(InventoryCategory.name.asc()).all()
    vendors = InventoryVendor.query.filter_by(active=True).order_by(InventoryVendor.name.asc()).all()
    items = InventoryItem.query.order_by(InventoryItem.category_name.asc(), InventoryItem.name.asc()).all()
    expense_logs = (
        InventoryExpenseLog.query.options(
            joinedload(InventoryExpenseLog.category),
            joinedload(InventoryExpenseLog.vendor),
            joinedload(InventoryExpenseLog.created_by),
        )
        .filter(
            InventoryExpenseLog.entry_date >= inventory_period_start,
            InventoryExpenseLog.entry_date <= inventory_period_end,
        )
        .order_by(InventoryExpenseLog.entry_date.desc(), InventoryExpenseLog.id.desc())
        .all()
    )
    if inventory_tracking_category_id:
        expense_logs = [row for row in expense_logs if row.category_id == inventory_tracking_category_id]
    filtered_items = _inventory_filtered_items(items, inventory_search, inventory_area, inventory_category_filter, inventory_status_filter)
    purchases = InventoryPurchase.query.order_by(InventoryPurchase.purchase_date.desc(), InventoryPurchase.id.desc()).limit(80).all()
    wastage_rows = InventoryWastage.query.order_by(InventoryWastage.wastage_date.desc(), InventoryWastage.id.desc()).limit(120).all()
    recipes = (
        InventoryRecipe.query.options(
            joinedload(InventoryRecipe.menu_item),
            joinedload(InventoryRecipe.ingredients).joinedload(InventoryRecipeItem.inventory_item),
        )
        .order_by(InventoryRecipe.id.desc())
        .all()
    )
    purchase_todos_active, purchase_todos_history = _purchase_todo_payloads()
    menu_items = MenuItem.query.filter_by(available=True, is_deleted=False).order_by(MenuItem.name.asc()).all()
    menu_item_meta = {
        item.id: {
            "name": item.name,
            "prep_station": item.prep_station,
            "prep_station_name": _workstation_display_name(item.prep_station),
            "sizes": _load_menu_item_size_variants(item),
        }
        for item in menu_items
    }
    recipe_payload_map = {
        recipe.menu_item_id: {
            "yield_qty": round(float(recipe.yield_qty or 0), 3),
            "yield_unit": recipe.yield_unit or "",
            "prep_time_minutes": int(recipe.prep_time_minutes or 0),
            "ingredients_note": recipe.ingredients_note or "",
            "preparation_steps": recipe.preparation_steps or "",
            "plating_notes": recipe.plating_notes or "",
            "quality_checks": recipe.quality_checks or "",
            "allergy_alerts": recipe.allergy_alerts or "",
            "training_notes": recipe.training_notes or "",
            "sop_photo_url": recipe.sop_photo_url or "",
            "size_notes": _recipe_size_note_map(recipe),
            "ingredient_ids": [int(ing.inventory_item_id) for ing in recipe.ingredients if ing.inventory_item_id],
            "ingredients": {
                int(ing.inventory_item_id): {
                    "qty": round(float(ing.qty_per_menu or 0), 3),
                    "unit": ing.unit or (ing.inventory_item.unit if ing.inventory_item else ""),
                }
                for ing in recipe.ingredients
                if ing.inventory_item_id
            },
        }
        for recipe in recipes
    }
    edit_item_id = request.args.get("edit_item_id", type=int)
    selected_item = InventoryItem.query.get(edit_item_id) if edit_item_id else (filtered_items[0] if filtered_items else None)
    edit_vendor_id = request.args.get("edit_vendor_id", type=int)
    selected_vendor = InventoryVendor.query.get(edit_vendor_id) if edit_vendor_id else (vendors[0] if vendors else None)
    edit_category_id = request.args.get("edit_category_id", type=int)
    selected_category = InventoryCategory.query.get(edit_category_id) if edit_category_id else (categories[0] if categories else None)

    category_stats = []
    for cat in categories:
        cat_items = [x for x in items if (x.category_name or "").strip().lower() == cat.name.lower()]
        stock_value = round(sum(float(x.current_amount or 0) * float(x.purchase_price or 0) for x in cat_items), 2)
        category_stats.append({"category": cat, "item_count": len(cat_items), "stock_value": stock_value})

    low_stock_items = [x for x in items if float(x.current_amount or 0) <= float(x.reorder_level or 0)]
    overstock_items = [x for x in items if _inventory_item_status(x) == "overstock"]
    items_without_vendor = [x for x in items if not x.vendor_id]
    total_stock_value = round(sum(float(x.current_amount or 0) * float(x.purchase_price or 0) for x in items), 2)
    today_purchase_spend = round(
        sum(float(p.total_amount or 0) for p in purchases if p.purchase_date == date.today()),
        2,
    )
    today_wastage_value = round(
        sum(float(w.quantity or 0) * float((w.item.purchase_price if w.item else 0) or 0) for w in wastage_rows if w.wastage_date == date.today()),
        2,
    )

    today_consumption = (
        db.session.query(db.func.coalesce(db.func.sum(InventoryDailyClosing.consumed_amount), 0.0))
        .filter(InventoryDailyClosing.closing_date == closing_date)
        .scalar()
        or 0.0
    )

    daily_rows = []
    for item in items:
        existing = InventoryDailyClosing.query.filter_by(item_id=item.id, closing_date=closing_date).first()
        prev_row = (
            InventoryDailyClosing.query.filter(
                InventoryDailyClosing.item_id == item.id,
                InventoryDailyClosing.closing_date < closing_date,
            )
            .order_by(InventoryDailyClosing.closing_date.desc())
            .first()
        )
        opening = float(prev_row.closing_stock) if prev_row else float(item.current_amount or 0)
        daily_rows.append({
            "item": item,
            "opening": opening,
            "existing": existing,
            "consumed": float(existing.consumed_amount) if existing else 0.0,
        })
    daily_rows_grouped = {}
    for row in daily_rows:
        key = row["item"].category_name or "Uncategorized"
        daily_rows_grouped.setdefault(key, []).append(row)

    vendor_purchase_map = {}
    for purchase in purchases:
        if purchase.vendor_id:
            data = vendor_purchase_map.setdefault(purchase.vendor_id, {"count": 0, "spend": 0.0})
            data["count"] += 1
            data["spend"] = round(data["spend"] + float(purchase.total_amount or 0), 2)

    top_stock_value_items = sorted(
        items,
        key=lambda x: float(x.current_amount or 0) * float(x.purchase_price or 0),
        reverse=True,
    )[:8]
    top_consumption_rows = sorted(
        daily_rows,
        key=lambda x: float(x["existing"].consumed_amount if x["existing"] else 0),
        reverse=True,
    )[:8]
    recent_purchase_lines = (
        InventoryPurchaseLine.query.options(joinedload(InventoryPurchaseLine.item), joinedload(InventoryPurchaseLine.purchase))
        .order_by(InventoryPurchaseLine.created_at.desc())
        .limit(12)
        .all()
    )

    inventory_analytics = {
        "total_items": len(items),
        "low_stock_count": len(low_stock_items),
        "overstock_count": len(overstock_items),
        "unassigned_vendor_count": len(items_without_vendor),
        "total_stock_value": total_stock_value,
        "today_purchase_spend": today_purchase_spend,
        "today_wastage_value": today_wastage_value,
        "today_consumption": round(today_consumption, 2),
        "to_purchase_count": len([row for row in purchase_todos_active if row["status"] != "purchased"]),
    }

    category_spend_map = {cat.id: 0.0 for cat in categories}
    category_entry_count_map = {cat.id: 0 for cat in categories}
    timeline_map: dict[date, float] = {}
    for log_row in expense_logs:
        category_spend_map[log_row.category_id] = round(category_spend_map.get(log_row.category_id, 0.0) + float(log_row.amount or 0), 2)
        category_entry_count_map[log_row.category_id] = category_entry_count_map.get(log_row.category_id, 0) + 1
        timeline_map[log_row.entry_date] = round(timeline_map.get(log_row.entry_date, 0.0) + float(log_row.amount or 0), 2)

    category_spend_rows = []
    for cat in categories:
        total = round(category_spend_map.get(cat.id, 0.0), 2)
        if inventory_tracking_category_id and cat.id != inventory_tracking_category_id:
            continue
        category_spend_rows.append(
            {
                "category": cat,
                "amount": total,
                "entry_count": category_entry_count_map.get(cat.id, 0),
            }
        )
    category_spend_rows.sort(key=lambda row: row["amount"], reverse=True)
    max_category_spend = max([row["amount"] for row in category_spend_rows], default=0.0)

    timeline_rows = []
    cursor = inventory_period_start
    while cursor <= inventory_period_end:
        timeline_rows.append({"date": cursor, "amount": round(timeline_map.get(cursor, 0.0), 2)})
        cursor += timedelta(days=1)
    max_timeline_amount = max([row["amount"] for row in timeline_rows], default=0.0)

    paid_orders = CafeOrder.query.filter(CafeOrder.status == "paid").all()
    period_revenue = 0.0
    for order in paid_orders:
        revenue_dt = order.paid_at or order.created_at
        revenue_local = _ist_from_utc_naive(revenue_dt)
        if not revenue_local:
            continue
        revenue_date = revenue_local.date()
        if inventory_period_start <= revenue_date <= inventory_period_end:
            period_revenue = round(period_revenue + float(order.total_amount or 0), 2)

    total_period_expense = round(sum(float(row.amount or 0) for row in expense_logs), 2)
    expense_vs_earning = {
        "expense": total_period_expense,
        "revenue": period_revenue,
        "difference": round(period_revenue - total_period_expense, 2),
    }
    average_daily_expense = round(total_period_expense / max((inventory_period_end - inventory_period_start).days + 1, 1), 2)
    recent_expense_logs = [
        {
            "expense_date": row.entry_date.strftime("%d %b %Y"),
            "category_name": row.category.name if row.category else "-",
            "vendor_name": row.vendor.name if row.vendor else "Other",
            "amount": round(float(row.amount or 0), 2),
            "transaction_mode": ((row.transaction_mode or "qr").strip().title() if (row.transaction_mode or "").strip() else "QR"),
            "logged_by": row.created_by.full_name if row.created_by else "-",
            "logged_at_ist": _format_ist(row.created_at, "%d %b %Y, %I:%M:%S %p"),
            "note": row.note or "",
        }
        for row in expense_logs[:12]
    ]

    return render_template(
        "cafe/inventory.html",
        section=section,
        categories=categories,
        category_stats=category_stats,
        vendors=vendors,
        items=items,
        purchases=purchases,
        recent_purchase_lines=recent_purchase_lines,
        recipes=recipes,
        menu_items=menu_items,
        wastage_rows=wastage_rows,
        low_stock_items=low_stock_items,
        overstock_items=overstock_items,
        items_without_vendor=items_without_vendor,
        daily_rows=daily_rows,
        daily_rows_grouped=daily_rows_grouped,
        closing_date=closing_date,
        inventory_analytics=inventory_analytics,
        filtered_items=filtered_items,
        selected_item=selected_item,
        selected_vendor=selected_vendor,
        selected_category=selected_category,
        inventory_search=inventory_search,
        inventory_period=inventory_period,
        inventory_period_options=_inventory_period_options(),
        inventory_period_start=inventory_period_start,
        inventory_period_end=inventory_period_end,
        inventory_tracking_category_id=inventory_tracking_category_id,
        inventory_area=inventory_area,
        inventory_area_options=inventory_area_options,
        inventory_area_name_map=inventory_area_name_map,
        inventory_category_filter=inventory_category_filter,
        inventory_status_filter=inventory_status_filter,
        top_stock_value_items=top_stock_value_items,
        top_consumption_rows=top_consumption_rows,
        vendor_purchase_map=vendor_purchase_map,
        menu_item_meta=menu_item_meta,
        recipe_payload_map=recipe_payload_map,
        expense_logs=expense_logs,
        recent_expense_logs=recent_expense_logs,
        purchase_todos_active=purchase_todos_active,
        purchase_todos_history=purchase_todos_history,
        category_spend_rows=category_spend_rows,
        max_category_spend=max_category_spend,
        timeline_rows=timeline_rows,
        max_timeline_amount=max_timeline_amount,
        expense_vs_earning=expense_vs_earning,
        average_daily_expense=average_daily_expense,
    )


@bp.route("/stats")
@roles_required("admin", "manager")
def stats():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    category_options = _stats_category_options()
    return render_template(
        "cafe/stats.html",
        stats_payload=payload,
        filters=payload["filters"],
        category_options=category_options,
        sales_source_options=_stats_sales_source_options(),
    )


@bp.route("/stats/export")
@roles_required("admin", "manager")
def export_stats():
    payload = _build_stats_payload(_parse_stats_filters(request.args), use_cache=False)
    filters = payload["filters"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(
        [
            "Order",
            "Table",
            "Order Type",
            "Sales Source Mix",
            "Status",
            "Payment Type",
            "Subtotal",
            "Packaging",
            "Delivery",
            "Total",
            "Created At",
        ]
    )
    for order in payload["orders"]:
        ws.append(
            [
                order["code"],
                order["table"],
                order["order_type_label"],
                order["sales_source_mix"],
                order["status"],
                order["payment_type"],
                order["line_subtotal"],
                order["packaging_charge"],
                order["delivery_charge"],
                order["total_amount"],
                order["created_at"],
            ]
        )
    ws2 = wb.create_sheet(title="Top Items")
    ws2.append(["Rank", "Item", "Category", "Qty", "Revenue", "Avg Price", "Contribution %"])
    for row in payload["item_analytics"]["top_items"]:
        ws2.append(
            [
                row["rank"],
                row["name"],
                row["category"],
                row["qty"],
                row["revenue"],
                row["avg_price"],
                row["contribution_pct"],
            ]
        )
    ws3 = wb.create_sheet(title="Summary")
    ws3.append(["Period", payload["period"]["label"]])
    ws3.append(["Preset", filters["preset"]])
    ws3.append(["Sales Source", filters["salesSource"]])
    ws3.append(["Order Type", filters["orderType"]])
    ws3.append(["Category", filters["category"]])
    ws3.append([])
    kpi = payload["summary"]
    ws3.append(["Total Sales", kpi["revenue"]["total_sales"]])
    ws3.append(["Total Orders", kpi["revenue"]["total_orders"]])
    ws3.append(["Average Order Value", kpi["revenue"]["average_order_value"]])
    for row in kpi["operations"]["workstation_rows"]:
        ws3.append([f'{row["label"]} Sales', row["sales"]])
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f'attachment; filename="cafe_stats_{filters["preset"]}.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


@bp.route("/reviews")
@roles_required("admin")
def review_summary():
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    table_filter = request.args.get("table_id", type=int)
    today_ist = datetime.now(IST_TZ).date()
    default_start = today_ist - timedelta(days=29)
    start_value = _parse_date_only(start_date) or default_start
    end_value = _parse_date_only(end_date) or today_ist
    if end_value < start_value:
        start_value, end_value = end_value, start_value
    payload = _build_review_summary_payload(start_value, end_value, table_filter)
    return render_template(
        "cafe/review_summary.html",
        table_options=CafeTable.query.filter_by(active=True).order_by(CafeTable.name.asc()).all(),
        selected_table_id=table_filter,
        start_date=start_value.isoformat(),
        end_date=end_value.isoformat(),
        export_url=url_for(
            "cafe.review_summary_export",
            start_date=start_value.isoformat(),
            end_date=end_value.isoformat(),
            table_id=table_filter,
        ),
        **payload,
    )


def _build_review_summary_payload(start_value: date, end_value: date, table_filter: int | None):
    if end_value < start_value:
        start_value, end_value = end_value, start_value
    start_dt = _utc_naive_from_ist(datetime.combine(start_value, time.min))
    end_dt = _utc_naive_from_ist(datetime.combine(end_value + timedelta(days=1), time.min))
    query = CafeFeedback.query.options(
        joinedload(CafeFeedback.table),
        joinedload(CafeFeedback.primary_order),
        joinedload(CafeFeedback.items),
        joinedload(CafeFeedback.items).joinedload(CafeFeedbackItem.menu_item),
    ).filter(
        CafeFeedback.submitted_at.is_not(None),
        CafeFeedback.submitted_at >= start_dt,
        CafeFeedback.submitted_at < end_dt,
    )
    if table_filter:
        query = query.filter(CafeFeedback.table_id == table_filter)
    feedback_entries = query.order_by(CafeFeedback.submitted_at.desc(), CafeFeedback.id.desc()).all()
    total_feedback_count = len(feedback_entries)
    average_service_rating = round(
        sum(int(entry.service_rating or 0) for entry in feedback_entries) / total_feedback_count,
        2,
    ) if total_feedback_count else 0.0
    item_rollup = {}
    table_rollup = {}
    day_rollup = {}
    chef_rollup = {}
    recent_reviews = []
    all_reviews = []
    source_rollup = {}
    total_item_ratings = 0
    total_item_score = 0
    reviews_with_comments = 0
    low_service_count = 0
    item_issue_count = 0
    positive_feedback_count = 0
    for chef in _chef_options(include_inactive=True):
        chef_rollup[chef.id] = {
            "chef_name": chef.full_name,
            "assigned_items": 0,
            "sales_total": 0.0,
            "sold_qty": 0,
            "review_count": 0,
            "rating_sum": 0,
            "complaint_count": 0,
        }
    for item in MenuItem.query.filter(
        MenuItem.chef_user_id.is_not(None),
        MenuItem.is_deleted.is_(False),
    ).all():
        row = chef_rollup.setdefault(
            item.chef_user_id,
            {
                "chef_name": item.chef.full_name if item.chef else "Unknown Staff",
                "assigned_items": 0,
                "sales_total": 0.0,
                "sold_qty": 0,
                "review_count": 0,
                "rating_sum": 0,
                "complaint_count": 0,
            },
        )
        row["assigned_items"] += 1
    paid_order_items = CafeOrderItem.query.options(
        joinedload(CafeOrderItem.menu_item),
        joinedload(CafeOrderItem.order),
    ).join(CafeOrder).filter(
        CafeOrder.status == "paid",
        CafeOrder.created_at >= start_dt,
        CafeOrder.created_at < end_dt,
    ).all()
    for order_item in paid_order_items:
        menu_item = order_item.menu_item
        if not menu_item or not menu_item.chef_user_id:
            continue
        row = chef_rollup.setdefault(
            menu_item.chef_user_id,
            {
                "chef_name": menu_item.chef.full_name if menu_item.chef else "Unknown Staff",
                "assigned_items": 0,
                "sales_total": 0.0,
                "sold_qty": 0,
                "review_count": 0,
                "rating_sum": 0,
                "complaint_count": 0,
            },
        )
        quantity = int(order_item.quantity or 0)
        row["sold_qty"] += quantity
        row["sales_total"] += float(order_item.unit_price or 0) * quantity
    for entry in feedback_entries:
        entry_day = _format_ist(entry.submitted_at or entry.created_at, "%Y-%m-%d")
        day_info = day_rollup.setdefault(entry_day, {"count": 0, "service_sum": 0, "item_sum": 0, "item_count": 0})
        day_info["count"] += 1
        service_rating = int(entry.service_rating or 0)
        day_info["service_sum"] += service_rating
        table_name = entry.table.name if entry.table else "Unknown"
        table_info = table_rollup.setdefault(
            table_name,
            {"count": 0, "service_sum": 0, "item_sum": 0, "item_count": 0, "low_service_count": 0, "item_issue_count": 0},
        )
        table_info["count"] += 1
        table_info["service_sum"] += service_rating
        source_label = (entry.source or "online").replace("_", " ").title()
        source_rollup[source_label] = source_rollup.get(source_label, 0) + 1
        has_comment = bool((entry.summary_text or "").strip())
        if has_comment:
            reviews_with_comments += 1
        if service_rating <= 2:
            low_service_count += 1
            table_info["low_service_count"] += 1
        raw_order_ids = []
        if entry.order_ids_json:
            try:
                parsed_order_ids = json.loads(entry.order_ids_json)
                if isinstance(parsed_order_ids, list):
                    raw_order_ids = [int(order_id) for order_id in parsed_order_ids if str(order_id).isdigit()]
            except Exception:
                raw_order_ids = []
        if not raw_order_ids and entry.primary_order_id:
            raw_order_ids = [int(entry.primary_order_id)]
        entry_item_ratings = []
        entry_items_preview = []
        entry_item_details = []
        for item in entry.items:
            item_rating = int(item.rating or 0)
            total_item_ratings += 1
            total_item_score += item_rating
            entry_item_ratings.append(item_rating)
            entry_items_preview.append(item.item_name)
            entry_item_details.append(
                {
                    "name": item.item_name,
                    "rating": item_rating,
                    "size_label": item.size_label or "",
                    "is_parcel": bool(item.is_parcel),
                }
            )
            item_key = item.menu_item_id or item.item_name
            item_info = item_rollup.setdefault(
                item_key,
                {"name": item.item_name, "count": 0, "rating_sum": 0, "parcel_count": 0, "low_rating_count": 0},
            )
            item_info["count"] += 1
            item_info["rating_sum"] += item_rating
            if item.is_parcel:
                item_info["parcel_count"] += 1
            if item_rating <= 2:
                item_info["low_rating_count"] += 1
            if item.menu_item and item.menu_item.chef_user_id:
                chef_row = chef_rollup.setdefault(
                    item.menu_item.chef_user_id,
                    {
                        "chef_name": item.menu_item.chef.full_name if item.menu_item.chef else "Unknown Staff",
                        "assigned_items": 0,
                        "sales_total": 0.0,
                        "sold_qty": 0,
                        "review_count": 0,
                        "rating_sum": 0,
                        "complaint_count": 0,
                    },
                )
                chef_row["review_count"] += 1
                chef_row["rating_sum"] += item_rating
                if item_rating <= 2:
                    chef_row["complaint_count"] += 1
            day_info["item_sum"] += item_rating
            day_info["item_count"] += 1
            table_info["item_sum"] += item_rating
            table_info["item_count"] += 1
        if any(rating <= 2 for rating in entry_item_ratings):
            item_issue_count += 1
            table_info["item_issue_count"] += 1
        item_average = round(sum(entry_item_ratings) / len(entry_item_ratings), 2) if entry_item_ratings else 0.0
        if service_rating >= 4 and (not entry_item_ratings or item_average >= 4):
            positive_feedback_count += 1
        recent_reviews.append(
            {
                "id": entry.id,
                "table": table_name,
                "service_rating": service_rating,
                "item_average": item_average,
                "summary_text": (entry.summary_text or "").strip(),
                "source": source_label,
                "submitted_at": _format_ist(entry.submitted_at or entry.created_at, "%d %b %Y, %I:%M %p"),
                "items": ", ".join(item.item_name for item in entry.items[:4]),
            }
        )
        all_reviews.append(
            {
                "id": entry.id,
                "table": table_name,
                "service_rating": service_rating,
                "item_average": item_average,
                "summary_text": (entry.summary_text or "").strip(),
                "summary_excerpt": ((entry.summary_text or "").strip()[:120] + "…") if len((entry.summary_text or "").strip()) > 120 else (entry.summary_text or "").strip(),
                "source": source_label,
                "submitted_at": _format_ist(entry.submitted_at or entry.created_at, "%d %b %Y, %I:%M %p"),
                "items_preview": ", ".join(entry_items_preview[:4]),
                "item_count": len(entry_item_details),
                "has_comment": has_comment,
                "order_count": len(raw_order_ids),
                "submitted_by_name": (entry.submitted_by_name or (entry.submitted_by_user.full_name if entry.submitted_by_user else "") or "-"),
            }
        )
    avg_item_rating = round(total_item_score / total_item_ratings, 2) if total_item_ratings else 0.0
    item_rows = sorted(
        [
            {
                "name": row["name"],
                "count": row["count"],
                "average_rating": round(row["rating_sum"] / row["count"], 2) if row["count"] else 0.0,
                "parcel_count": row["parcel_count"],
                "low_rating_count": row["low_rating_count"],
            }
            for row in item_rollup.values()
        ],
        key=lambda row: (-row["average_rating"], -row["count"], row["name"].lower()),
    )
    table_rows = sorted(
        [
            {
                "name": name,
                "count": row["count"],
                "service_average": round(row["service_sum"] / row["count"], 2) if row["count"] else 0.0,
                "item_average": round(row["item_sum"] / row["item_count"], 2) if row["item_count"] else 0.0,
                "low_service_count": row["low_service_count"],
                "item_issue_count": row["item_issue_count"],
            }
            for name, row in table_rollup.items()
        ],
        key=lambda row: (-row["count"], row["name"].lower()),
    )
    day_rows = sorted(
        [
            {
                "day": day,
                "count": row["count"],
                "service_average": round(row["service_sum"] / row["count"], 2) if row["count"] else 0.0,
                "item_average": round(row["item_sum"] / row["item_count"], 2) if row["item_count"] else 0.0,
            }
            for day, row in day_rollup.items()
        ],
        key=lambda row: row["day"],
        reverse=True,
    )
    chef_rows = sorted(
        [
            {
                "chef_name": row["chef_name"],
                "assigned_items": row["assigned_items"],
                "sales_total": round(row["sales_total"], 2),
                "sold_qty": row["sold_qty"],
                "review_count": row["review_count"],
                "average_rating": round(row["rating_sum"] / row["review_count"], 2) if row["review_count"] else 0.0,
                "complaint_count": row["complaint_count"],
            }
            for row in chef_rollup.values()
            if row["assigned_items"] or row["sold_qty"] or row["review_count"]
        ],
        key=lambda row: (-row["sales_total"], row["chef_name"].lower()),
    )
    source_rows = sorted(
        [{"source": source, "count": count} for source, count in source_rollup.items()],
        key=lambda row: (-row["count"], row["source"].lower()),
    )
    attention_item_rows = sorted(
        [row for row in item_rows if row["low_rating_count"] or row["average_rating"] <= 3],
        key=lambda row: (-row["low_rating_count"], row["average_rating"], -row["count"], row["name"].lower()),
    )
    comments_rate = round((reviews_with_comments / total_feedback_count) * 100, 1) if total_feedback_count else 0.0
    positive_rate = round((positive_feedback_count / total_feedback_count) * 100, 1) if total_feedback_count else 0.0
    attention_table = None
    if table_rows:
        attention_table = sorted(
            table_rows,
            key=lambda row: (-row["low_service_count"], -row["item_issue_count"], row["service_average"], -row["count"], row["name"].lower()),
        )[0]
    best_table = None
    if table_rows:
        best_table = sorted(
            table_rows,
            key=lambda row: (-row["service_average"], -row["item_average"], -row["count"], row["name"].lower()),
        )[0]
    top_issue_item = attention_item_rows[0] if attention_item_rows else None
    return {
        "feedback_entries": feedback_entries,
        "total_feedback_count": total_feedback_count,
        "average_service_rating": average_service_rating,
        "average_item_rating": avg_item_rating,
        "low_service_count": low_service_count,
        "item_issue_count": item_issue_count,
        "reviews_with_comments": reviews_with_comments,
        "comments_rate": comments_rate,
        "positive_feedback_count": positive_feedback_count,
        "positive_rate": positive_rate,
        "item_rows": item_rows,
        "attention_item_rows": attention_item_rows,
        "table_rows": table_rows,
        "day_rows": day_rows,
        "chef_rows": chef_rows,
        "recent_reviews": recent_reviews[:20],
        "all_reviews": all_reviews,
        "source_rows": source_rows,
        "top_issue_item": top_issue_item,
        "attention_table": attention_table,
        "best_table": best_table,
    }


@bp.route("/reviews/export")
@roles_required("admin")
def review_summary_export():
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    table_filter = request.args.get("table_id", type=int)
    today_ist = datetime.now(IST_TZ).date()
    default_start = today_ist - timedelta(days=29)
    start_value = _parse_date_only(start_date) or default_start
    end_value = _parse_date_only(end_date) or today_ist
    if end_value < start_value:
        start_value, end_value = end_value, start_value
    payload = _build_review_summary_payload(start_value, end_value, table_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Review Dashboard Export"])
    ws.append(["Start Date", start_value.isoformat()])
    ws.append(["End Date", end_value.isoformat()])
    ws.append(["Table Filter", next((table.name for table in CafeTable.query.filter_by(active=True).all() if table.id == table_filter), "All Tables") if table_filter else "All Tables"])
    ws.append([])
    ws.append(["Total Feedback", payload["total_feedback_count"]])
    ws.append(["Average Service Rating", payload["average_service_rating"]])
    ws.append(["Average Item Rating", payload["average_item_rating"]])
    ws.append(["Low Service Alerts", payload["low_service_count"]])
    ws.append(["Item Issue Alerts", payload["item_issue_count"]])
    ws.append(["Reviews With Comments", payload["reviews_with_comments"]])
    ws.append(["Positive Review Rate %", payload["positive_rate"]])

    ws_reviews = wb.create_sheet(title="Reviews")
    ws_reviews.append(["When", "Table", "Source", "Submitted By", "Service Rating", "Item Average", "Order Count", "Items", "Summary"])
    for row in payload["all_reviews"]:
        ws_reviews.append([
            row["submitted_at"],
            row["table"],
            row["source"],
            row["submitted_by_name"],
            row["service_rating"],
            row["item_average"],
            row["order_count"],
            row["items_preview"],
            row["summary_text"],
        ])

    ws_items = wb.create_sheet(title="Items")
    ws_items.append(["Item", "Average Rating", "Reviews", "Parcel Count", "Low Rating Count"])
    for row in payload["item_rows"]:
        ws_items.append([row["name"], row["average_rating"], row["count"], row["parcel_count"], row["low_rating_count"]])

    ws_tables = wb.create_sheet(title="Tables")
    ws_tables.append(["Table", "Feedback Count", "Service Average", "Item Average", "Low Service Count", "Item Issue Count"])
    for row in payload["table_rows"]:
        ws_tables.append([row["name"], row["count"], row["service_average"], row["item_average"], row["low_service_count"], row["item_issue_count"]])

    ws_days = wb.create_sheet(title="Days")
    ws_days.append(["Day", "Feedback Count", "Service Average", "Item Average"])
    for row in payload["day_rows"]:
        ws_days.append([row["day"], row["count"], row["service_average"], row["item_average"]])

    ws_chefs = wb.create_sheet(title="Preparation Staff")
    ws_chefs.append(["Responsible Staff", "Assigned Items", "Paid Sales", "Qty Sold", "Review Count", "Average Rating", "Complaints"])
    for row in payload["chef_rows"]:
        ws_chefs.append([row["chef_name"], row["assigned_items"], row["sales_total"], row["sold_qty"], row["review_count"], row["average_rating"], row["complaint_count"]])

    ws_sources = wb.create_sheet(title="Sources")
    ws_sources.append(["Source", "Count"])
    for row in payload["source_rows"]:
        ws_sources.append([row["source"], row["count"]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename_suffix = f"{start_value.isoformat()}_{end_value.isoformat()}"
    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f'attachment; filename="cafe_reviews_{filename_suffix}.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


@bp.route("/reviews/<int:feedback_id>")
@roles_required("admin")
def review_summary_detail(feedback_id):
    entry = CafeFeedback.query.options(
        joinedload(CafeFeedback.table),
        joinedload(CafeFeedback.primary_order),
        joinedload(CafeFeedback.submitted_by_user),
        joinedload(CafeFeedback.items).joinedload(CafeFeedbackItem.menu_item),
    ).filter(CafeFeedback.id == feedback_id).first()
    if not entry:
        return jsonify({"ok": False, "error": "Review not found."}), 404
    order_ids = []
    if entry.order_ids_json:
        try:
            parsed = json.loads(entry.order_ids_json)
            if isinstance(parsed, list):
                order_ids = [str(order_id) for order_id in parsed]
        except Exception:
            order_ids = []
    if not order_ids and entry.primary_order_id:
        order_ids = [str(entry.primary_order_id)]
    item_rows = []
    item_ratings = []
    for item in entry.items:
        rating = int(item.rating or 0)
        item_ratings.append(rating)
        item_rows.append(
            {
                "name": item.item_name,
                "rating": rating,
                "size_label": item.size_label or "",
                "is_parcel": bool(item.is_parcel),
            }
        )
    item_average = round(sum(item_ratings) / len(item_ratings), 2) if item_ratings else 0.0
    return jsonify(
        {
            "ok": True,
            "review": {
                "id": entry.id,
                "table": entry.table.name if entry.table else "Unknown",
                "submitted_at": _format_ist(entry.submitted_at or entry.created_at, "%d %b %Y, %I:%M %p"),
                "source": (entry.source or "online").replace("_", " ").title(),
                "submitted_by_name": (entry.submitted_by_name or (entry.submitted_by_user.full_name if entry.submitted_by_user else "") or "-"),
                "service_rating": int(entry.service_rating or 0),
                "item_average": item_average,
                "summary_text": (entry.summary_text or "").strip(),
                "order_ids": order_ids,
                "primary_order_reference": (
                    (entry.primary_order.display_code or _format_internal_order_code(entry.primary_order))
                    if entry.primary_order
                    else ""
                ),
                "items": item_rows,
            },
        }
    )


_STATS_CACHE = {}
_STATS_CACHE_TTL_SECONDS = 120
_STAT_FILTER_PRESETS = {
    "today": "Today",
    "yesterday": "Yesterday",
    "last_7_days": "Last 7 Days",
    "last_30_days": "Last 30 Days",
    "this_month": "This Month",
    "last_month": "Last Month",
    "this_quarter": "This Quarter",
    "last_quarter": "Last Quarter",
    "this_year": "This Year",
    "custom": "Custom Date Range",
}


def _parse_date_only(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value.strip())
    except (TypeError, ValueError, AttributeError):
        return None


def _parse_datetime_flexible(value):
    if not value:
        return None
    raw = value.strip()
    for parser in (datetime.fromisoformat,):
        try:
            return parser(raw)
        except ValueError:
            continue
    parsed_date = _parse_date_only(raw)
    if parsed_date:
        return datetime.combine(parsed_date, time.min)
    return None


def _quarter_start(d: date):
    q = ((d.month - 1) // 3) * 3 + 1
    return date(d.year, q, 1)


def _month_start(d: date):
    return date(d.year, d.month, 1)


def _next_month_start(d: date):
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _resolve_period_from_filters(filters):
    today = datetime.now(IST_TZ).date()
    preset = filters["preset"]
    if preset == "yesterday":
        start = datetime.combine(today - timedelta(days=1), time.min)
        end = start + timedelta(days=1)
        label = "Yesterday"
        granularity = "hour"
    elif preset == "last_7_days":
        start = datetime.combine(today - timedelta(days=6), time.min)
        end = datetime.combine(today + timedelta(days=1), time.min)
        label = "Last 7 Days"
        granularity = "day"
    elif preset == "last_30_days":
        start = datetime.combine(today - timedelta(days=29), time.min)
        end = datetime.combine(today + timedelta(days=1), time.min)
        label = "Last 30 Days"
        granularity = "day"
    elif preset == "this_month":
        month_start = _month_start(today)
        start = datetime.combine(month_start, time.min)
        end = datetime.combine(today + timedelta(days=1), time.min)
        label = today.strftime("%B %Y")
        granularity = "day"
    elif preset == "last_month":
        this_month_start = _month_start(today)
        prev_month_end = this_month_start - timedelta(days=1)
        prev_month_start = _month_start(prev_month_end)
        start = datetime.combine(prev_month_start, time.min)
        end = datetime.combine(this_month_start, time.min)
        label = prev_month_start.strftime("%B %Y")
        granularity = "day"
    elif preset == "this_quarter":
        qstart = _quarter_start(today)
        start = datetime.combine(qstart, time.min)
        end = datetime.combine(today + timedelta(days=1), time.min)
        quarter_no = ((today.month - 1) // 3) + 1
        label = f"Q{quarter_no} {today.year}"
        granularity = "day"
    elif preset == "last_quarter":
        this_q_start = _quarter_start(today)
        prev_q_end = this_q_start - timedelta(days=1)
        prev_q_start = _quarter_start(prev_q_end)
        start = datetime.combine(prev_q_start, time.min)
        end = datetime.combine(this_q_start, time.min)
        quarter_no = ((prev_q_start.month - 1) // 3) + 1
        label = f"Q{quarter_no} {prev_q_start.year}"
        granularity = "day"
    elif preset == "this_year":
        year_start = date(today.year, 1, 1)
        start = datetime.combine(year_start, time.min)
        end = datetime.combine(today + timedelta(days=1), time.min)
        label = str(today.year)
        granularity = "month"
    elif preset == "custom":
        start_date = _parse_date_only(filters.get("startDate"))
        end_date = _parse_date_only(filters.get("endDate"))
        if not start_date:
            start_date = today
        if not end_date:
            end_date = start_date
        if end_date < start_date:
            start_date, end_date = end_date, start_date
        start = datetime.combine(start_date, time.min)
        end = datetime.combine(end_date + timedelta(days=1), time.min)
        label = f"{start_date.isoformat()} to {end_date.isoformat()}"
        days = (end - start).days
        granularity = "hour" if days <= 1 else ("day" if days <= 92 else "month")
    else:
        start = datetime.combine(today, time.min)
        end = start + timedelta(days=1)
        label = "Today"
        granularity = "hour"
    return start, end, label, granularity


def _parse_stats_filters(args):
    mode = (args.get("mode") or "").strip().lower()
    day_str = (args.get("day") or "").strip()
    from_str = (args.get("from") or "").strip()
    to_str = (args.get("to") or "").strip()
    preset = (args.get("preset") or "").strip().lower()
    if not preset and mode in ["day", "week", "month", "range"]:
        if mode == "day":
            preset = "custom" if day_str else "today"
            if day_str:
                args = dict(args)
                args["startDate"] = day_str
                args["endDate"] = day_str
        elif mode == "week":
            preset = "last_7_days"
        elif mode == "month":
            preset = "this_month"
        else:
            preset = "custom"
            if from_str:
                parsed = _parse_datetime_flexible(from_str)
                if parsed:
                    args = dict(args)
                    args["startDate"] = parsed.date().isoformat()
            if to_str:
                parsed = _parse_datetime_flexible(to_str)
                if parsed:
                    args = dict(args)
                    args["endDate"] = parsed.date().isoformat()
    if preset not in _STAT_FILTER_PRESETS:
        preset = "today"
    sales_source = (args.get("salesSource") or "all").strip().lower()
    if sales_source not in {row["value"] for row in _stats_sales_source_options()}:
        sales_source = "all"
    order_type = (args.get("orderType") or "all").strip().lower()
    if order_type not in ["all", "dine_in", "takeaway", "online_order"]:
        order_type = "all"
    category = (args.get("category") or "all").strip()
    return {
        "preset": preset,
        "startDate": (args.get("startDate") or "").strip(),
        "endDate": (args.get("endDate") or "").strip(),
        "salesSource": sales_source,
        "orderType": order_type,
        "category": category or "all",
    }


def _stats_category_options():
    defaults = ["Coffee", "Tea", "Shake", "Mocktail", "Pizza", "Pasta", "Noodles", "Snacks", "Other"]
    from_db = [c.name for c in MenuCategory.query.order_by(MenuCategory.name.asc()).all() if c.name]
    out = []
    seen = set()
    for name in (defaults + from_db):
        key = name.strip().lower()
        if key and key not in seen:
            out.append(name.strip())
            seen.add(key)
    return out


def _menu_item_category_names_for_stats(item: MenuItem, category_map: dict[int, str]):
    names = _get_item_category_names(item, category_map)
    return names or ["Other"]


def _order_type_key(order: CafeOrder, approved_items):
    if order.is_delivery:
        return "online_order"
    if approved_items and all(bool(x.is_parcel) for x in approved_items):
        return "takeaway"
    return "dine_in"


def _order_type_label(key):
    return {
        "dine_in": "Dine In",
        "takeaway": "Takeaway",
        "online_order": "Online Order",
    }.get(key, "Dine In")


def _growth(curr, prev):
    if prev == 0:
        if curr == 0:
            return 0.0
        return 100.0
    return round(((curr - prev) / prev) * 100.0, 2)


def _bucket_key(dt: datetime, granularity: str):
    if granularity == "hour":
        return dt.replace(minute=0, second=0, microsecond=0)
    if granularity == "month":
        return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return dt.replace(hour=0, minute=0, second=0, microsecond=0)


def _bucket_label(dt: datetime, granularity: str):
    if granularity == "hour":
        return dt.strftime("%H:00")
    if granularity == "month":
        return dt.strftime("%b %Y")
    return dt.strftime("%d %b")


def _build_stats_payload(filters, use_cache=True):
    start_dt, end_dt, period_label, granularity = _resolve_period_from_filters(filters)
    query_start_dt = _utc_naive_from_ist(start_dt)
    query_end_dt = _utc_naive_from_ist(end_dt)
    prev_start = start_dt - (end_dt - start_dt)
    prev_end = start_dt
    prev_query_start = _utc_naive_from_ist(prev_start)
    prev_query_end = _utc_naive_from_ist(prev_end)
    cache_key = (
        start_dt.isoformat(),
        end_dt.isoformat(),
        filters["salesSource"],
        filters["orderType"],
        filters["category"].strip().lower(),
    )
    now_ts = datetime.utcnow().timestamp()
    if use_cache:
        existing = _STATS_CACHE.get(cache_key)
        if existing and (now_ts - existing["ts"] < _STATS_CACHE_TTL_SECONDS):
            return existing["payload"]

    category_rows = MenuCategory.query.order_by(MenuCategory.name.asc()).all()
    category_map = {c.id: c.name for c in category_rows}
    selected_category = (filters["category"] or "all").strip().lower()
    sales_source = filters["salesSource"]
    order_type_filter = filters["orderType"]
    station_registry = _stats_station_registry(include_unassigned=True)
    station_label_map = {row["slug"]: row["label"] for row in station_registry}
    station_color_map = {row["slug"]: row["color"] for row in station_registry}
    station_slugs = [row["slug"] for row in station_registry]

    orders = (
        CafeOrder.query.options(
            joinedload(CafeOrder.table),
            joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item),
        )
        .filter(
            CafeOrder.status == "paid",
            CafeOrder.paid_at.is_not(None),
            CafeOrder.paid_at >= query_start_dt,
            CafeOrder.paid_at < query_end_dt,
        )
        .order_by(CafeOrder.paid_at.asc(), CafeOrder.created_at.asc())
        .all()
    )
    prev_orders = (
        CafeOrder.query.options(joinedload(CafeOrder.order_items).joinedload(CafeOrderItem.menu_item))
        .filter(
            CafeOrder.status == "paid",
            CafeOrder.paid_at.is_not(None),
            CafeOrder.paid_at >= prev_query_start,
            CafeOrder.paid_at < prev_query_end,
        )
        .all()
    )

    def _filtered_rows(order_rows):
        filtered = []
        for order in order_rows:
            approved_items = [oi for oi in order.order_items if (oi.approval_status or "pending") != "rejected" and oi.menu_item]
            if not approved_items:
                continue
            otype = _order_type_key(order, approved_items)
            if order_type_filter != "all" and otype != order_type_filter:
                continue
            all_line_subtotal = round(sum(float(oi.unit_price or 0) * int(oi.quantity or 0) for oi in approved_items), 2)
            matched_items = []
            for oi in approved_items:
                station_slug = _station_slug_for_stats(oi.menu_item.prep_station)
                if sales_source != "all" and station_slug != sales_source:
                    continue
                categories = _menu_item_category_names_for_stats(oi.menu_item, category_map)
                if selected_category != "all" and selected_category not in [c.lower() for c in categories]:
                    continue
                matched_items.append((oi, categories, station_slug))
            if not matched_items:
                continue
            matched_subtotal = round(sum(float(oi.unit_price or 0) * int(oi.quantity or 0) for oi, _categories, _station_slug in matched_items), 2)
            extra = float(order.packaging_charge or 0) + float(order.delivery_charge or 0)
            ratio = (matched_subtotal / all_line_subtotal) if all_line_subtotal > 0 else 1.0
            matched_total = round(matched_subtotal + (extra * ratio), 2)
            filtered.append(
                {
                    "order": order,
                    "order_type": otype,
                    "approved_items": approved_items,
                    "matched_items": matched_items,
                    "matched_subtotal": matched_subtotal,
                    "matched_total": matched_total,
                    "ratio": ratio,
                }
            )
        return filtered

    filtered_orders = _filtered_rows(orders)
    filtered_prev_orders = _filtered_rows(prev_orders)

    total_sales = round(sum(row["matched_total"] for row in filtered_orders), 2)
    total_orders = len(filtered_orders)
    avg_order_value = round(total_sales / total_orders, 2) if total_orders else 0.0
    highest_order = round(max((row["matched_total"] for row in filtered_orders), default=0), 2)
    lowest_order = round(min((row["matched_total"] for row in filtered_orders), default=0), 2)
    prev_total_sales = round(sum(row["matched_total"] for row in filtered_prev_orders), 2)
    prev_total_orders = len(filtered_prev_orders)
    prev_avg = round(prev_total_sales / prev_total_orders, 2) if prev_total_orders else 0.0

    station_sales = {slug: 0.0 for slug in station_slugs}
    station_order_counts = {slug: 0 for slug in station_slugs}
    station_items = {slug: {} for slug in station_slugs}

    total_items_sold = 0
    unique_item_ids = set()
    item_stats = {}
    orders_payload = []
    order_type_counts = {"dine_in": 0, "takeaway": 0, "online_order": 0}
    bucket_sales = {}
    bucket_orders = {}
    station_trend_buckets = {slug: {} for slug in station_slugs}
    peak_hour_orders = {h: 0 for h in range(24)}
    peak_hour_sales = {h: 0.0 for h in range(24)}

    category_stats = {}
    category_order_ids = {}
    customers = set()
    mobile_customers = set()

    for row in filtered_orders:
        order = row["order"]
        revenue_dt = order.paid_at or order.created_at
        order_local_dt = _ist_from_utc_naive(revenue_dt)
        order_type_counts[row["order_type"]] += 1
        bucket = _bucket_key(order_local_dt.replace(tzinfo=None), granularity)
        bucket_sales[bucket] = round(bucket_sales.get(bucket, 0.0) + row["matched_total"], 2)
        bucket_orders[bucket] = bucket_orders.get(bucket, 0) + 1
        peak_hour_orders[order_local_dt.hour] += 1
        peak_hour_sales[order_local_dt.hour] = round(peak_hour_sales[order_local_dt.hour] + row["matched_total"], 2)
        if order.is_delivery and (order.delivery_customer_mobile or "").strip():
            key = f"m:{order.delivery_customer_mobile.strip()}"
            customers.add(key)
            mobile_customers.add(order.delivery_customer_mobile.strip())
        else:
            customers.add(f"o:{order.id}")

        order_station_mix = {slug: 0.0 for slug in station_slugs}
        for oi, categories, station_slug in row["matched_items"]:
            qty = int(oi.quantity or 0)
            line_amount = round(float(oi.unit_price or 0) * qty, 2)
            total_items_sold += qty
            unique_item_ids.add(oi.menu_item_id)

            stat = item_stats.setdefault(
                oi.menu_item_id,
                {
                    "id": oi.menu_item_id,
                    "name": oi.menu_item.name,
                    "category": ", ".join(categories[:2]),
                    "qty": 0,
                    "revenue": 0.0,
                    "order_ids": set(),
                    "station": station_slug,
                },
            )
            stat["qty"] += qty
            stat["revenue"] = round(stat["revenue"] + line_amount, 2)
            stat["order_ids"].add(order.id)

            order_station_mix[station_slug] = round(order_station_mix.get(station_slug, 0.0) + line_amount, 2)
            st_item = station_items[station_slug].setdefault(
                oi.menu_item_id,
                {"name": oi.menu_item.name, "qty": 0, "revenue": 0.0},
            )
            st_item["qty"] += qty
            st_item["revenue"] = round(st_item["revenue"] + line_amount, 2)

            cats_for_agg = categories or ["Other"]
            split = 1.0 / len(cats_for_agg)
            for cname in cats_for_agg:
                c = category_stats.setdefault(cname, {"revenue": 0.0, "qty": 0.0})
                c["revenue"] = round(c["revenue"] + line_amount * split, 2)
                c["qty"] = round(c["qty"] + qty * split, 2)
                category_order_ids.setdefault(cname, set()).add(order.id)

        order_mix_parts = []
        for station in station_slugs:
            val = order_station_mix[station]
            if val > 0:
                order_mix_parts.append(f"{station_label_map.get(station, _workstation_display_name(station))} ₹{val:.2f}")
                station_order_counts[station] += 1
        mix_label = ", ".join(order_mix_parts) if order_mix_parts else "-"

        line_subtotal = round(sum(float(oi.unit_price or 0) * int(oi.quantity or 0) for oi, _categories, _station_slug in row["matched_items"]), 2)
        ratio = (line_subtotal / row["matched_subtotal"]) if row["matched_subtotal"] > 0 else 0
        packaging_m = round(float(order.packaging_charge or 0) * row["ratio"], 2)
        delivery_m = round(float(order.delivery_charge or 0) * row["ratio"], 2)
        orders_payload.append(
            {
                "id": order.id,
                "code": order.display_code or _format_internal_order_code(order),
                "pickup_no": _format_pickup_number(order),
                "table": "For Delivery" if order.is_delivery else (order.table.name if order.table else "-"),
                "order_type": row["order_type"],
                "order_type_label": _order_type_label(row["order_type"]),
                "sales_source_mix": mix_label,
                "status": order.status,
                "payment_type": order.payment_type or "-",
                "line_subtotal": line_subtotal,
                "packaging_charge": packaging_m,
                "delivery_charge": delivery_m,
                "total_amount": row["matched_total"],
                "settled_at": _format_ist(revenue_dt),
                "items": [
                    {
                        "name": oi.menu_item.name,
                        "qty": int(oi.quantity or 0),
                        "unit_price": float(oi.unit_price or 0),
                        "size_label": oi.size_label or "",
                        "is_parcel": bool(oi.is_parcel),
                    }
                    for oi, _, _station_slug in row["matched_items"]
                ],
            }
        )

        total_line_revenue_for_alloc = sum(order_station_mix.values())
        if total_line_revenue_for_alloc <= 0:
            total_line_revenue_for_alloc = 1.0
        extra = row["matched_total"] - row["matched_subtotal"]
        for station in station_slugs:
            station_line = order_station_mix[station]
            alloc = extra * (station_line / total_line_revenue_for_alloc)
            station_total = station_line + alloc
            station_sales[station] = round(station_sales.get(station, 0.0) + station_total, 2)
            station_trend_buckets[station][bucket] = round(
                station_trend_buckets[station].get(bucket, 0.0) + station_total,
                2,
            )

    prev_station_sales = {slug: 0.0 for slug in station_slugs}
    for row in filtered_prev_orders:
        order_station_mix = {slug: 0.0 for slug in station_slugs}
        for oi, _categories, station_slug in row["matched_items"]:
            order_station_mix[station_slug] = round(
                order_station_mix[station_slug] + float(oi.unit_price or 0) * int(oi.quantity or 0),
                2,
            )
        total_line_revenue = sum(order_station_mix.values()) or 1.0
        extra = row["matched_total"] - row["matched_subtotal"]
        for station in station_slugs:
            prev_station_sales[station] = round(
                prev_station_sales[station]
                + order_station_mix[station]
                + (extra * (order_station_mix[station] / total_line_revenue)),
                2,
            )

    recurring_mobile_set = set()
    if mobile_customers:
        prior_mobile_rows = (
            db.session.query(CafeOrder.delivery_customer_mobile)
            .filter(
                CafeOrder.delivery_customer_mobile.in_(list(mobile_customers)),
                CafeOrder.status == "paid",
                CafeOrder.paid_at.is_not(None),
                CafeOrder.paid_at < query_start_dt,
            )
            .all()
        )
        recurring_mobile_set = {row[0].strip() for row in prior_mobile_rows if row and row[0]}
    returning = len(recurring_mobile_set)
    new_customers = max(len(customers) - returning, 0)

    item_rows = []
    for _, s in item_stats.items():
        avg_price = round((s["revenue"] / s["qty"]), 2) if s["qty"] else 0.0
        contribution = round((s["revenue"] / total_sales) * 100, 2) if total_sales else 0.0
        item_rows.append(
            {
                "id": s["id"],
                "name": s["name"],
                "category": s["category"] or "Other",
                "qty": int(s["qty"]),
                "revenue": round(float(s["revenue"]), 2),
                "avg_price": avg_price,
                "contribution_pct": contribution,
                "station": s["station"],
                "order_count": len(s["order_ids"]),
            }
        )
    top_by_qty = sorted(item_rows, key=lambda x: (-x["qty"], -x["revenue"], x["name"].lower()))
    for idx, row in enumerate(top_by_qty, start=1):
        row["rank"] = idx
    top_items = top_by_qty[:15]
    bottom_items = sorted(item_rows, key=lambda x: (x["qty"], x["revenue"], x["name"].lower()))[:10]
    menu_q = MenuItem.query.filter(MenuItem.available.is_(True), MenuItem.is_deleted.is_(False))
    if sales_source == "unassigned":
        menu_q = menu_q.filter(db.or_(MenuItem.prep_station.is_(None), MenuItem.prep_station == ""))
    elif sales_source != "all":
        menu_q = menu_q.filter(MenuItem.prep_station == sales_source)
    menu_all = menu_q.all()
    zero_sales = []
    sold_ids = {x["id"] for x in item_rows}
    for item in menu_all:
        if item.id in sold_ids:
            continue
        c_names = _menu_item_category_names_for_stats(item, category_map)
        if selected_category != "all" and selected_category not in [c.lower() for c in c_names]:
            continue
        zero_sales.append({"id": item.id, "name": item.name, "category": ", ".join(c_names[:2])})
    zero_sales = sorted(zero_sales, key=lambda x: x["name"].lower())[:20]

    sorted_buckets = sorted(bucket_sales.keys())
    revenue_trend_labels = [_bucket_label(b, granularity) for b in sorted_buckets]
    revenue_trend_values = [round(bucket_sales.get(b, 0.0), 2) for b in sorted_buckets]
    order_trend_values = [int(bucket_orders.get(b, 0)) for b in sorted_buckets]
    station_trends = {
        slug: [round(station_trend_buckets[slug].get(b, 0.0), 2) for b in sorted_buckets]
        for slug in station_slugs
    }
    station_contribution_datasets = []
    for station in station_slugs:
        values = []
        trend_values = station_trends[station]
        for index in range(len(sorted_buckets)):
            total_bucket = sum(station_trends[slug][index] for slug in station_slugs)
            values.append(round((trend_values[index] / total_bucket) * 100, 2) if total_bucket > 0 else 0.0)
        station_contribution_datasets.append(
            {
                "slug": station,
                "label": station_label_map.get(station, _workstation_display_name(station)),
                "values": values,
                "color": station_color_map.get(station, "#6f4a35"),
            }
        )

    peak_hours = [
        {
            "hour": f"{h:02d}:00-{(h + 1) % 24:02d}:00",
            "orders": int(peak_hour_orders[h]),
            "sales": round(peak_hour_sales[h], 2),
        }
        for h in range(24)
    ]

    def _best_day_for_range(day_start_dt, day_end_dt):
        day_start_query = _utc_naive_from_ist(day_start_dt)
        day_end_query = _utc_naive_from_ist(day_end_dt)
        rows = (
            db.session.query(
                db.func.date(CafeOrder.paid_at).label("day"),
                db.func.count(CafeOrder.id),
                db.func.coalesce(db.func.sum(CafeOrder.total_amount), 0.0),
            )
            .filter(
                CafeOrder.status == "paid",
                CafeOrder.paid_at.is_not(None),
                CafeOrder.paid_at >= day_start_query,
                CafeOrder.paid_at < day_end_query,
            )
            .group_by(db.func.date(CafeOrder.paid_at))
            .all()
        )
        if not rows:
            return {"day": "-", "revenue": 0.0, "orders": 0}
        remapped = []
        for row_day, row_count, row_revenue in rows:
            parsed_day = datetime.fromisoformat(str(row_day))
            remapped.append(
                {
                    "day": _format_ist(parsed_day, "%Y-%m-%d"),
                    "revenue": round(float(row_revenue or 0), 2),
                    "orders": int(row_count or 0),
                }
            )
        best = max(remapped, key=lambda x: x["revenue"])
        return best

    today = datetime.now(IST_TZ).date()
    month_start = datetime.combine(_month_start(today), time.min)
    month_end = datetime.combine(_next_month_start(today), time.min)
    prev_month_start_date = _month_start(_month_start(today) - timedelta(days=1))
    prev_month_start = datetime.combine(prev_month_start_date, time.min)
    prev_month_end = datetime.combine(_month_start(today), time.min)
    best_day_ever = _best_day_for_range(datetime(2020, 1, 1), datetime.combine(today + timedelta(days=1), time.min))
    best_day_this_month = _best_day_for_range(month_start, month_end)
    best_day_last_month = _best_day_for_range(prev_month_start, prev_month_end)

    category_rows_out = []
    for name, data in sorted(category_stats.items(), key=lambda kv: kv[1]["revenue"], reverse=True):
        qty = float(data["qty"])
        rev = float(data["revenue"])
        category_rows_out.append(
            {
                "name": name,
                "revenue": round(rev, 2),
                "orders": len(category_order_ids.get(name, set())),
                "qty": round(qty, 2),
                "avg_price": round(rev / qty, 2) if qty else 0.0,
            }
        )

    def _best_station_item(station):
        rows = list(station_items[station].values())
        if not rows:
            return {"name": "-", "qty": 0, "revenue": 0.0}
        best = max(rows, key=lambda x: x["revenue"])
        return {"name": best["name"], "qty": int(best["qty"]), "revenue": round(best["revenue"], 2)}

    recipes = (
        InventoryRecipe.query.options(joinedload(InventoryRecipe.ingredients).joinedload(InventoryRecipeItem.inventory_item))
        .filter(InventoryRecipe.active.is_(True))
        .all()
    )
    recipe_map = {r.menu_item_id: r for r in recipes}
    ingredient_usage = {}
    for row in filtered_orders:
        for oi, _categories, _station_slug in row["matched_items"]:
            recipe = recipe_map.get(oi.menu_item_id)
            if not recipe:
                continue
            for ing in recipe.ingredients:
                if not ing.inventory_item:
                    continue
                key = ing.inventory_item.name
                total_qty = float(ing.qty_per_menu or 0) * int(oi.quantity or 0)
                u = ingredient_usage.setdefault(
                    key,
                    {
                        "ingredient": key,
                        "unit": ing.unit or ing.inventory_item.unit or "unit",
                        "consumption": 0.0,
                    },
                )
                u["consumption"] = round(u["consumption"] + total_qty, 3)
    inventory_consumption_rows = sorted(ingredient_usage.values(), key=lambda x: x["consumption"], reverse=True)

    item_contrib_rows = sorted(item_rows, key=lambda x: x["revenue"], reverse=True)
    top10 = item_contrib_rows[:10]
    top10_sum = sum(x["revenue"] for x in top10)
    others_sum = round(max(total_sales - top10_sum, 0), 2)

    workstation_rows = []
    for station in station_slugs:
        sales = round(station_sales.get(station, 0.0), 2)
        contribution_pct = round((sales / total_sales) * 100, 2) if total_sales else 0.0
        order_count = int(station_order_counts.get(station, 0))
        workstation_rows.append(
            {
                "slug": station,
                "label": station_label_map.get(station, _workstation_display_name(station)),
                "sales": sales,
                "contribution_pct": contribution_pct,
                "order_count": order_count,
                "best_item": _best_station_item(station),
                "average_ticket_size": round(sales / order_count, 2) if order_count else 0.0,
                "growth_pct": _growth(sales, prev_station_sales.get(station, 0.0)),
                "color": station_color_map.get(station, "#6f4a35"),
            }
        )

    summary = {
        "revenue": {
            "total_sales": round(total_sales, 2),
            "total_orders": total_orders,
            "average_order_value": round(avg_order_value, 2),
            "highest_order_value": highest_order,
            "lowest_order_value": lowest_order,
        },
        "customers": {
            "total_customers_served": len(customers),
            "new_customers": int(new_customers),
            "returning_customers": int(returning),
        },
        "products": {
            "total_items_sold": int(total_items_sold),
            "unique_items_sold": len(unique_item_ids),
        },
        "operations": {
            "workstation_rows": workstation_rows,
        },
        "growth": {
            "total_sales_pct": _growth(total_sales, prev_total_sales),
            "orders_pct": _growth(total_orders, prev_total_orders),
            "average_order_value_pct": _growth(avg_order_value, prev_avg),
        },
    }

    payload = {
        "filters": filters,
        "period": {
            "label": period_label,
            "start": query_start_dt.isoformat(),
            "end": query_end_dt.isoformat(),
            "previous_start": prev_query_start.isoformat(),
            "previous_end": prev_query_end.isoformat(),
            "granularity": granularity,
            "display_start_ist": start_dt.isoformat(),
            "display_end_ist": end_dt.isoformat(),
        },
        "summary": summary,
        "revenue": {
            "trend": {"labels": revenue_trend_labels, "values": revenue_trend_values},
            "split": {
                "labels": [row["label"] for row in workstation_rows if row["sales"] > 0],
                "values": [row["sales"] for row in workstation_rows if row["sales"] > 0],
                "colors": [row["color"] for row in workstation_rows if row["sales"] > 0],
            },
            "comparison": {
                "labels": [row["label"] for row in workstation_rows],
                "values": [row["sales"] for row in workstation_rows],
                "colors": [row["color"] for row in workstation_rows],
            },
            "contribution_trend": {
                "labels": revenue_trend_labels,
                "datasets": station_contribution_datasets,
            },
        },
        "orders_analytics": {
            "orders_over_time": {"labels": revenue_trend_labels, "values": order_trend_values},
            "order_type_distribution": order_type_counts,
            "peak_hours": peak_hours,
            "best_sales_days": {
                "best_day_ever": best_day_ever,
                "best_day_this_month": best_day_this_month,
                "best_day_last_month": best_day_last_month,
            },
        },
        "item_analytics": {
            "top_items": top_items,
            "bottom_items": bottom_items,
            "zero_sales_items": zero_sales,
            "item_revenue_contribution": {
                "labels": [x["name"] for x in top10] + (["Others"] if others_sum > 0 else []),
                "values": [round(x["revenue"], 2) for x in top10] + ([others_sum] if others_sum > 0 else []),
            },
            "profitability": [
                {
                    "item_id": x["id"],
                    "name": x["name"],
                    "revenue": x["revenue"],
                    "estimated_cost": 0.0,
                    "profit": x["revenue"],
                    "profit_margin_pct": 100.0 if x["revenue"] > 0 else 0.0,
                }
                for x in top_items[:20]
            ],
        },
        "category_analytics": {
            "rows": category_rows_out,
            "pie": {"labels": [x["name"] for x in category_rows_out], "values": [x["revenue"] for x in category_rows_out]},
            "bar": {"labels": [x["name"] for x in category_rows_out], "values": [x["revenue"] for x in category_rows_out]},
            "trend": {"labels": revenue_trend_labels, "values": revenue_trend_values},
        },
        "kitchen_vs_barista": {
            "workstation_rows": workstation_rows,
            "contribution_trend": {
                "labels": revenue_trend_labels,
                "datasets": station_contribution_datasets,
            },
        },
        "inventory_consumption": {
            "estimated_consumption": inventory_consumption_rows[:30],
            "most_consumed": inventory_consumption_rows[:10],
            "trend": {"labels": revenue_trend_labels, "values": [0 for _ in revenue_trend_labels]},
        },
        "orders": list(reversed(orders_payload))[:400],
        "generated_at": _format_ist(datetime.utcnow()),
    }

    _STATS_CACHE[cache_key] = {"ts": now_ts, "payload": payload}
    if len(_STATS_CACHE) > 250:
        for key in list(_STATS_CACHE.keys())[:80]:
            _STATS_CACHE.pop(key, None)
    return payload


@bp.route("/api/statistics/summary")
@roles_required("admin", "manager")
def api_stats_summary():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "period": payload["period"], "filters": payload["filters"], "summary": payload["summary"]})


@bp.route("/api/statistics/revenue")
@roles_required("admin", "manager")
def api_stats_revenue():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "period": payload["period"], "filters": payload["filters"], "revenue": payload["revenue"]})


@bp.route("/api/statistics/orders")
@roles_required("admin", "manager")
def api_stats_orders():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "period": payload["period"], "filters": payload["filters"], "orders_analytics": payload["orders_analytics"]})


@bp.route("/api/statistics/items")
@roles_required("admin", "manager")
def api_stats_items():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    item_id = request.args.get("itemId", type=int)
    data = payload["item_analytics"]
    if item_id:
        start_dt = datetime.fromisoformat(payload["period"]["start"])
        end_dt = datetime.fromisoformat(payload["period"]["end"])
        rows = (
            CafeOrder.query.options(joinedload(CafeOrder.order_items))
            .filter(
                CafeOrder.status == "paid",
                CafeOrder.paid_at.is_not(None),
                CafeOrder.paid_at >= start_dt,
                CafeOrder.paid_at < end_dt,
            )
            .order_by(CafeOrder.paid_at.asc(), CafeOrder.created_at.asc())
            .all()
        )
        daily = {}
        for order in rows:
            order_local_day = _format_ist(order.paid_at or order.created_at, "%Y-%m-%d")
            for oi in order.order_items:
                if oi.menu_item_id != item_id or (oi.approval_status or "pending") == "rejected":
                    continue
                key = order_local_day
                d = daily.setdefault(key, {"qty": 0, "revenue": 0.0})
                d["qty"] += int(oi.quantity or 0)
                d["revenue"] = round(d["revenue"] + float(oi.unit_price or 0) * int(oi.quantity or 0), 2)
        labels = sorted(daily.keys())
        trend = {
            "labels": labels,
            "qty_values": [daily[k]["qty"] for k in labels],
            "revenue_values": [daily[k]["revenue"] for k in labels],
        }
        return jsonify({"ok": True, "filters": payload["filters"], "item_trend": trend, "item_analytics": data})
    return jsonify({"ok": True, "filters": payload["filters"], "item_analytics": data})


@bp.route("/api/statistics/categories")
@roles_required("admin", "manager")
def api_stats_categories():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "filters": payload["filters"], "category_analytics": payload["category_analytics"]})


@bp.route("/api/statistics/kitchen-vs-barista")
@roles_required("admin", "manager")
def api_stats_kitchen_barista():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "filters": payload["filters"], "kitchen_vs_barista": payload["kitchen_vs_barista"]})


@bp.route("/api/statistics/inventory-consumption")
@roles_required("admin", "manager")
def api_stats_inventory_consumption():
    payload = _build_stats_payload(_parse_stats_filters(request.args))
    return jsonify({"ok": True, "filters": payload["filters"], "inventory_consumption": payload["inventory_consumption"]})


@bp.route("/staff", methods=["GET", "POST"])
@roles_required("admin", "manager")
def staff():
    _ensure_role_templates_exist()
    ensure_leave_defaults()
    run_leave_maintenance()
    allowed_sections = {
        "add_new_staff",
        "active_staff",
        "archived_staff",
        "attendance_calendar",
        "attendance_entry",
        "leave_requests",
        "leave_settings",
        "rulebook",
        "docs_review",
        "payroll_summary",
    }
    active_staff_section = (request.args.get("section") or "add_new_staff").strip().lower()
    if active_staff_section not in allowed_sections:
        active_staff_section = "add_new_staff"

    def _staff_redirect(section=None, **kwargs):
        target_section = (section or active_staff_section).strip().lower()
        if target_section not in allowed_sections:
            target_section = "add_new_staff"
        params = {"section": target_section}
        params.update(kwargs)
        return redirect(url_for("cafe.staff", **params))

    if request.method == "POST":
        action = request.form.get("action", "create")

        if action == "save_rulebook":
            if not user_has_any_role(g.current_user, "admin"):
                flash("Only Cafe Admin can publish attendance rule-book updates.", "error")
                return _staff_redirect("rulebook")
            title = request.form.get("rulebook_title", "Staff Attendance Rule Book").strip()
            content_text = request.form.get("rulebook_content", "").strip()
            uploaded = request.files.get("rulebook_file")
            if not title:
                title = "Staff Attendance Rule Book"
            if not content_text and (not uploaded or not uploaded.filename):
                flash("Add rule-book text or upload an updated rule-book file.", "error")
                return _staff_redirect("rulebook")
            file_path = None
            file_name = None
            if uploaded and uploaded.filename:
                extension = os.path.splitext(secure_filename(uploaded.filename))[1].lower()
                if extension not in {".pdf", ".txt", ".md", ".doc", ".docx"}:
                    flash("Rule-book uploads must be PDF, TXT, MD, DOC, or DOCX files.", "error")
                    return _staff_redirect("rulebook")
                version = next_rulebook_version()
                file_path = _save_uploaded_file(uploaded, "rulebooks", f"attendance-rules-v{version}")
                file_name = secure_filename(uploaded.filename)
            else:
                version = next_rulebook_version()
            AttendanceRuleBook.query.filter_by(active=True).update({"active": False})
            db.session.add(
                AttendanceRuleBook(
                    version=version,
                    title=title,
                    content_text=content_text or None,
                    file_path=file_path,
                    file_name=file_name,
                    active=True,
                    published_by_user_id=g.current_user.id,
                )
            )
            db.session.commit()
            flash(f"Attendance rule book v{version} is now published.", "success")
            return _staff_redirect("rulebook")

        if action == "create":
            email = request.form["email"].strip().lower()
            existing = User.query.filter_by(email=email).first()
            if existing:
                flash("A user with this email already exists.", "error")
                return _staff_redirect("add_new_staff")
            roles = _parse_roles_from_form()

            new_user = User(
                full_name=request.form["full_name"].strip(),
                email=email,
                password_hash=generate_password_hash(request.form["password"]),
                role=roles[0],
                active=True,
            )
            new_user.set_assigned_roles(roles)
            _assign_user_type_from_role(new_user)
            db.session.add(new_user)
            db.session.flush()

            profile = _ensure_staff_profile(new_user)
            profile.joining_date = (
                date.fromisoformat(request.form["joining_date"])
                if request.form.get("joining_date")
                else None
            )
            profile.dob = date.fromisoformat(request.form["dob"]) if request.form.get("dob") else None
            profile.marital_status = request.form.get("marital_status", "").strip() or None
            profile.gender = request.form.get("gender", "").strip() or None
            profile.shift_start_time = _parse_time_only(request.form.get("shift_start_time"))
            profile.shift_end_time = _parse_time_only(request.form.get("shift_end_time"))
            profile.salary_type = request.form.get("salary_type", "").strip() or "monthly"
            salary_amount_raw = request.form.get("salary_amount", "").strip()
            profile.salary_amount = float(salary_amount_raw) if salary_amount_raw else None
            profile.probation_end_date = date.fromisoformat(request.form["probation_end_date"]) if request.form.get("probation_end_date") else None
            profile.emergency_contact = request.form.get("emergency_contact", "").strip() or None
            profile.archived = False
            db.session.commit()
            flash("Staff member added.", "success")
            return _staff_redirect("active_staff")

        if action == "update":
            user = User.query.get_or_404(int(request.form["user_id"]))
            profile = _ensure_staff_profile(user)

            user.full_name = request.form["full_name"].strip()
            user.email = request.form["email"].strip().lower()
            duplicate = User.query.filter(User.email == user.email, User.id != user.id).first()
            if duplicate:
                flash("Another user already has this email.", "error")
                return _staff_redirect("active_staff")
            roles = _parse_roles_from_form()
            user.set_assigned_roles(roles)
            _assign_user_type_from_role(user)
            user.active = True if request.form.get("active") else False
            new_password = request.form.get("password", "").strip()
            if new_password:
                user.password_hash = generate_password_hash(new_password)

            profile.joining_date = (
                date.fromisoformat(request.form["joining_date"])
                if request.form.get("joining_date")
                else None
            )
            profile.dob = date.fromisoformat(request.form["dob"]) if request.form.get("dob") else None
            profile.marital_status = request.form.get("marital_status", "").strip() or None
            profile.gender = request.form.get("gender", "").strip() or None
            profile.shift_start_time = _parse_time_only(request.form.get("shift_start_time"))
            profile.shift_end_time = _parse_time_only(request.form.get("shift_end_time"))
            profile.phone = request.form.get("phone", "").strip() or None
            profile.alternate_contact = request.form.get("alternate_contact", "").strip() or None
            profile.emergency_contact = request.form.get("emergency_contact", "").strip() or None
            profile.address = request.form.get("address", "").strip() or None
            profile.salary_type = request.form.get("salary_type", "").strip() or None
            salary_amount_raw = request.form.get("salary_amount", "").strip()
            profile.salary_amount = float(salary_amount_raw) if salary_amount_raw else None
            profile.probation_end_date = date.fromisoformat(request.form["probation_end_date"]) if request.form.get("probation_end_date") else None
            profile.pan_number = request.form.get("pan_number", "").strip() or None
            profile.bank_account_name = request.form.get("bank_account_name", "").strip() or None
            profile.bank_account_number = request.form.get("bank_account_number", "").strip() or None
            profile.bank_ifsc = request.form.get("bank_ifsc", "").strip() or None
            profile.bank_name = request.form.get("bank_name", "").strip() or None
            profile.govt_id_type = request.form.get("govt_id_type", "").strip() or None
            profile.govt_id_number = request.form.get("govt_id_number", "").strip() or None
            govt_doc = request.files.get("govt_id_file")
            if govt_doc and govt_doc.filename:
                profile.govt_id_file_path = _save_uploaded_file(govt_doc, "staff_docs", f"govt-{user.id}")
            photo_file = request.files.get("photo_file")
            if photo_file and photo_file.filename:
                profile.photo_file_path = _save_uploaded_file(photo_file, "staff_photos", f"photo-{user.id}")
            profile.archived = not user.active
            db.session.commit()
            flash("Staff member updated.", "success")
            return _staff_redirect("active_staff")

        if action == "delete":
            user = User.query.get_or_404(int(request.form["user_id"]))
            if _is_protected_admin(user):
                flash("Cafe Admin cannot be deleted.", "error")
                return _staff_redirect("active_staff")
            result = retire_staff_account(user)
            if result == "archived":
                flash("Staff member has historical records, so the login was disabled and the account was archived.", "success")
            else:
                flash("Staff member deleted.", "success")
            return _staff_redirect("active_staff")

        if action == "archive":
            user = User.query.get_or_404(int(request.form["user_id"]))
            if user_has_any_role(user, "admin") and user.email == "admin@brownberries.local":
                flash("Cafe Admin cannot be archived.", "error")
                return _staff_redirect("active_staff")
            profile = _ensure_staff_profile(user)
            profile.archived = True
            user.active = False
            db.session.commit()
            flash("Staff member archived.", "success")
            return _staff_redirect("active_staff")

        if action == "restore":
            user = User.query.get_or_404(int(request.form["user_id"]))
            profile = _ensure_staff_profile(user)
            profile.archived = False
            user.active = True
            db.session.commit()
            flash("Staff member restored.", "success")
            return _staff_redirect("archived_staff")

        if action == "leave_decision":
            leave = StaffLeaveRequest.query.get_or_404(int(request.form["leave_id"]))
            decision = request.form.get("decision", "pending")
            if decision not in ["approved", "rejected", "pending", "requested_changes", "cancelled"]:
                decision = "pending"
            leave.admin_remarks = request.form.get("admin_remarks", "").strip() or None
            success, errors = apply_leave_decision(leave, decision, g.current_user)
            if not success:
                db.session.rollback()
                flash(" ".join(errors), "error")
                return _staff_redirect("leave_requests")
            db.session.add(
                StaffNotification(
                    user_id=leave.user_id,
                    category="leave",
                    message=f"Leave request #{leave.id} is now {leave.status.replace('_', ' ').title()}.",
                )
            )
            db.session.commit()
            flash("Leave request updated and attendance/balance records synchronized.", "success")
            return _staff_redirect("leave_requests")

        if action == "save_leave_settings":
            if not user_has_any_role(g.current_user, "admin"):
                flash("Only Cafe Admin can change leave policy settings.", "error")
                return _staff_redirect("leave_settings")
            policy = leave_policy()
            try:
                policy.monthly_earned_credit = max(0, float(request.form.get("monthly_earned_credit", 1)))
                policy.month_end_earned_credit = max(0, float(request.form.get("month_end_earned_credit", 1)))
                policy.max_continuous_days = max(1, int(request.form.get("max_continuous_days", 7)))
                policy.max_company_leaves = max(1, int(request.form.get("max_company_leaves", 2)))
                policy.urgent_leaves_per_year = max(0, float(request.form.get("urgent_leaves_per_year", 12)))
                policy.max_monthly_urgent_leaves = max(0, float(request.form.get("max_monthly_urgent_leaves", 3)))
                policy.role_cooldown_days = max(0, int(request.form.get("role_cooldown_days", 1)))
            except (TypeError, ValueError):
                db.session.rollback()
                flash("Leave policy values must be valid numbers.", "error")
                return _staff_redirect("leave_settings")
            policy.urgent_requires_probation = bool(request.form.get("urgent_requires_probation"))
            weekly = weekly_off_config()
            weekly.enabled = bool(request.form.get("weekly_off_enabled"))
            try:
                weekly.weekday = min(6, max(0, int(request.form.get("weekly_off_weekday", 6))))
            except (TypeError, ValueError):
                weekly.weekday = 6
            weekly.label = request.form.get("weekly_off_label", "Weekly Off").strip() or "Weekly Off"
            db.session.commit()
            flash("Leave policy and weekly-off settings saved.", "success")
            return _staff_redirect("leave_settings")

        if action == "add_holiday":
            if not user_has_any_role(g.current_user, "admin"):
                flash("Only Cafe Admin can manage company holidays.", "error")
                return _staff_redirect("leave_settings")
            try:
                holiday_date = date.fromisoformat(request.form["holiday_date"])
            except (KeyError, TypeError, ValueError):
                flash("Please enter a valid holiday date.", "error")
                return _staff_redirect("leave_settings")
            if CompanyHoliday.query.filter_by(holiday_date=holiday_date).first():
                flash("A company holiday already exists for that date.", "error")
                return _staff_redirect("leave_settings")
            db.session.add(CompanyHoliday(holiday_date=holiday_date, name=request.form.get("holiday_name", "").strip() or "Company Holiday", active=True))
            db.session.commit()
            flash("Company holiday added.", "success")
            return _staff_redirect("leave_settings")

        if action == "delete_holiday":
            if not user_has_any_role(g.current_user, "admin"):
                flash("Only Cafe Admin can manage company holidays.", "error")
                return _staff_redirect("leave_settings")
            holiday = CompanyHoliday.query.get_or_404(int(request.form["holiday_id"]))
            db.session.delete(holiday)
            db.session.commit()
            flash("Company holiday removed.", "success")
            return _staff_redirect("leave_settings")

        if action == "update_role_leave_rule":
            if not user_has_any_role(g.current_user, "admin"):
                flash("Only Cafe Admin can manage role leave rules.", "error")
                return _staff_redirect("leave_settings")
            role_name = request.form.get("role_name", "").strip().lower()
            if not role_name:
                flash("Role name is required.", "error")
                return _staff_redirect("leave_settings")
            rule = RoleLeaveRule.query.filter_by(role_name=role_name).first()
            if not rule:
                rule = RoleLeaveRule(role_name=role_name)
                db.session.add(rule)
            rule.enabled = bool(request.form.get("enabled"))
            try:
                rule.max_concurrent_on_leave = max(1, int(request.form.get("max_concurrent_on_leave", 1)))
                rule.cooldown_days = max(0, int(request.form.get("cooldown_days", 1)))
            except (TypeError, ValueError):
                db.session.rollback()
                flash("Role rule values must be valid numbers.", "error")
                return _staff_redirect("leave_settings")
            db.session.commit()
            flash("Role leave rule updated.", "success")
            return _staff_redirect("leave_settings")

        if action == "attendance_for_user":
            target_user_id = int(request.form["target_user_id"])
            attendance_date = date.fromisoformat(request.form["attendance_date"])
            status = (request.form.get("status") or "").strip()
            valid_statuses = {value for value, _ in ATTENDANCE_STATUS_OPTIONS if value}
            if status and status not in valid_statuses:
                status = ""
            notes = request.form.get("notes", "").strip() or None
            check_in_time = attendance_datetime_for(attendance_date, request.form.get("check_in_time"))
            check_out_time = attendance_datetime_for(attendance_date, request.form.get("check_out_time"))
            if check_in_time and check_out_time and check_out_time < check_in_time:
                flash("Check-out time cannot be before check-in time.", "error")
                return _staff_redirect("attendance_entry", attendance_user_id=target_user_id)
            existing = StaffAttendance.query.filter_by(
                user_id=target_user_id, attendance_date=attendance_date
            ).first()
            if existing:
                existing.check_in_at = check_in_time
                existing.check_out_at = check_out_time
                existing.notes = notes
                existing.manager_override = True
                refresh_attendance_row(existing, manual_status=status)
            else:
                existing = StaffAttendance(
                    user_id=target_user_id,
                    attendance_date=attendance_date,
                    status=status or "absent",
                    check_in_at=check_in_time,
                    check_out_at=check_out_time,
                    manager_override=True,
                    notes=notes,
                )
                db.session.add(existing)
                refresh_attendance_row(existing, manual_status=status)
            db.session.commit()
            flash("Attendance override saved for selected staff member.", "success")
            return _staff_redirect("attendance_entry", attendance_user_id=target_user_id)

        if action == "checkout_active_session":
            attendance_id = int(request.form["attendance_id"])
            existing = StaffAttendance.query.get_or_404(attendance_id)
            if not existing.check_in_at or existing.check_out_at:
                flash("This session is already closed.", "error")
                return _staff_redirect("attendance_entry", attendance_user_id=existing.user_id)
            checkout_time = attendance_datetime_for(
                existing.attendance_date,
                request.form.get("check_out_time"),
            ) or datetime.now()
            if checkout_time < existing.check_in_at:
                checkout_time = existing.check_in_at
            existing.check_out_at = checkout_time
            existing.check_out_method = "admin_override"
            existing.manager_override = True
            admin_note = f"Checked out by {g.current_user.full_name}"
            existing.notes = (
                f"{existing.notes} | {admin_note}"[:255]
                if existing.notes
                else admin_note
            )
            refresh_attendance_row(existing)
            db.session.commit()
            flash("Active session checked out.", "success")
            return _staff_redirect("attendance_entry", attendance_user_id=existing.user_id)

        if action == "checkout_all_active_sessions":
            rows = StaffAttendance.query.filter(
                StaffAttendance.check_in_at.is_not(None),
                StaffAttendance.check_out_at.is_(None),
            ).all()
            if not rows:
                flash("There are no active sessions to close.", "error")
                return _staff_redirect("attendance_entry")
            raw_time = request.form.get("bulk_check_out_time")
            for existing in rows:
                checkout_time = attendance_datetime_for(existing.attendance_date, raw_time)
                if not checkout_time:
                    checkout_time = datetime.now() if existing.attendance_date == date.today() else datetime.combine(existing.attendance_date, time(18, 0))
                if checkout_time < existing.check_in_at:
                    checkout_time = existing.check_in_at
                existing.check_out_at = checkout_time
                existing.check_out_method = "admin_bulk_checkout"
                existing.manager_override = True
                admin_note = f"Bulk checked out by {g.current_user.full_name}"
                existing.notes = (
                    f"{existing.notes} | {admin_note}"[:255]
                    if existing.notes
                    else admin_note
                )
                refresh_attendance_row(existing)
            db.session.commit()
            flash(f"Closed {len(rows)} active session(s).", "success")
            return _staff_redirect("attendance_entry")

    excluded_staff_emails = ["qr.guest@brownberries.local", "delivery.guest@brownberries.local"]
    staff_users = (
        User.query.filter(~User.email.in_(excluded_staff_emails))
        .order_by(User.full_name.asc())
        .all()
    )
    for staff_user in staff_users:
        _ensure_staff_profile(staff_user)
    db.session.commit()
    staff_profiles = [u.staff_profile for u in staff_users if u.staff_profile]
    active_profiles = [s for s in staff_profiles if not s.archived]
    archived_profiles = [s for s in staff_profiles if s.archived]
    attendance_today = {
        row.user_id: row
        for row in StaffAttendance.query.filter_by(attendance_date=date.today()).all()
    }
    attendance_latest = {}
    latest_rows = (
        StaffAttendance.query.order_by(
            StaffAttendance.user_id.asc(), StaffAttendance.attendance_date.desc()
        ).all()
    )
    for row in latest_rows:
        if row.user_id not in attendance_latest:
            attendance_latest[row.user_id] = row
    leave_summary = {}
    for staff_user in staff_users:
        pending = StaffLeaveRequest.query.filter_by(user_id=staff_user.id, status="pending").count()
        approved = StaffLeaveRequest.query.filter_by(user_id=staff_user.id, status="approved").count()
        leave_summary[staff_user.id] = {"pending": pending, "approved": approved}
    leave_requests = (
        StaffLeaveRequest.query.join(User, StaffLeaveRequest.user_id == User.id)
        .order_by(StaffLeaveRequest.created_at.desc())
        .limit(80)
        .all()
    )
    active_attendance_sessions = (
        StaffAttendance.query.join(User, StaffAttendance.user_id == User.id)
        .options(joinedload(StaffAttendance.user))
        .filter(
            User.active.is_(True),
            StaffAttendance.check_in_at.is_not(None),
            StaffAttendance.check_out_at.is_(None),
        )
        .order_by(StaffAttendance.check_in_at.asc())
        .all()
    )

    selected_user_id = request.args.get("attendance_user_id", type=int)
    selected_month = request.args.get("attendance_month", type=int) or date.today().month
    selected_year = request.args.get("attendance_year", type=int) or date.today().year
    if selected_month < 1 or selected_month > 12:
        selected_month = date.today().month
    if selected_year < 2000 or selected_year > 2100:
        selected_year = date.today().year
    if not selected_user_id and active_profiles:
        selected_user_id = active_profiles[0].user_id

    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    month_rows = calendar.Calendar(firstweekday=0).monthdayscalendar(selected_year, selected_month)
    selected_attendance_map = {}
    if selected_user_id:
        rows = StaffAttendance.query.filter(
            StaffAttendance.user_id == selected_user_id,
            StaffAttendance.attendance_date >= date(selected_year, selected_month, 1),
            StaffAttendance.attendance_date <= date(selected_year, selected_month, days_in_month),
        ).all()
        selected_attendance_map = {row.attendance_date.day: row.status for row in rows}
        selected_start = date(selected_year, selected_month, 1)
        selected_end = date(selected_year, selected_month, days_in_month)
        for holiday in CompanyHoliday.query.filter(
            CompanyHoliday.active.is_(True),
            CompanyHoliday.holiday_date >= selected_start,
            CompanyHoliday.holiday_date <= selected_end,
        ).all():
            selected_attendance_map.setdefault(holiday.holiday_date.day, "company_holiday")
        weekly_config = weekly_off_config()
        if weekly_config.enabled:
            for day_number in range(1, days_in_month + 1):
                day_value = date(selected_year, selected_month, day_number)
                if day_value.weekday() == weekly_config.weekday:
                    selected_attendance_map.setdefault(day_number, "weekly_off")

    payroll_month = request.args.get("payroll_month", type=int) or date.today().month
    payroll_year = request.args.get("payroll_year", type=int) or date.today().year
    if payroll_month < 1 or payroll_month > 12:
        payroll_month = date.today().month
    if payroll_year < 2000 or payroll_year > 2100:
        payroll_year = date.today().year
    payroll_days = calendar.monthrange(payroll_year, payroll_month)[1]
    payroll_start = date(payroll_year, payroll_month, 1)
    payroll_end = date(payroll_year, payroll_month, payroll_days)
    payroll_rows = []
    total_payroll_estimate = 0.0
    today_attendance_rows = StaffAttendance.query.filter_by(attendance_date=date.today()).all()
    doc_pending_count = StaffDocument.query.filter(
        StaffDocument.verification_status.in_(["pending", "rejected"])
    ).count()
    pending_leave_count = StaffLeaveRequest.query.filter_by(status="pending").count()
    pending_documents = (
        StaffDocument.query.join(User, StaffDocument.user_id == User.id)
        .options(joinedload(StaffDocument.user), joinedload(StaffDocument.uploaded_by))
        .filter(StaffDocument.verification_status.in_(["pending", "rejected"]))
        .order_by(StaffDocument.created_at.desc())
        .all()
    )
    leave_policy_row = leave_policy()
    weekly_off_row = weekly_off_config()
    current_rulebook = ensure_rulebook_default()
    rulebook_versions = AttendanceRuleBook.query.order_by(AttendanceRuleBook.version.desc()).limit(20).all()
    company_holidays = CompanyHoliday.query.filter_by(active=True).order_by(CompanyHoliday.holiday_date.asc()).all()
    role_leave_rules = RoleLeaveRule.query.order_by(RoleLeaveRule.role_name.asc()).all()
    staff_leave_balances = {
        row.user_id: row
        for row in LeaveBalance.query.join(User).filter(User.active.is_(True)).all()
    }
    for staff_user in staff_users:
        profile = staff_user.staff_profile
        month_rows = StaffAttendance.query.filter(
            StaffAttendance.user_id == staff_user.id,
            StaffAttendance.attendance_date >= payroll_start,
            StaffAttendance.attendance_date <= payroll_end,
        ).all()
        present_days = sum(1 for row in month_rows if row.status in ["present_all_day", "late_entry", "early_exit", "weekly_off", "on_leave", "earned_leave", "company_holiday"])
        half_days = sum(1 for row in month_rows if row.status in ["first_half", "second_half", "half_day_leave", "half_day_earned_leave", "half_day_urgent_leave"])
        sick_days = sum(1 for row in month_rows if row.status == "sick_leave")
        unpaid_days = sum(max(0.0, 1.0 - attendance_pay_fraction(row.status)) for row in month_rows)
        payable_days = sum(attendance_pay_fraction(row.status) for row in month_rows)
        salary_amount = float(profile.salary_amount or 0)
        per_day_salary = round((salary_amount / payroll_days), 2) if salary_amount else 0.0
        estimated_pay = round(payable_days * per_day_salary, 2)
        total_payroll_estimate += estimated_pay
        payroll_rows.append(
            {
                "user": staff_user,
                "profile": profile,
                "present_days": present_days,
                "half_days": half_days,
                "sick_days": sick_days,
                "unpaid_days": round(unpaid_days, 2),
                "payable_days": round(payable_days, 2),
                "per_day_salary": per_day_salary,
                "estimated_pay": estimated_pay,
            }
        )

    return render_template(
        "cafe/staff_admin.html",
        active_profiles=active_profiles,
        archived_profiles=archived_profiles,
        attendance_today=attendance_today,
        attendance_latest=attendance_latest,
        leave_summary=leave_summary,
        leave_requests=leave_requests,
        attendance_user_id=selected_user_id,
        attendance_month=selected_month,
        attendance_year=selected_year,
        attendance_month_rows=month_rows,
        selected_attendance_map=selected_attendance_map,
        staff_role_options=_get_role_options(),
        user_types=UserType.query.order_by(UserType.name.asc()).all(),
        active_staff_section=active_staff_section,
        attendance_status_options=ATTENDANCE_STATUS_OPTIONS,
        active_attendance_sessions=active_attendance_sessions,
        payroll_rows=payroll_rows,
        payroll_month=payroll_month,
        payroll_year=payroll_year,
        total_payroll_estimate=round(total_payroll_estimate, 2),
        today_attendance_count=len(today_attendance_rows),
        document_pending_count=doc_pending_count,
        pending_documents=pending_documents,
        pending_leave_count=pending_leave_count,
        leave_policy=leave_policy_row,
        weekly_off_config=weekly_off_row,
        company_holidays=company_holidays,
        role_leave_rules=role_leave_rules,
        staff_leave_balances=staff_leave_balances,
        current_rulebook=current_rulebook,
        rulebook_versions=rulebook_versions,
    )


@bp.route("/user-types", methods=["GET", "POST"])
@roles_required("admin")
def user_types():
    _ensure_role_templates_exist()
    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            name = request.form.get("name", "").strip()
            if not name:
                flash("User type name is required.", "error")
                return redirect(url_for("cafe.user_types"))
            if UserType.query.filter(db.func.lower(UserType.name) == name.lower()).first():
                flash("User type already exists.", "error")
                return redirect(url_for("cafe.user_types"))
            ut = UserType(name=name)
            for field in [
                "can_access_cafe", "can_access_library", "can_manage_staff", "can_manage_menu",
                "can_manage_orders", "can_manage_kitchen", "can_manage_inventory", "can_manage_cashier",
                "can_manage_stats", "can_manage_library_members", "can_manage_library_books",
                "can_manage_library_loans", "can_manage_library_payments", "can_manage_library_plans",
                "can_view_staff_profiles", "can_upload_salary", "can_view_delivery_locations",
            ]:
                setattr(ut, field, True if request.form.get(field) else False)
            # Operational requirement: all roles can take orders and view running orders.
            ut.can_manage_orders = True
            ut.can_manage_kitchen = True
            db.session.add(ut)
            db.session.commit()
            flash("User type created.", "success")
            return redirect(url_for("cafe.user_types"))
        if action == "update":
            ut = UserType.query.get_or_404(int(request.form["user_type_id"]))
            ut.name = request.form.get("name", ut.name).strip() or ut.name
            for field in [
                "can_access_cafe", "can_access_library", "can_manage_staff", "can_manage_menu",
                "can_manage_orders", "can_manage_kitchen", "can_manage_inventory", "can_manage_cashier",
                "can_manage_stats", "can_manage_library_members", "can_manage_library_books",
                "can_manage_library_loans", "can_manage_library_payments", "can_manage_library_plans",
                "can_view_staff_profiles", "can_upload_salary", "can_view_delivery_locations",
            ]:
                setattr(ut, field, True if request.form.get(f"{field}_{ut.id}") else False)
            ut.can_manage_orders = True
            ut.can_manage_kitchen = True
            db.session.commit()
            flash("User type updated.", "success")
            return redirect(url_for("cafe.user_types"))
        if action == "delete":
            ut = UserType.query.get_or_404(int(request.form["user_type_id"]))
            in_use = User.query.filter(
                db.or_(User.user_type_id == ut.id, db.func.lower(User.role) == ut.name.lower())
            ).count()
            if in_use > 0:
                flash(
                    f"Cannot delete role '{ut.name}' because it is assigned to {in_use} user(s). Reassign users first.",
                    "error",
                )
                return redirect(url_for("cafe.user_types"))
            db.session.delete(ut)
            db.session.commit()
            flash("Role template deleted.", "success")
            return redirect(url_for("cafe.user_types"))
    return render_template("cafe/user_types.html", user_types=UserType.query.order_by(UserType.name.asc()).all())


@bp.route("/staff/attendance/export")
@roles_required("admin", "manager")
def export_staff_attendance():
    user_id = request.args.get("user_id", type=int)
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    user = User.query.get_or_404(user_id)
    if month is None or month < 1 or month > 12:
        month = date.today().month
    if year is None or year < 2000 or year > 2100:
        year = date.today().year
    days_in_month = calendar.monthrange(year, month)[1]
    from_dt = date(year, month, 1)
    to_dt = date(year, month, days_in_month)
    logs = (
        StaffAttendance.query.filter(
            StaffAttendance.user_id == user.id,
            StaffAttendance.attendance_date >= from_dt,
            StaffAttendance.attendance_date <= to_dt,
        )
        .order_by(StaffAttendance.attendance_date.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Staff Name", "Email", "Date", "Status", "Check In", "Check Out", "Worked Hours", "Flags", "Notes"])
    for row in logs:
        ws.append([
            user.full_name,
            user.email,
            row.attendance_date.isoformat(),
            attendance_status_label(row.status),
            row.check_in_at.strftime("%I:%M:%S %p") if row.check_in_at else "",
            row.check_out_at.strftime("%I:%M:%S %p") if row.check_out_at else "",
            worked_hours_for_row(row),
            ", ".join(attendance_flags_for_row(row)),
            row.notes or "",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = user.full_name.lower().replace(" ", "-")
    filename = f"{safe_name}-attendance-{year}-{month:02d}.xlsx"
    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


@bp.route("/staff/<int:user_id>/id-card")
@roles_required("admin", "manager")
def download_staff_id_card(user_id):
    user = User.query.get_or_404(user_id)
    profile = _ensure_staff_profile(user)
    output = _build_staff_id_card(user, profile)
    filename = f"{user.full_name.lower().replace(' ', '-')}-staff-id.png"
    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "image/png",
        },
    )


@bp.route("/staff/<int:user_id>/govt-id")
@roles_required("admin", "manager")
def download_staff_govt_id(user_id):
    user = User.query.get_or_404(user_id)
    profile = _ensure_staff_profile(user)
    if not profile.govt_id_file_path:
        flash("No Govt ID file uploaded for this staff member.", "error")
        return redirect(url_for("cafe.staff"))
    full_path = os.path.join(current_app.config["UPLOADS_ROOT"], profile.govt_id_file_path)
    if not os.path.exists(full_path):
        flash("Govt ID file is missing from storage.", "error")
        return redirect(url_for("cafe.staff"))
    with open(full_path, "rb") as f:
        data = f.read()
    filename = os.path.basename(full_path)
    return Response(
        data,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/octet-stream",
        },
    )


@bp.route("/my-staff", methods=["GET", "POST"])
@login_required
def my_staff():
    user = g.current_user
    run_leave_maintenance()
    if _is_staff_user(user):
        profile = _ensure_staff_profile(user)
        db.session.commit()
    else:
        profile = user.staff_profile
    if request.method == "POST":
        action = request.form.get("action")
        if action == "check_in":
            flash("Please use the staff attendance QR to check in from the cafe premises.", "error")
            return redirect(url_for("main.staff_attendance_check_in"))

        if action == "check_out":
            existing = (
                StaffAttendance.query.filter(
                    StaffAttendance.user_id == user.id,
                    StaffAttendance.check_in_at.is_not(None),
                    StaffAttendance.check_out_at.is_(None),
                )
                .order_by(StaffAttendance.attendance_date.desc(), StaffAttendance.check_in_at.desc())
                .first()
            )
            if not existing or not existing.check_in_at:
                flash("No active check-in session found.", "error")
                return redirect(url_for("cafe.my_staff"))
            existing.check_out_at = datetime.now()
            existing.check_out_method = "profile"
            existing.manager_override = False
            refresh_attendance_row(existing)
            flash("Check-out recorded.", "success")
            db.session.commit()
            return redirect(url_for("cafe.my_staff"))

        if action == "leave":
            try:
                start_date = date.fromisoformat(request.form["start_date"])
                end_date = date.fromisoformat(request.form["end_date"])
            except (KeyError, TypeError, ValueError):
                flash("Please enter valid leave dates.", "error")
                return redirect(url_for("cafe.my_staff"))
            leave_type = request.form.get("leave_type", "earned").strip().lower()
            from_shift = request.form.get("from_shift", "first_half").strip().lower()
            to_shift = request.form.get("to_shift", "second_half").strip().lower()
            errors, duration = validate_leave_request(user, leave_type, start_date, end_date, from_shift, to_shift)
            if errors:
                db.session.rollback()
                flash(" ".join(errors), "error")
                return redirect(url_for("cafe.my_staff"))
            db.session.add(
                StaffLeaveRequest(
                    user_id=user.id,
                    leave_type=leave_type,
                    start_date=start_date,
                    end_date=end_date,
                    from_shift=from_shift,
                    to_shift=to_shift,
                    duration_days=duration,
                    reason=request.form.get("reason", "").strip() or None,
                    status="pending",
                )
            )
            db.session.commit()
            flash("Leave request submitted.", "success")
            return redirect(url_for("cafe.my_staff"))

        if action == "change_password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not check_password_hash(user.password_hash, current_password):
                flash("Current password is incorrect.", "error")
                return redirect(url_for("cafe.my_staff"))
            if len(new_password) < 6:
                flash("New password must be at least 6 characters.", "error")
                return redirect(url_for("cafe.my_staff"))
            if new_password != confirm_password:
                flash("New password and confirmation do not match.", "error")
                return redirect(url_for("cafe.my_staff"))
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash("Password changed successfully.", "success")
            return redirect(url_for("cafe.my_staff"))

    attendance_logs = (
        StaffAttendance.query.filter_by(user_id=user.id)
        .order_by(StaffAttendance.attendance_date.desc())
        .limit(40)
        .all()
    )
    attendance_changed = False
    for row in attendance_logs:
        before = row.status
        refresh_attendance_row(row, manual_status=before if (row.manager_override and not (row.check_in_at or row.check_out_at)) else None)
        if row.status != before:
            attendance_changed = True
    if attendance_changed:
        db.session.commit()
    leave_logs = (
        StaffLeaveRequest.query.filter_by(user_id=user.id)
        .order_by(StaffLeaveRequest.created_at.desc())
        .limit(40)
        .all()
    )
    month_start = date.today().replace(day=1)
    month_logs = [row for row in attendance_logs if row.attendance_date >= month_start]
    attendance_summary = build_attendance_summary(month_logs)
    today_attendance = next((row for row in attendance_logs if row.attendance_date == date.today()), None)
    active_attendance_session = next((row for row in attendance_logs if row.check_in_at and not row.check_out_at), None)
    next_attendance_action = "completed"
    if active_attendance_session:
        next_attendance_action = "check_out"
    elif not today_attendance or not today_attendance.check_in_at:
        next_attendance_action = "check_in"
    return render_template(
        "cafe/staff_self.html",
        profile=profile,
        attendance_logs=attendance_logs,
        leave_logs=leave_logs,
        attendance_summary=attendance_summary,
        today_attendance=today_attendance,
        active_attendance_session=active_attendance_session,
        next_attendance_action=next_attendance_action,
        attendance_status_label=attendance_status_label,
        attendance_flags_for_row=attendance_flags_for_row,
        worked_hours_for_row=worked_hours_for_row,
        leave_type_options=SELF_LEAVE_TYPE_OPTIONS,
    )
